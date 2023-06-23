import openpyxl
import os
import sys
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.styles.borders import Border, Side
from openpyxl.workbook.defined_name import DefinedName


def create_reports_sday_v10(sl_object_all: dict, node_param_rus: str):
    igrek = 24
    igrek_delta = 24
    sl_param = dict()
    #  os.path.dirname(sys.argv[0])
    if os.path.exists(os.path.join(os.path.abspath(os.curdir), 'SCADA', 'Daily')):
        all_obj = tuple([i[0] for i in sl_object_all])
        for file in os.listdir(os.path.join(os.path.abspath(os.curdir), 'SCADA', 'Daily')):
            if file[:file.find('_')] in all_obj and file.endswith('DailyList.csv'):
                # print(f"---------------------{file[:file.find('_')]}--------------------------")
                with open(os.path.join(os.path.abspath(os.curdir), 'SCADA', 'Daily', file), 'r') as f_:
                    for line in f_:
                        if 'Ошибка' in line or ';' not in line:
                            continue
                        tmp_list = line.strip().split(';')
                        # print(tmp_list)
                        if 'find' in tmp_list:
                            continue
                        folder_tag = tmp_list[1]
                        obj = folder_tag[:folder_tag.find('.')]
                        ind_end_str = folder_tag.rfind('.Value.V') if folder_tag.endswith('.Value.V') else len(
                            folder_tag)
                        tag = folder_tag[folder_tag.find('.') + 1:ind_end_str]
                        description = tmp_list[2]
                        eunit = tmp_list[3]
                        precision = eunit[:eunit.find('|')]
                        eunit = '-' if 'Не используется' in eunit else eunit[eunit.find('|') + 1:]
                        short_name = tmp_list[0]

                        if obj not in sl_param:
                            sl_param[obj] = {}
                        if tag not in sl_param[obj]:
                            sl_param[obj][tag] = (short_name, description, eunit, precision)
    else:
        print('Папка SCADA/Daily не найдена')

    sl_name_object = {alg_name[0]: alg_name[1] for alg_name in sl_object_all}
    # print(sl_param)

    thin_border = Border(left=Side(style='thin'),  # thick - толстые границы, thin - тонкие
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    tuple_sh = ('W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH')
    str_sh = 'IJKLMNOPQRST'

    sl_param = {i: value for i, value in sl_param.items() if i}
    # Для каждого объекта
    for obj, sl_par in sl_param.items():
        num_cel_start = 8
        num_cel = 8
        num_par = 1
        # Создаём эксельку
        wb = openpyxl.Workbook()
        # Выбираем активный лист
        sheet = wb.active
        sheet.title = 'Лист 1'
        # Заполняем стартовую информацию
        for col, w in {'F': 5, 'G': 20, 'H': 10}.items():
            sheet.column_dimensions[col].width = w
        sheet['E2'] = f'="{obj}."'

        for cell, value in {'A2': '=NOW()', 'B2': '="№ ГПА"', 'A3': '=reportdate', 'A4': '=smena_key', 'A5': '="smena=1"',
                            'A6': '="smena=2"', 'B3': '="дата"', 'B4': '="смена"', 'B5': '="дневная"',
                            'B6': '="ночная"', 'C7': '="узел"', 'D7': '="ед.изм"', 'E7': '="значение"',
                            'D3': '="начало пути"', 'E3': '="Получение данных с OPC UA@"',
                            'D4': '="значение"', 'E4': '=".Value;1"', 'E5': '=".Value.100;1"',
                            'A9': '=YEAR(A3)', 'B9': '="год"', 'A10': '=MONTH(A3)', 'B10': '="месяц"',
                            'A11': '=DAY(A3)', 'B11': '="день"', 'A12': '=DATE(A9, A10, A11)',
                            'A14': '=IF(A4=1, TIME(8, 0, 0), TIME(20, 0, 0))', 'A15': '=TIME(0, 3, 0)',
                            'B14': '="выбор смены"', 'B15': '="диапазон"'}.items():
            sheet[cell] = value

        # Заполняем временные формулы
        for cell, value in {'I5': '=A12+A14', 'I6': '=I5+$A$15'}.items():  # 'I6': '=I5+$A$15'
            sheet[cell] = value
        cell_ind_str = 'JKLMNOPQRST'
        formul_ind_str = 'IJKLMNOPQRS'
        for i in range(len(cell_ind_str)):
            sheet[f'{cell_ind_str[i]}5'] = f'={formul_ind_str[i]}5+1/24'
            sheet[f'{cell_ind_str[i]}6'] = f'={cell_ind_str[i]}5+$A$15'
        for coll in 'IJKLMNOPQRST':
            sheet[f'{coll}4'] = f'=IF($A$2>{coll}5, true, false)'
        # Выставляем форматы
        for cell, value in {'A2': 22, 'A3': 22, 'A12': 22, 'A14': 21, 'A15': 21}.items():
            sheet[cell].number_format = numbers.BUILTIN_FORMATS[value]
        for coll in 'IJKLMNOPQRST':
            sheet[f'{coll}5'].number_format = numbers.BUILTIN_FORMATS[22]
            sheet[f'{coll}6'].number_format = numbers.BUILTIN_FORMATS[22]
        # Заполняем шапку формул данных
        sheet['V5'] = '="текущие значения"'
        for i in range(len(tuple_sh)):
            sheet[f'{tuple_sh[i]}5'] = f'={str_sh[i]}7'
            sheet[f'{tuple_sh[i]}5'].number_format = numbers.BUILTIN_FORMATS[22]

        # Создаём заголовок
        sheet.row_dimensions[2].height = 20
        sheet.row_dimensions[1].height = 25
        sheet.merge_cells('G2:L2')
        sheet[f'G2'] = f'="Сменная ведомость {sl_name_object.get(obj,str(777))} на "'
        sheet[f'G2'].font = Font(size=16, name='Arial', bold=True)
        sheet[f'G2'].alignment = Alignment(horizontal="right", vertical="top")
        sheet.merge_cells('M2:N2')
        sheet[f'M2'] = f'=A3'
        sheet[f'M2'].font = Font(size=16, name='Arial', bold=True)
        sheet[f'M2'].alignment = Alignment(horizontal="center", vertical="top")
        sheet[f'M2'].number_format = 'dd.mm.yyyy'

        for cell, i in {'F7': '="№"', 'G7': '="Наименование  "',
                        'H7': '="ед.изм"'}.items():
            sheet[cell] = i
            sheet[cell].font = Font(size=9, name='Arial', bold=True)
            sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
            sheet[cell].border = thin_border

        # Проставляем время каждый час в первой шапке
        for coll in str_sh:
            sheet[f'{coll}7'] = f'=TIME(HOUR({coll}5), 0, 0)'
            sheet[f'{coll}7'].font = Font(size=9, name='Arial', bold=True)
            sheet[f'{coll}7'].alignment = Alignment(horizontal="center", vertical="center")
            sheet[f'{coll}7'].border = thin_border
            sheet[f'{coll}7'].number_format = numbers.BUILTIN_FORMATS[20]
            sheet.column_dimensions[coll].width = 9

        # ...для каждого параметра...
        tuple_page = tuple()
        num_page = 1
        for par, property_par in sl_par.items():
            # Если отсчитали igrek параметров, то добавляем шапку для следующего листа
            if not (num_par - 1) % igrek and num_par != 1:
                num_cel += 2
                tuple_page += (num_page,)
                num_page += 1
                igrek_delta = 24
                for _ in range(3):
                    sheet.row_dimensions[num_cel].height = 15
                    # for cell, i in {'G': '="должность"', 'J': '="ФИО"', 'M': '="подпись"'}.items():
                    #     sheet[f'{cell}{num_cel}'] = i
                    #     sheet[f'{cell}{num_cel}'].font = Font(size=5, name='Arial')
                    #     sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
                    # for coll in 'GHIJKLMN':
                    #     sheet[f'{coll}{num_cel}'].border = Border(top=Side(style='thin'))
                    # num_cel += 1

                num_cel += 1
                sheet.row_dimensions[num_cel].height = 25
                sheet.merge_cells(f'G{num_cel}:L{num_cel}')
                sheet[f'G{num_cel}'] = f'="Сменная ведомость {sl_name_object.get(obj,str(777))} на "'
                sheet[f'G{num_cel}'].font = Font(size=16, name='Arial', bold=True)
                sheet[f'G{num_cel}'].alignment = Alignment(horizontal="right", vertical="top")
                sheet.merge_cells(f'M{num_cel}:N{num_cel}')
                sheet[f'M{num_cel}'] = f'=M2'
                sheet[f'M{num_cel}'].font = Font(size=16, name='Arial', bold=True)
                sheet[f'M{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
                sheet[f'M{num_cel}'].number_format = 'dd.mm.yyyy'
                num_cel += 2
                sheet.row_dimensions[num_cel].height = 20
                for cell, i in {'F': '="№"', 'G': '="Наименование  "',
                                'H': '="ед.изм"'}.items():
                    sheet[f'{cell}{num_cel}'] = i
                    sheet[f'{cell}{num_cel}'].font = Font(size=9, name='Arial', bold=True)
                    sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="center")
                    sheet[f'{cell}{num_cel}'].border = thin_border
                for coll in 'IJKLMNOPQRST':
                    sheet[f'{coll}{num_cel}'] = f'={coll}7'
                    sheet[f'{coll}{num_cel}'].font = Font(size=9, name='Arial', bold=True)
                    sheet[f'{coll}{num_cel}'].alignment = Alignment(horizontal="center", vertical="center")
                    sheet[f'{coll}{num_cel}'].border = thin_border
                    sheet[f'{coll}{num_cel}'].number_format = numbers.BUILTIN_FORMATS[20]
                num_cel += 1
            # Заполняем параметры
            sheet[f'C{num_cel}'] = f'="{par}"'
            sheet[f'F{num_cel}'] = f'="{num_par}"'
            rus_name = property_par[0]
            sheet[f'G{num_cel}'] = f'="{rus_name}  "'  # Русское наименование

            # Заполняем формулы для вывода данных
            sheet[f'V{num_cel}'] = f'=CurrAttrValue(E{num_cel}, 0)'
            for i in range(len(tuple_sh)):
                sheet[f'{tuple_sh[i]}{num_cel}'] = f'=ArchiveAttributeValue($E{num_cel}, ' \
                                                   f'0, {str_sh[i]}$5, {str_sh[i]}$6, 1)'
                # sheet[f'{tuple_sh[i]}{num_cel}'] = f'=ValueOnDate($E{num_cel}, 0, {str_sh[i]}$5)'
            for i in range(len(str_sh)):
                sheet[f'{str_sh[i]}{num_cel}'] = f'=IF({str_sh[i]}$4, IF(ISNUMBER({tuple_sh[i]}{num_cel}), ' \
                                                 f'{tuple_sh[i]}{num_cel}, "нет"), "-")'  # предпоследний "-" это был $V{num_cel}
                sheet[f'{str_sh[i]}{num_cel}'].number_format = numbers.BUILTIN_FORMATS[2]
                sheet[f'{str_sh[i]}{num_cel}'].font = Font(size=9, name='Arial')
                sheet[f'{str_sh[i]}{num_cel}'].border = thin_border
                sheet[f'{str_sh[i]}{num_cel}'].alignment = Alignment(horizontal="center", vertical="center")
            # Заполняем ссылки
            for i, value in {'D': f'=CONCATENATE($E$3, $E$2, $C{num_cel}, $E$5)',
                             'E': f'=CONCATENATE($E$3, $E$2, $C{num_cel}, $E$4)'}.items():
                sheet[f'{i}{num_cel}'] = value
            # Заполняем формулы
            for i, value in {'H': f'=CurrAttrValue(D{num_cel}, 0)'}.items():
                sheet[f'{i}{num_cel}'] = value

            # Устанавливаем для ячеек размер шрифта 14, Arial и выставляем границу
            for i in 'FGH':
                sheet[f'{i}{num_cel}'].font = Font(size=9, name='Arial')
                sheet[f'{i}{num_cel}'].border = thin_border
            # Выводим в центр нумерацию, значения и единицы измерения
            for i, vert_alig in {'F': 'center', 'G': 'left', 'H': 'center', 'S': 'center'}.items():
                sheet[f'{i}{num_cel}'].alignment = Alignment(horizontal=vert_alig, vertical="center",
                                                             wrap_text=True)
            sheet.row_dimensions[num_cel].height = 15
            num_cel += 1
            num_par += 1
            igrek_delta -= 1

        # Прячем столбцы с привязками
        for col in ('A', 'B', 'C', 'D', 'E'):
            sheet.column_dimensions[col].hidden = True
        for row in (4, 5, 6):
            sheet.row_dimensions[row].hidden = True
        for col in ('V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK',
                    'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX'):
            sheet.column_dimensions[col].hidden = True
        # Добавляем подпись в конце
        num_cel += 2
        tuple_page += (num_page,)
        # for _ in range(3):
        #     sheet.row_dimensions[num_cel].height = 15
        #     for cell, i in {'G': '="должность"', 'J': '="ФИО"', 'M': '="подпись"'}.items():
        #         sheet[f'{cell}{num_cel}'] = i
        #         sheet[f'{cell}{num_cel}'].font = Font(size=5, name='Arial')
        #         sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
        #     for coll in 'GHIJKLMN':
        #         sheet[f'{coll}{num_cel}'].border = Border(top=Side(style='thin'))
        #     num_cel += 1
        # Добавляем общие имена (локальные переменные)
        new_range = DefinedName('smena', attr_text='"1"', comment='2')
        wb.defined_names.add(new_range)
        new_range = DefinedName('reportdate', attr_text='"44987"', comment='0')
        wb.defined_names.add(new_range)

        for _ in range(igrek_delta):
            sheet.row_dimensions[num_cel].height = 15
            num_cel += 1
        cc = igrek + num_cel_start
        for _ in tuple_page:
            # print(_, cc)
            # print(sheet[f'G{_}'].value, end=' ')
            # tmp = sheet[f'G{_}'].value.replace('=', '').replace('"', '')
            sheet[f'U{cc}'].value = f'="{_}/{len(tuple_page)}"'
            sheet[f'U{cc}'].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            cc += (igrek + 6)

        # Добавляем колонтитул
        footer_text = ('_' * 144 + ' ' * 2 + 'должность' + ' ' * 130 + 'ФИО' + ' ' * 50 + 'подпись' + ' '*50) * 3
        sheet.oddFooter.center.text = footer_text
        sheet.oddFooter.center.size = 10
        sheet.oddFooter.center.font = "Arial,Bold"

        wb.save(os.path.join('File_for_Import', 'Reports', f'{node_param_rus} {sl_name_object.get(obj,str(777))}.xlsx'))

    return
