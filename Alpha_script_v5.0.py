# import datetime
import os.path

import openpyxl
import logging
import warnings
# import lxml.etree
import sys
from func_for_v3 import *
from create_trends import is_create_trends
# from read_jtr import is_create_trends_jtr
from alpha_index_v3 import create_index, read_mko_cpu_index
from alpha_index_v3u2 import create_index_u2
from Create_mnemo_v3 import create_mnemo_param, create_mnemo_pz, create_mnemo_drv, create_mnemo_drv_general
from Create_mnemo_visual import create_mnemo_visual
from create_reports import create_reports, create_reports_pz, create_reports_sday
from create_reports_sday import create_reports_sday_v10
from create_reports_v80 import create_reports_v80, create_reports_pz_v80, create_reports_sday_v80
from create_reports_v80 import create_reports_sut_v80
from json import load as json_load
# from collections import ChainMap
from pathlib import Path
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

try:
    pref_IP = ()  # кортеж префиксов IP
    sl_object_all = {}  # {(Объект, рус имя объекта, индекс объекта):
    # {контроллер: (ip основной, ip резервный, индекс объекта)} }

    # # path_config = input('Укажите путь до конфигуратора\n')
    # with open('Source_list_config.txt', 'r', encoding='UTF-8') as f:
    #     path_config = f.readline().strip()
    # print(Path.cwd())
    # path_config = os.path.dirname(sys.argv[0])
    path_config = os.path.abspath(os.curdir)
    file_config = ''
    # Ищем файл конфигуратора в текущем каталоге
    for file in os.listdir(path_config):
        if file.endswith('.xlsm') or file.endswith('.xls'):
            file_config = file
            break
    if not file_config:
        print('Файл конфигуратора не найден, выполнение будет прекращено')
        sleep(3)
        sys.exit()

    print(datetime.datetime.now(), '- Начало сборки файлов')

    print(datetime.datetime.now(), '- Открытие файла конфигуратора')
    book = openpyxl.open(os.path.join(path_config, file_config))  # , read_only=True
    # читаем список всех контроллеров
    print(datetime.datetime.now(), '- Файл открыт. Начало сборки')
    sheet = book['Настройки']  # worksheets[1]

    # Читаем префиксы IP адреса ПЛК
    # Также читаем имя проекта для отработки спец-добавок
    cells = sheet['A1': 'B' + str(sheet.max_row)]
    name_prj = ''
    for p in cells:
        if p[0].value == 'Наменование проекта':
            name_prj = p[1].value
        if p[0].value == 'Cетевая часть адреса основной сети (связь с CPU)':
            pref_IP += (p[1].value + '.',)
        if p[0].value == 'Cетевая часть адреса резервной сети (связь с CPU)':
            pref_IP += (p[1].value + '.',)
            break
    # Читаем размер мнемосхем
    size_shirina = 1920
    size_vysota = 900
    for p in cells:
        if p[0].value == 'Ширина мнемосхемы, в пикселях':
            size_shirina = int(p[1].value)
            # print(size_shirina, 'size_shirina')
        if p[0].value == 'Высота мнемосхемы, в пикселях':
            size_vysota = int(p[1].value)
            # print(size_vysota, 'size_vysota')
            break
    buff_size = '50'
    for p in cells:
        if p[0].value == 'Количество хранимых записей МЭК':
            buff_size = p[1].value
    # Читаем архитектуру
    architecture = 'сингл'
    server_name_osn = 'server11'
    server_name_rez = 'server12'
    for p in cells:
        if p[0].value == 'Архитектура':
            lst_architecture = p[1].value.split(';')
            architecture = 'клиент-сервер' if lst_architecture[0] == 'ON' else 'сингл'
            if len(lst_architecture) == 49:
                server_name_osn = lst_architecture[1]
                server_name_rez = lst_architecture[5]
            elif len(lst_architecture) == 47:
                server_name_osn = lst_architecture[1]
                server_name_rez = lst_architecture[4]
            break
    # print(architecture, server_name_osn, server_name_rez)

    # Читаем межконтроллерку (МКО)
    lst_mko = list()

    sl_mko = {}
    for p in cells:
        if p[0].value == 'Обмен между контроллерами':
            lst_mko = p[1].value.split(';')
            break
    if ''.join(lst_mko):
        for mko in lst_mko:
            mko = mko.replace('(^)', '')
            tmp_mko_list = mko.split('->')
            # print(mko, tmp_mko_list)
            i = tmp_mko_list[0]
            obj1 = i[i.find('(')+1:i.find(')')]
            cpu1 = i[:i.find('(')]
            j = tmp_mko_list[1]
            obj2 = j[j.find('(') + 1:j.find(')')]
            cpu2 = j[:j.find('(')]
            # print(obj1, obj2)
            if obj1 == obj2:
                obj_key = obj1
            else:
                obj_key = (obj1, obj2)
            obj_value_one = (cpu1, cpu2)
            if obj_key not in sl_mko:
                sl_mko[obj_key] = (obj_value_one,)
            elif obj_value_one not in sl_mko.get(obj_key, ''):
                sl_mko[obj_key] += (obj_value_one,)
    set_mko_par_cpu = {x for j in [value for _, value in sl_mko.items()] for x in j}
    # print(set_mko_par_cpu)
    # Читаем состав объектов и заполняем sl_object_all
    cells = sheet['A23': 'U23']
    index_on1 = is_f_ind(cells[0], 'ON1')
    cells = sheet['A24': 'U38']
    for p in cells:
        if p[1].value is None:
            break
        else:
            # промежуточный словарь, для загрузки в общий
            # {контроллер: (ip основной, ip резервный)}
            sl_tmp = {}
            for i in range(index_on1, index_on1+6):
                if p[i].value == 'ON':
                    last_dig_ip = p[i - 5].value
                    tmp_dig_ip1 = (last_dig_ip if '(' not in last_dig_ip else last_dig_ip[:last_dig_ip.find('(')])
                    tmp_dig_ip2 = (last_dig_ip if '(' not in last_dig_ip else last_dig_ip[last_dig_ip.find('(') + 1:-1])
                    sl_tmp[p[i - 10].value] = (tmp_dig_ip1, tmp_dig_ip2)
            sl_object_all[p[1].value, p[2].value, p[0].value.replace('Объект', '').strip()] = sl_tmp

    # Создаём словарь вида {алг.имя объекта: рус имя объекта}
    sl_object_rus_name = {i[0]: i[1] for i in sl_object_all}

    sl_project_settings_default = {
        "Module": {
            "Unet_Client": [0, 1]
        },
        "Protocol": {
            "Unet_version": 1
        }
    }
    try:
        if os.path.exists(os.path.join('Template_Alpha', f'Project_settings.json')):
            with open(os.path.join('Template_Alpha', 'Project_settings.json'), 'r',
                      encoding='UTF-8') as f_sig:
                text_json = json_load(f_sig)
            sl_project_settings = text_json
        else:
            print('Файл Systemach/Project_settings.json не найден, '
                  'Настройки будут использованы по умолчанию')
            sl_project_settings = sl_project_settings_default
        u = sl_project_settings.get("Module", {}).get("Unet_Client", [0, 1])
        unet_start = int(u[0])
        n_unet = int(u[1])
        unet_version = int(sl_project_settings.get("Protocol", {}).get("Unet_version", 1))
    except (Exception, KeyError):
        print('Файл Systemach/Project_settings.json заполнен некорректно, '
              'Настройки будут использованы по умолчанию')
        sl_project_settings = sl_project_settings_default
        u = sl_project_settings.get("Module", {}).get("Unet_Client", [0, 1])
        unet_start = u[0]
        n_unet = u[1]
        unet_version = sl_project_settings.get("Protocol", {}).get("Unet_version", 1)

    # all_cpu - кортеж со всеми контроллерами с цифрой объекта
    all_cpu = tuple([f"{cpu}!{obj[2]}".rstrip('') for obj in sl_object_all for cpu in sl_object_all[obj]])
    # all_cpu_select - словарь кортежей контроллеров с цифрой объекта с разбивкой
    # по указанному количеству элементов в каждом кортеже, ключи - порядковый номер UNET-модуля, учитывается стартовый
    all_cpu_select = [all_cpu[i:i + n_unet] for i in range(0, len(all_cpu), n_unet)]
    all_cpu_select = dict(zip(range(unet_start, len(all_cpu_select) + unet_start), all_cpu_select))

    # Мониторинг ТР и АПР в контроллере
    sl_TR = {}
    if os.path.exists(os.path.join('Template_Alpha', 'TR', 'Tun_TR.txt')):
        with open(os.path.join('Template_Alpha', 'TR', 'Tun_TR.txt'), 'r', encoding='UTF-8') as f_tr:
            for line in f_tr:
                if '#' in line or ':' not in line:
                    continue
                line = line.split(':')
                sl_TR[line[0]] = (tuple([branch.strip() for branch in line[1].split(',')]) if line[1] else tuple())

    choice_tr = ''
    # О: sl_CPU_spec - Словарь спец. добавок, ключ - cpu, значение - кортеж ('ТР', 'АПР') при наличии таковых в cpu
    sl_CPU_spec = {}
    cells = sheet['B1':'R1']
    index_flr_on = is_f_ind(cells[0], 'FLR')
    index_type_tr = is_f_ind(cells[0], 'Тип ТР')
    index_apr_on = is_f_ind(cells[0], 'APR')
    index_sar_on = is_f_ind(cells[0], 'SAR')
    index_dks_on = is_f_ind(cells[0], 'DKS')
    index_path = is_f_ind(cells[0], 'Path')
    index_name_cpu = is_f_ind(cells[0], 'Наименование CPU')
    index_ppu = is_f_ind(cells[0], 'PPU')
    index_alr = is_f_ind(cells[0], 'ALR')
    index_mde = is_f_ind(cells[0], 'MDE')
    index_main = is_f_ind(cells[0], 'Main')
    cells = sheet['B2':'R21']
    # Словарь {ПЛК: путь к проекту ПЛК}
    sl_cpu_path = {}
    for p in cells:
        if p[0].value is None:
            break
        else:
            sl_CPU_spec[p[0].value] = ()
            if p[index_flr_on].value == 'ON':
                sl_CPU_spec[p[0].value] += ('ТР',)
                choice_tr = p[index_type_tr].value
                if choice_tr not in sl_TR and sl_TR:
                    choice_tr = ''
                    print('В файле Tun_TR.txt не указан выбранный тип топливного регулятора')
            if p[index_apr_on].value == 'ON':
                sl_CPU_spec[p[0].value] += ('АПР',)
            if p[index_ppu].value == 'ON':
                sl_CPU_spec[p[0].value] += ('PPU',)
            if p[index_alr].value == 'ON':
                sl_CPU_spec[p[0].value] += ('ALR',)
            if p[index_mde].value == 'ON':
                sl_CPU_spec[p[0].value] += ('MDE',)
            if p[index_main].value == 'ON':
                sl_CPU_spec[p[0].value] += ('Main',)
            if p[index_sar_on].value == 'ON':
                sl_CPU_spec[p[0].value] += ('САР',)
            if p[index_dks_on].value == 'ON':
                sl_CPU_spec[p[0].value] += ('ДКС',)
            # Узнаём пути к контроллерам и забиваем в словарь
            sl_cpu_path[p[index_name_cpu].value] = p[index_path].value
            # os.path.dirname(sys.argv[0])

    # МКО из конфигуратора, потом проверить
    # for _ in sl_mko:
    #     print(_)
    #     print(sl_mko[_])

    # sl_mko = {кортеж объектов: (кортеж кортежей контроллеров с обменом друг с другом)}
    # sl_cpu_mko_index =
    # {cpu: ((полное имя переменной диагностики МКО, индекс, с кем обмен(кортеж, в случае указания объекта), )}}
    sl_cpu_mko_index = read_mko_cpu_index(
        tuple_all_cpu=tuple([cpu for obj in sl_object_all for cpu in sl_object_all[obj]]),
        sl_cpu_path=sl_cpu_path,
    )
    # print()
    # for c in sl_cpu_mko_index:
    #     for par in sl_cpu_mko_index[c]:
    #         print(c, *par)
    #     print()
    # sl_obj_cpu_mko_index = {алг.имя объекта:
    # {cpu: ((полное имя переменной диагностики МКО, индекс, с кем обмен(кортеж, в случае указания объекта), )}}
    sl_obj_cpu_mko_index = {}
    for obj_, obj_cpu in {obj[0]: tuple(sl_cpu.keys()) for obj, sl_cpu in sl_object_all.items()}.items():
        if obj_ not in sl_obj_cpu_mko_index:
            sl_obj_cpu_mko_index[obj_] = {}
            for cpu in sl_cpu_mko_index:
                if cpu in obj_cpu:
                    sl_obj_cpu_mko_index[obj_].update({cpu: sl_cpu_mko_index[cpu]})

    # print('sl_object_all', sl_object_all)
    sl_obj_all_cpu = {obj[0]: tuple(sl_cpu.keys()) for obj, sl_cpu in sl_object_all.items()}
    # print('sl_obj_all_cpu', sl_obj_all_cpu)

    # for obj_ in sl_obj_cpu_mko_index:
    #     for cpu in sl_obj_cpu_mko_index[obj_]:
    #         for par in sl_obj_cpu_mko_index[obj_][cpu]:
    #             print(obj_, cpu, *par)
    #         print()
    #     print()

    if os.path.exists(os.path.join(os.path.abspath(os.curdir), 'Template_Alpha', 'Systemach', 'cpu_spec.txt')):
        with open(os.path.join(os.path.abspath(os.curdir), 'Template_Alpha', 'Systemach', 'cpu_spec.txt'),
                  'r', encoding='UTF-8') as f_signal:
            for line in f_signal:
                if '#' in line:
                    continue
                if not line.strip():
                    break
                cpu_s = line.split(':')[0].strip()
                spec = line.split(':')[1].strip()
                if cpu_s in sl_CPU_spec:
                    if spec not in sl_CPU_spec[cpu_s]:
                        sl_CPU_spec[cpu_s] += (spec,)
                else:
                    sl_CPU_spec[cpu_s] = (spec,)
    # print(sl_CPU_spec)
    # Определение заведённых драйверов
    # print('Определение заведённых драйверов, Также Определяем список объявленных мнемосхем')
    # Также Определяем список объявленных мнемосхем
    cells = sheet['A1': 'A' + str(sheet.max_row)]
    drv_eng, drv_rus = [], []
    tuple_mnemo = tuple()
    for p in cells:
        if p[0].value == 'Наименование драйвера (Eng)':
            jj = 1
            while sheet[p[0].row][jj].value and sheet[p[0].row + 1][jj].value:
                drv_eng.append(sheet[p[0].row][jj].value)
                drv_rus.append(sheet[p[0].row + 1][jj].value)
                jj += 1
        if p[0].value == 'Мнемосхемы проекта':
            jj = 1
            while sheet[p[0].row + 1][jj].value:
                tuple_mnemo += (sheet[p[0].row + 1][jj].value,)
                jj += 1
    sl_all_drv = dict(zip(drv_eng, drv_rus))
    # sl_all_drv = {Англ. имя драйвера: русское имя}
    # Если не объявлен IEC, то добавляем
    if 'IEC' not in sl_all_drv:
        sl_all_drv['IEC'] = 'Прочие архивируемые параметры'

    # Если нет папки File_for_Import, то создадим её
    if not os.path.exists('File_for_Import'):
        os.mkdir('File_for_Import')
    # Если нет папки File_for_Import/PLC_Aspect_importDomain, то создадим её
    if not os.path.exists(os.path.join('File_for_Import', 'PLC_Aspect_importDomain')):
        os.mkdir(os.path.join('File_for_Import', 'PLC_Aspect_importDomain'))
    # Если нет папки File_for_Import/IOS_Aspect_in_ApplicationServer, то создадим её
    if not os.path.exists(os.path.join('File_for_Import', 'IOS_Aspect_in_ApplicationServer')):
        os.mkdir(os.path.join('File_for_Import', 'IOS_Aspect_in_ApplicationServer'))
    # Если нет папки File_for_Import/Trends, то создадим её
    if not os.path.exists(os.path.join('File_for_Import', 'Trends')):
        os.mkdir(os.path.join('File_for_Import', 'Trends'))
    # Если нет папки File_for_Import/Mnemo, то создадим её
    if not os.path.exists(os.path.join('File_for_Import', 'Mnemo')):
        os.mkdir(os.path.join('File_for_Import', 'Mnemo'))
    # Если нет папки File_for_Import/Mnemo/Control_Mnemo, то создадим её
    if not os.path.exists(os.path.join('File_for_Import', 'Mnemo', 'Control_Mnemo')):
        os.mkdir(os.path.join('File_for_Import', 'Mnemo', 'Control_Mnemo'))
    # Если нет папки File_for_Import/Mnemo/Control_Mnemo/Systemach, то создадим её
    if not os.path.exists(os.path.join('File_for_Import', 'Mnemo', 'Control_Mnemo', 'Systemach')):
        os.mkdir(os.path.join('File_for_Import', 'Mnemo', 'Control_Mnemo', 'Systemach'))
    # Если нет папки File_for_Import/Reports, то создадим её
    if not os.path.exists(os.path.join('File_for_Import', 'Reports')):
        os.mkdir(os.path.join('File_for_Import', 'Reports'))

    # Чистим папку Мнемосхем, чтобы далее создать новые!!!
    # os.path.dirname(sys.argv[0])
    for file in os.listdir(os.path.join(os.path.abspath(os.curdir), 'File_for_Import', 'Mnemo')):
        if file.endswith('.omobj'):
            os.remove(os.path.join(os.path.abspath(os.curdir), 'File_for_Import', 'Mnemo', file))
    # for file in os.listdir(os.path.join(os.path.abspath(os.curdir), 'File_for_Import', 'Mnemo', 'Control_Mnemo')):
    #     if file.endswith('.omobj'):
    #         os.remove(os.path.join(os.path.abspath(os.curdir), 'File_for_Import', 'Mnemo', 'Control_Mnemo', file))
    # Очистка файла изменений, если он есть и содержит больше 100 строк
    if os.path.exists('Required_change.txt'):
        with open('Required_change.txt', 'r') as f_test:
            check_test = f_test.readlines()
        if len(check_test) > 100:
            with open('Required_change.txt', 'w') as f_test:
                last_change = [i for i in ''.join(check_test).split(f"{'-'* 70}\n") if i][-1]
                f_test.write(f"{'-'* 70}\n{last_change.strip()}\n")

    sl_agreg = {'Agregator_Important_IOS': 'Types.MSG_Agregator.Agregator_Important_IOS',
                'Agregator_LessImportant_IOS': 'Types.MSG_Agregator.Agregator_LessImportant_IOS',
                'Agregator_N_IOS': 'Types.MSG_Agregator.Agregator_N_IOS',
                'Agregator_Repair_IOS': 'Types.MSG_Agregator.Agregator_Repair_IOS'}

    # print(datetime.datetime.now(), ' - Начинаем парсить файл')
    # Измеряемые
    # return_sl = {cpu: {алг_пар: (тип параметра в студии, русское имя, ед. изм., короткое имя, количество знаков,
    # количество знаков для истории)}}
    # sl_mnemo = {узел: список параметров узла}
    return_sl_ai, sl_mnemo_ai, return_sl_sday, return_sl_ai_diff, sl_cpu_fast_ai = {}, {}, {}, {}, {}
    return_sl_ai_add_pars = {}
    if 'Измеряемые' in book.sheetnames:
        return_sl_ai, sl_mnemo_ai, return_sl_sday, return_sl_ai_diff, sl_cpu_fast_ai, return_sl_ai_add_pars = \
            is_read_ai_ae_set(sheet=book['Измеряемые'], type_signal='AI')
    # print(datetime.datetime.now(), ' - Разпарсили Анпары, парсим Расчётные')
    # Расчетные
    return_sl_ae, sl_mnemo_ae, return_ae_sday, return_sl_ae_diff, sl_cpu_fast_ae = {}, {}, {}, {}, {}
    return_sl_ae_add_pars = {}
    if 'Расчетные' in book.sheetnames:
        return_sl_ae, sl_mnemo_ae, return_ae_sday, return_sl_ae_diff, sl_cpu_fast_ae, return_sl_ae_add_pars = \
            is_read_ai_ae_set(sheet=book['Расчетные'], type_signal='AE')
    # print(datetime.datetime.now(), ' - Разпарсили Расчётные, парсим Дискретные')
    # Дискретные
    return_sl_di, sl_wrn_di, sl_mnemo_di, return_sl_di_diff = {}, {}, {}, {}
    if 'Входные' in book.sheetnames:
        return_sl_di, sl_wrn_di, sl_mnemo_di, return_sl_di_diff = is_read_di(sheet=book['Входные'])
    # print(datetime.datetime.now(), ' - Разпарсили Дискретные, парсим ИМы и ИМАО')
    # ИМ
    return_sl_im, sl_cnt, return_sl_im_diff = {}, {}, {}
    return_sl_all_add_pars = {}
    if 'ИМ' in book.sheetnames and 'ИМ(АО)' in book.sheetnames:
        return_sl_im, sl_cnt, return_sl_im_diff, return_sl_all_add_pars = is_read_im(sheet=book['ИМ'],
                                                                                     sheet_imao=book['ИМ(АО)'])

    # Создаём и набиваем словарь для мнемосхемы наработок
    sl_mnemo_cnt = {plc: {'Наработка': list(), 'Перестановки': list()} for plc in sl_cnt}
    for cpu, sl_c in sl_cnt.items():
        for nar in sl_c:
            if 'WorkTime' in nar:
                sl_mnemo_cnt[cpu]['Наработка'].append(nar)
            elif 'Swap' in nar:
                sl_mnemo_cnt[cpu]['Перестановки'].append(nar)

    # # sl_cnt = {CPU: {алг.имя : русское имя}}
    # sl_cnt_xml = {CPU: {алг.имя : (тип модуля в студии, русское имя,)}}
    sl_cnt_xml = {cpu: {alg_par: ('CNT.CNT_PLC_View', val) for alg_par, val in value.items()}
                  for cpu, value in sl_cnt.items()}
    # print(datetime.datetime.now(), ' - Разпарсили ИМы и ИМАО, собираем диагностику модулей')
    # Диагностика
    # sl_modules_cpu {имя CPU: {имя модуля: (тип модуля в студии, тип модуля, [каналы])}}
    # sl_cpu_res = {имя CPU: (агл.имя основного, алг.имя резервного)}
    sl_modules_cpu, sl_for_diag, sl_cpu_res, sl_cpu_rus_name = {}, {}, {}, {}
    if {'Измеряемые', 'Входные', 'Выходные', 'ИМ(АО)'} <= set(book.sheetnames):
        sl_modules_cpu, sl_for_diag, sl_cpu_res, sl_cpu_rus_name = is_read_create_diag(book, name_prj, 'Измеряемые',
                                                                                       'Входные', 'Выходные', 'ИМ(АО)')
    # print('sl_cpu_res ', sl_cpu_res)
    # print('sl_for_diag ', sl_for_diag)
    # print('sl_modules_cpu', sl_modules_cpu)
    # print('sl_cpu_rus_name', sl_cpu_rus_name)
    # print(datetime.datetime.now(), ' - Собрали диагностику, парсим уставки')
    # Уставки
    return_sl_set, sl_set_mnemo_not_use, return_set_sday, return_sl_set_diff, sl_cpu_fast_set = {}, {}, {}, {}, {}
    return_sl_set_add_pars = {}
    if 'Уставки' in book.sheetnames:
        return_sl_set, sl_set_mnemo_not_use, return_set_sday, return_sl_set_diff, sl_cpu_fast_set, return_sl_set_add_pars = \
            is_read_ai_ae_set(sheet=book['Уставки'], type_signal='SET')
    # Собираем словари сменной ведомости в один
    # return_sl_sday = {cpu: {Префикс папки.alg_par: русское имя}}
    for sl in (return_ae_sday, return_set_sday):
        for cpu in sl:
            if cpu in return_sl_sday:
                return_sl_sday[cpu].update(sl[cpu])
            else:
                return_sl_sday[cpu] = sl[cpu]
    # print(datetime.datetime.now(), ' - Разпарсили уставки, парсим кнопки')
    # Кнопки
    return_sl_btn, return_sl_btn_diff = {}, {}
    if 'Кнопки' in book.sheetnames:
        return_sl_btn, return_sl_btn_diff = is_read_btn(sheet=book['Кнопки'])
    # print(datetime.datetime.now(), ' - Разпарсили кнопки, парсим защиты и сигналы')
    # Защиты
    # В чтении защит передаём словари AI и AE с единицами измерения,
    # чтобы определить единицы измерения у защиты при необходимости
    sl_pz, sl_pz_xml = {}, {}
    if 'Сигналы' in book.sheetnames:
        sl_pz, sl_pz_xml = is_read_pz(sheet=book['Сигналы'],
                                      sl_ai={cpu: {par: value[2] for par, value in sl_par.items()}
                                             for cpu, sl_par in return_sl_ai.items()},
                                      sl_ae={cpu: {par: value[2] for par, value in sl_par.items()}
                                             for cpu, sl_par in return_sl_ae.items()})
    # sl_pz - словарь, в котором ключ - cpu, значение - кортеж алг. имён A+000 и т.д. словарь для индексов
    # sl_pz_xml - {cpu: {алг_имя(A000): (рус. имя, ед измерения, Проверяется при ПЗ - Да/Нет)}}

    # return_sl_pru = {cpu: {алг_пар: (Тип сигнала, русское имя,)}}
    return_sl_pru = {}
    if 'ПРУ' in book.sheetnames:
        return_sl_pru = is_read_pru(sheet=book['ПРУ'])

    # print('return_sl_pru ', return_sl_pru)

    # Сигналы остальные
    return_ts, return_ppu, return_alr, return_alg, return_wrn, return_modes = {}, {}, {}, {}, {}, {}
    sl_cpu_archive = {}
    return_sl_diff_sig = {}
    if 'Сигналы' in book.sheetnames:
        return_ts, return_ppu, return_alr, return_alg, return_wrn, return_modes, sl_cpu_archive, return_sl_diff_sig = \
            is_read_signals(
                sheet=book['Сигналы'],
                sl_wrn_di=sl_wrn_di,
                sl_cpu_spec=sl_CPU_spec
            )
    # print(return_sl_diff_sig)
    # print('return_wrn', return_wrn)
    # О
    # print(sl_cpu_archive)
    # print(datetime.datetime.now(), ' - Разпарсили защиты и сигналы, парсим Драйвера')
    # Драйвера
    # return_sl_cpu_drv = {cpu: {(Драйвер, рус имя драйвера):
    # {алг.пар: (Тип переменной в студии, рус имя, тип сообщения, цвет отключения, цвет включения,
    # ед. изм., кол-во знаков, Сохраняемый - Да/Нет, кол-во знаков для истории) }}}
    # sl_cpu_drv_signal = {cpu: {Драйвер: (кортеж переменных)}}
    # return_ios_drv = {(Драйвер, рус имя драйвера): {cpu:
    # {алг.пар: (Тип переменной в студии, рус имя, тип сообщения, цвет отключения, цвет включения,
    # ед. изм., кол-во знаков) }}}
    sl_cpu_drv_signal, return_sl_cpu_drv, return_ios_drv = {}, {}, {}
    if 'Драйвера' in book.sheetnames:
        sl_cpu_drv_signal, return_sl_cpu_drv, return_ios_drv = is_read_drv(
            sheet=book['Драйвера'],
            sl_all_drv=sl_all_drv
        )

    # print(return_sl_all_add_pars)
    # print(return_sl_ai_add_pars)
    # print(return_sl_ai)
    # print('sl_cpu_drv_signal', sl_cpu_drv_signal)
    # print('return_sl_cpu_drv', return_sl_cpu_drv)
    # print('return_ios_drv', return_ios_drv)

    # Переменные алгоритмов
    # словарь команд по cpu sl_command_in_cpu = {cpu: sl_command}
    # sl_command = {(Режим_alg, русское имя режима): {номер шага: {команда_alg: русский текст команды}}}
    sl_grh, return_alg_grh, sl_command_in_cpu = {}, {}, {}
    sl_mod_cpu = {}
    # print(datetime.datetime.now(), ' - Разпарсили Драйвера, парсим Алгоритмы')
    if 'Алгоритмы' in book.sheetnames:
        # Проверяем существование режимов, на тот случай, если лист Алгоритмы может быть пустой
        if 'Режим' in (p[0].value for p in book['Алгоритмы']['A1': 'A' + str(sheet.max_row)]):
            sl_grh, return_alg_grh, sl_command_in_cpu, sl_mod_cpu = \
                is_read_create_grh(sheet=book['Алгоритмы'], sl_object_all=sl_object_all)

    # Подготавливаем словарь такого же вида, как sl_command_in_cpu
    # только для условий перехода и без кортежа в режимах - там просто режим_alg
    sl_condition_in_cpu = {cpu: {mod[0]: {step: {} for step in sl_step} for mod, sl_step in sl_.items()}
                           for cpu, sl_ in sl_command_in_cpu.items()}

    # print('sl_command_in_cpu ', sl_command_in_cpu)
    # print('sl_mod_cpu ', sl_mod_cpu)

    # Создаём кортеж всех команд, чтобы анализировать и отсеивать их
    tuple_all_command = tuple([command for cpu in sl_command_in_cpu
                               for mod in sl_command_in_cpu[cpu]
                               for step in sl_command_in_cpu[cpu][mod]
                               for command in sl_command_in_cpu[cpu][mod][step]])

    for m in [mod[0] for cpu in sl_command_in_cpu for mod in sl_command_in_cpu[cpu]]:
        tuple_all_command += (f'{m}_START', f'{m}_END', f'{m}_Point')

    for cpu, sl_alg_grh in return_alg_grh.items():
        # Определяем все режимы в контроллере
        all_mod_cpu = [i[0] for i in sl_mod_cpu.get(cpu, ' ')]
        # Пробегаемся по алгоритмическим переменным, за исключением команд и Tmv_in в составе
        for cond in [co for co in sl_alg_grh if co not in tuple_all_command and 'Tmv_In' not in co]:
            # Определяем, к какому режиму относится
            mode = ''
            for m in all_mod_cpu:
                if cond.startswith(f"{m}_Cmd_In_") or cond.startswith(f"{m}_Tmv") or cond.startswith(f"{m}_Par"):
                    mode = m
                    break
            # Определяем к какому шагу это относится
            if not mode:
                continue
            step = int([i for i in cond.replace(mode, '').split('_') if i.isdigit()][0])
            # print(f'Команда {cond}, Режим {mode}, Шаг {step}')
            # Закидываем в словарь условий перехода
            # print(mode, cond, f'ШАГ: {step}', f'Переменная : {sl_alg_grh[cond][1]}')
            sl_condition_in_cpu[cpu][mode][int(step)].update({cond: sl_alg_grh[cond][1]})

    # # пробегаемся по общему словарю и выдёргиваем оттуда всё, что не касается, старта, стопа, шага и команд
    # # если не понадобиться, можно удалить Tmv_In если они не нужны - исключено
    # for cpu, sl_alg_grh in return_alg_grh.items():
    #     for alg, tuple_property_alg in sl_alg_grh.items():
    #         if "START" not in alg and "END" not in alg and "Point" not in alg and 'Tmv_In' not in alg \
    #                 and alg not in tuple_all_command:
    #             mod = alg[:alg.find('_')]  # узнаём к какому режиму относится alg-шка
    #             step = [i for i in alg.split('_') if i.isdigit()][0]  # узнаём какой шаг алгоритма
    #             # Засовываем в словарь условия с проверкой вхождения
    #             if cpu in sl_condition_in_cpu and mod in sl_condition_in_cpu[cpu] \
    #                     and int(step) in sl_condition_in_cpu[cpu][mod]:
    #                 sl_condition_in_cpu[cpu][mod][int(step)].update({alg: tuple_property_alg[1]})

    print(datetime.datetime.now(), '- Фиксики начинают создавать и проверять выходные файлы')
    is_create_rlock(sl_object_all=sl_object_all, sl_cpu_archive=sl_cpu_archive)
    # Сеть(коммутаторы) - уже новая функция
    return_sl_net = {}
    if 'Сеть' in book.sheetnames:
        return_sl_net = is_create_net(sl_object_all=sl_object_all, sheet_net=book['Сеть'])

    # print(return_sl_ai)
    # print(sl_mnemo_ai)
    return_sl_cdo, return_sl_mnemo_cdo, return_sl_diff_cdo = {}, {}, {}
    if 'Выходные' in book.sheetnames:
        return_sl_cdo, return_sl_mnemo_cdo, return_sl_diff_cdo = is_read_out(sheet=book['Выходные'], type_signal='CDO')
    # print(return_sl_cdo)
    # print(return_sl_mnemo_cdo)
    # print()

    # print(datetime.datetime.now(), '- Фиксики начинают создавать и проверять выходные файлы')
    sl_w = {
        # return_sl_ai =
        # {cpu: {алг_пар: (тип параметра в студии, русское имя, ед. изм., короткое имя, количество знаков)}}
        'AI': {'dict': return_sl_ai,
               'tuple_attr': ('unit.System.Attributes.Description',
                              'Attributes.EUnit', 'Attributes.ShortName', 'Attributes.FracDigits',
                              'Attributes.Module_Canal'),
               'dict_agreg_IOS': sl_agreg,
               'dict_diff': return_sl_ai_diff
               },
        # return_sl_ae =
        # {cpu: {алг_пар: (тип параметра в студии, русское имя, ед. изм., короткое имя, количество знаков)}}
        'AE': {'dict': return_sl_ae,
               'tuple_attr': ('unit.System.Attributes.Description',
                              'Attributes.EUnit', 'Attributes.ShortName', 'Attributes.FracDigits'),
               'dict_agreg_IOS': sl_agreg,
               'dict_diff': return_sl_ae_diff
               },
        # return_sl_di = {cpu: {алг_пар: (тип параметра в студии, русское имя, sColorOff, sColorOn)}}
        'DI': {'dict': return_sl_di,
               'tuple_attr': ('unit.System.Attributes.Description', 'Attributes.sColorOff', 'Attributes.sColorOn',
                              'Attributes.Module_Canal', 'Attributes.Inverse_Canal'),
               'dict_agreg_IOS': sl_agreg,
               'dict_diff': return_sl_di_diff
               },
        # return_sl_im = {cpu: {алг_пар: (тип ИМа в студии, русское имя, StartView, Gender, ед изм.,
        # количество знаков(для д.ИМ=0), количество знаков для истории)}}
        'IM': {'dict': return_sl_im,
               'tuple_attr': ('unit.System.Attributes.Description',
                              'Attributes.StartView', 'Attributes.Gender', 'Attributes.EUnit', 'Attributes.FracDigits'),
               'dict_agreg_IOS': sl_agreg,
               # 'dict_diff': return_sl_im_diff
               'dict_diff': {_: {p: _ for p, _ in sl_val.items() if not (p.endswith(('_Swap', '_WorkTime')))} for _, sl_val in return_sl_im_diff.items()}
               },
        # return_sl_set =
        # {cpu: {алг_пар: (тип параметра в студии, русское имя, ед. изм., короткое имя, количество знаков)}}
        'SET': {'dict': return_sl_set,
                'tuple_attr': ('unit.System.Attributes.Description',
                               'Attributes.EUnit', 'Attributes.ShortName', 'Attributes.FracDigits'),
                'dict_agreg_IOS': sl_agreg,
                'dict_diff': return_sl_set_diff
                },
        # return_sl_btn = {cpu: {алг_пар: (Тип кнопки в студии, русское имя, )}}
        'BTN': {'dict': return_sl_btn,
                'tuple_attr': ('unit.System.Attributes.Description', ),
                'dict_agreg_IOS': {},
                'dict_diff': return_sl_btn_diff
                },
        # return_sl_pru = {cpu: {алг_пар: (Тип сигнала, русское имя, )}}
        'PRU': {'dict': return_sl_pru,
                'tuple_attr': ('unit.System.Attributes.Description',),
                'dict_agreg_IOS': {}
                },
        # sl_pz_xml = {cpu: {алг_имя(A000): (тип защиты в студии, рус.имя, ед измерения)}}
        'PZ': {'dict': sl_pz_xml,
               'tuple_attr': ('unit.System.Attributes.Description', 'Attributes.EUnit'),
               'dict_agreg_IOS': sl_agreg
               },
        # sl_cnt_xml = {CPU: {алг.имя : (тип модуля в студии, русское имя,)}}
        'CNT': {'dict': sl_cnt_xml,
                'tuple_attr': ('unit.System.Attributes.Description', ),
                'dict_agreg_IOS': {},
                'dict_diff': {_: {p: _ for p, _ in sl_val.items() if p.endswith(('_Swap', '_WorkTime'))} for _, sl_val in return_sl_im_diff.items()}
                },
        # return_ts = {cpu: {алг_пар: (ТИП TS в студии, русское имя, )}}
        'TS': {'dict': return_ts,
               'tuple_attr': ('unit.System.Attributes.Description',),
               'dict_agreg_IOS': {}
               },
        # return_ppu = {cpu: {алг_пар: (ТИП PPU в студии, русское имя, )}}
        'PPU': {'dict': return_ppu,
                'tuple_attr': ('unit.System.Attributes.Description',),
                'dict_agreg_IOS': {}
                },
        # return_alr = {cpu: {алг_пар: (ТИП ALR в студии, русское имя, )}}
        'ALR': {'dict': return_alr,
                'tuple_attr': ('unit.System.Attributes.Description',),
                'dict_agreg_IOS': {'Agregator_Important_IOS': 'Types.MSG_Agregator.Agregator_Important_IOS'}
                },
        # return_alg = {cpu: {алг_пар: (ТИП ALG в студии, русское имя, )}}
        'ALG': {'dict': return_alg,
                'tuple_attr': ('unit.System.Attributes.Description',),
                'dict_agreg_IOS': {}
                },
        # return_wrn = {cpu: {алг_пар: (ТИП WRN в студии, русское имя, )}}
        'WRN': {'dict': return_wrn,
                'tuple_attr': ('unit.System.Attributes.Description',),
                'dict_agreg_IOS': {'Agregator_LessImportant_IOS': 'Types.MSG_Agregator.Agregator_LessImportant_IOS'}
                },
        # return_modes = {cpu: {алг_пар: (ТИП переменной режима в студии, русское имя, )}}
        'MODES': {'dict': return_modes,
                  'tuple_attr': ('unit.System.Attributes.Description',),
                  'dict_agreg_IOS': {}
                  },
        # return_alg_grh = {cpu: {алг_пар: (тип переменной в студии, русское имя )}}
        'GRH': {'dict': return_alg_grh,
                'tuple_attr': ('unit.System.Attributes.Description', 'Attributes.FracDigits'),
                'dict_agreg_IOS': {}
                },
        # return_sl_cdo = {cpu: {алг_пар: (ТИП ALG в студии, русское имя, Нестандартный канал Да/Нет, модуль, канал)}}
        'CDO': {'dict': return_sl_cdo,
                'tuple_attr': ('unit.System.Attributes.Description', 'Attributes.State_Out'),
                'dict_agreg_IOS': {}
                },
        # return_sl_cdo = {cpu: {алг_пар: (ТИП ALG в студии, русское имя, Нестандартный канал Да/Нет, модуль, канал)}}
        'Pars': {'dict': is_update_dict(main_dict=return_sl_all_add_pars, sub_dict=return_sl_ai_add_pars),
                 'tuple_attr': ('unit.System.Attributes.Description', 'Attributes.EUnit'),
                 'dict_agreg_IOS': {}
                 },
    }
    # sl_object_all = {}  #
    # {(Объект, рус имя объекта, индекс объекта): {контроллер: (ip основной, ip резервный, индекс объекта)} }
    # print(return_alg_grh)
    # {cpu: {алг_имя тюнинга: (плк_тип, рус. описание(имя))}}
    sl_tun_apr = {}
    sl_add_obj_cpu_mko = {}
    sl_add_cpu_mko = {}
    # Для каждого объекта...
    progress_ios_plc = len(sl_object_all) + sum([len(sl_) for _, sl_ in sl_object_all.items()])
    # print(progress_ios_plc)
    progress_bar = ' ' * 100
    progress_percent = 0
    for objects in sl_object_all:
        # ...создаём корневой узел xml для IOS-аспекта
        root_ios_aspect = ET.Element('omx', xmlns="system", xmlns_dp="automation.deployment",
                                     xmlns_trei="trei", xmlns_ct="automation.control")
        child_object = ET.SubElement(root_ios_aspect, 'ct_object', name=f"{objects[0]}", access_level="public")
        ET.SubElement(child_object, 'attribute', type=f"unit.System.Attributes.Description",
                      value=f"{objects[1]}")
        # ...добавляем агрегаторы
        for agreg, type_agreg in sl_agreg.items():
            ET.SubElement(child_object, 'ct_object', name=f'{agreg}', base_type=f"{type_agreg}",
                          aspect="Types.IOS_Aspect", access_level="public")
        # Для каждого контроллера в объекте
        for cpu in sl_object_all[objects]:
            # ...создаём корневой узел xml для PLC-аспекта
            # добавляем в него адаптеры с указанными IP
            # добавляем unet-сервер и привязку к карте адресов
            # добавляем APP с именем Tree
            root_plc_aspect = ET.Element('omx', xmlns="system", xmlns_dp="automation.deployment",
                                         xmlns_trei="trei", xmlns_ct="automation.control")
            child = ET.SubElement(root_plc_aspect, 'trei_trei', name=f"PLC_{cpu}_{objects[2]}")
            if unet_version == 2:
                child_cpu = ET.SubElement(child, 'trei_master-module', name="CPU",
                                          soft_version=f"UnimodPro{unet_version}")
            else:
                child_cpu = ET.SubElement(child, 'trei_master-module', name="CPU")

            if pref_IP[0] != '000.000.000.':
                ip1 = '.'.join([a.lstrip('0') for a in f'{pref_IP[0]}{sl_object_all[objects][cpu][0]}'.split('.')])
                ET.SubElement(child_cpu, 'trei_ethernet-adapter', name="Eth1", address=ip1)
            if pref_IP[1] != '000.000.000.':
                ip2 = '.'.join([a.lstrip('0') for a in f'{pref_IP[1]}{sl_object_all[objects][cpu][0]}'.split('.')])
                ET.SubElement(child_cpu, 'trei_ethernet-adapter', name="Eth2", address=ip2)

                # os.path.dirname(sys.argv[0])
            if os.path.exists(os.path.join(os.path.abspath(os.curdir), 'Template_Alpha', 'Systemach',
                                           'cpu', 'unet_tcp_port.txt')):
                sl_unet = {}
                with open(os.path.join(os.path.abspath(os.curdir), 'Template_Alpha', 'Systemach',
                                       'cpu', 'unet_tcp_port.txt'), 'r', encoding='UTF-8') as f_:
                    for line in f_:
                        if '#' in line:
                            continue
                        if not line.strip():
                            break
                        cpu_unet = line.split(':')[0].strip()
                        ports = tuple(sorted(set([i.strip() for i in line.split(':')[1].strip().split(',')
                                                  if i.strip().isdigit()])))
                        sl_unet[cpu_unet] = ports
                if cpu in sl_unet:
                    for port in sl_unet[cpu]:
                        ET.SubElement(child_cpu, 'trei_unet-server', name=f"UnetServer_{port}",
                                      address_map=f"PLC_{cpu}_{objects[2]}.CPU.Tree.UnetAddressMap", port=f"{port}")
                else:
                    ET.SubElement(child_cpu, 'trei_unet-server', name="UnetServer",
                                  address_map=f"PLC_{cpu}_{objects[2]}.CPU.Tree.UnetAddressMap", port="6021")
            else:
                ET.SubElement(child_cpu, 'trei_unet-server', name="UnetServer",
                              address_map=f"PLC_{cpu}_{objects[2]}.CPU.Tree.UnetAddressMap", port="6021")
            child_app = ET.SubElement(child_cpu, 'dp_application-object', name="Tree", access_level="public")
            if unet_version == 2:
                ET.SubElement(child_app, 'trei_unet-address-map', name="UnetAddressMap",
                              protocol_version=f"{'UNET2' if unet_version == 2 else 'UNET'}")
            else:
                ET.SubElement(child_app, 'trei_unet-address-map', name="UnetAddressMap")

            if cpu in sl_cpu_res:
                child_cpu_res = ET.SubElement(child, 'trei_redundant-master-module', name="R-CPU", cpu="CPU")
                if pref_IP[0] != '000.000.000.':
                    ip1 = '.'.join([a.lstrip('0') for a in f'{pref_IP[0]}{sl_object_all[objects][cpu][1]}'.split('.')])
                    ET.SubElement(child_cpu_res, 'trei_ethernet-adapter', name="Eth1", address=ip1)
                if pref_IP[1] != '000.000.000.':
                    ip2 = '.'.join([a.lstrip('0') for a in f'{pref_IP[1]}{sl_object_all[objects][cpu][1]}'.split('.')])
                    ET.SubElement(child_cpu_res, 'trei_ethernet-adapter', name="Eth2", address=ip2)

            # Если есть контроллер САР, то добавляем в него структуру САР (PLC-аспект)
            if 'САР' in sl_CPU_spec.get(cpu, {}):
                child_sar = ET.SubElement(child_app, 'ct_object', name=f"SAR", access_level="public")
                child_sar_im = ET.SubElement(child_sar, 'ct_object', name=f"IM", access_level="public")
                child_sar_reload = ET.SubElement(child_sar, 'ct_object', name=f"Reload", access_level="public")

                flag_reload_standart = False
                if os.path.exists(os.path.join('Template_Alpha', 'SAR', 'Reload_checkbox')):
                    with open(os.path.join('Template_Alpha', 'SAR', 'Reload_checkbox'), 'r', encoding='UTF-8') as f_:
                        for line in f_:
                            if '#' in line:
                                continue
                            if not line.strip():
                                break
                            lst_sar_reload = [i.strip() for i in line.strip().split(';')]
                            if len(lst_sar_reload) == 3:
                                reload_sar = ET.SubElement(
                                    child_sar_reload, 'ct_object', name=f"{lst_sar_reload[1]}",
                                    base_type=f"Types.SAR_Switch_Reload.SAR_Switch_Reload_PLC_View",
                                    aspect="Types.PLC_Aspect", access_level="public"
                                )
                                ET.SubElement(reload_sar, 'attribute', type="unit.System.Attributes.Description",
                                              value=f"{lst_sar_reload[2]}")
                            else:
                                flag_reload_standart = True
                                break
                else:
                    flag_reload_standart = True
                if flag_reload_standart:
                    for i in range(1, 10):
                        reload_km = ET.SubElement(child_sar_reload, 'ct_object', name=f"GPA_KM{i}",
                                                  base_type=f"Types.SAR_Switch_Reload.SAR_Switch_Reload_PLC_View",
                                                  aspect="Types.PLC_Aspect", access_level="public")
                        ET.SubElement(reload_km, 'attribute', type="unit.System.Attributes.Description",
                                      value=f"Загрузка ГПА{i} в магистраль")
                        reload_mk = ET.SubElement(child_sar_reload, 'ct_object', name=f"GPA_MK{i}",
                                                  base_type=f"Types.SAR_Switch_Reload.SAR_Switch_Reload_PLC_View",
                                                  aspect="Types.PLC_Aspect", access_level="public")
                        ET.SubElement(reload_mk, 'attribute', type="unit.System.Attributes.Description",
                                      value=f"Разгрузка ГПА{i} на кольцо")

                ET.SubElement(child_sar_im, 'ct_object', name=f"KHR",
                              base_type=f"Types.SAR_Classic.KHR_PLC_View",
                              aspect="Types.PLC_Aspect", access_level="public")
                ET.SubElement(child_sar, 'ct_object', name=f"REGUL",
                              base_type=f"Types.SAR_Classic.REGUL_PLC_View",
                              aspect="Types.PLC_Aspect", access_level="public")
                ET.SubElement(child_sar, 'ct_object', name=f"WRN",
                              base_type=f"Types.SAR_Classic.WRN_PLC_View",
                              aspect="Types.PLC_Aspect", access_level="public")

                # os.path.dirname(sys.argv[0])
                if os.path.exists(os.path.join(os.path.abspath(os.curdir), 'Template_Alpha', 'SAR', 'Tun_SAR.txt')):
                    child_tun = ET.SubElement(child_sar, 'ct_object', name=f"Tuning", access_level="public")
                    with open(os.path.join(os.path.abspath(os.curdir), 'Template_Alpha', 'SAR', 'Tun_SAR.txt'),
                              'r', encoding='UTF-8') as f_signal:
                        for line in f_signal:
                            if '#' in line:
                                continue
                            if not line.strip():
                                break
                            lst_tun = line.strip().split(';')
                            if lst_tun[0] == 'sSetWord':
                                types_tun = f'SAR_tuning_sSetWord'
                            else:
                                types_tun = f'SAR_tuning' if lst_tun[1] == 'REAL' else f'SAR_tuning_int'
                            object_sar_tuning = ET.SubElement(child_tun, 'ct_object', name=f"{lst_tun[0]}",
                                                              base_type=f"Types.{types_tun}.{types_tun}_PLC_View",
                                                              aspect="Types.PLC_Aspect", access_level="public")
                            ET.SubElement(object_sar_tuning, 'attribute', type="unit.System.Attributes.Description",
                                          value=f"{lst_tun[3]}")
                else:
                    print('Не найден файл тюнингов Template_Alpha/SAR/Tun_SAR.txt, тюнинги не будут добавлены')

            # return_sl = {cpu: {алг_пар: (русское имя, ед изм., короткое имя, количество знаков)}}
            # sl_attr_par - словарь атрибутов параметра {алг_пар: {тип атрибута: значение атрибута}}
            # Добавляем в кортежи параметров на первое место тип сигнала в студии

            # Пробегаемся по словарю ключами анпаров, расчётными и дискретными
            for node in ('AI', 'AE', 'DI', 'IM'):
                # если у текущего контроллера есть анпары или...
                if cpu in sl_w[node]['dict']:
                    tuple_attr = sl_w[node]['tuple_attr']
                    sl_w_node_dict_cpu_copy = copy(sl_w[node]['dict'][cpu])
                    if sl_w[node].get('dict_diff'):
                        # print(objects[1], sl_w[node]['dict_diff'][cpu])
                        for alg_par, value_diff in sl_w[node]['dict_diff'][cpu].items():
                            # print(node, alg_par, value_diff, sl_w_node_dict_cpu_copy.get(alg_par), f'{objects[1]}(нет)' in value_diff)
                            if f'{objects[1]}(нет)' in value_diff:
                                # print(objects[1], alg_par)
                                sl_w_node_dict_cpu_copy[alg_par] = \
                                    sl_w_node_dict_cpu_copy[alg_par][:1] + (sl_w_node_dict_cpu_copy[alg_par][1] +
                                                                            '(РЕЗЕРВ)',) + \
                                    sl_w_node_dict_cpu_copy[alg_par][2:]
                    add_xml_par_plc(name_group=node,
                                    sl_par=sl_w[node]['dict'][cpu],
                                    parent_node=child_app,
                                    sl_attr_par={alg_par: dict(zip(tuple_attr, value[1:]))
                                                 for alg_par, value in sl_w_node_dict_cpu_copy.items()})

            # Если есть АПР в данном контроллере...
            if 'АПР' in sl_CPU_spec.get(cpu, {}):
                # Добавляем ИМ, просто ссылаемся на структуру студии, если что, можно заменить
                child_apr = ET.SubElement(child_app, 'ct_object', name="APR", access_level="public")
                child_apr_im = ET.SubElement(child_apr, 'ct_object', name="IM",
                                             base_type="Types.APR_IM.APR_IM_PLC_View",
                                             aspect="Types.PLC_Aspect", access_level="public")
                ET.SubElement(child_apr_im, 'attribute', type="unit.System.Attributes.Description", value="АПР")
                child_apr_tuninig = ET.SubElement(child_apr, 'ct_object', name="Tuning", access_level="public")
                # Если есть файл с описанием необходимых тюнингов
                if os.path.exists(os.path.join('Template_Alpha', 'APR', 'Tun_APR.txt')):
                    # Создаём кортеж в словаре cpu-тюнингов
                    sl_tun_apr[cpu] = {}
                    # открываем данный файл
                    with open(os.path.join('Template_Alpha', 'APR', 'Tun_APR.txt'), 'r', encoding='UTF-8') as f_in:
                        # и по файлу бежим
                        for line in f_in:
                            if '#' in line:
                                continue
                            in_f = [i.strip() for i in line.split(';')]
                            # Если структура в файле состоит из двух элементов
                            if len(in_f) == 2:
                                # добавляем в кортеж, чтобы использовать потом
                                sl_tun_apr[cpu].update({in_f[0]: ('APR_tuning.APR_tuning_PLC_View', in_f[1])})
                                # То записываем структуру
                                one_tun = ET.SubElement(child_apr_tuninig, 'ct_object', name=f"{in_f[0]}",
                                                        base_type="Types.APR_tuning.APR_tuning_PLC_View",
                                                        aspect="Types.PLC_Aspect", access_level="public")
                                ET.SubElement(one_tun, 'attribute', type="unit.System.Attributes.Description",
                                              value=f"{in_f[1]}")

            # Создаём узел Diag
            child_diag = ET.SubElement(child_app, 'ct_object', name="Diag", access_level="public")
            # sl_modules_cpu {имя CPU: {имя модуля: (тип модуля в студии, тип модуля, [каналы])}}
            # sl_attr_par - словарь атрибутов параметра {алг_пар: {тип атрибута: значение атрибута}}
            if cpu in sl_modules_cpu:
                tuple_attr = ('unit.System.Attributes.Description', 'Attributes.ShortName')
                name_cpu_main = sl_cpu_res[cpu][0] if cpu in sl_cpu_res else '-'
                name_cpu_res = sl_cpu_res[cpu][1] if cpu in sl_cpu_res else '-'
                sl_tt = {alg_par: dict(zip((tuple_attr + ('Attributes.CPU_Name_Main', 'Attributes.CPU_Name_Res')
                                            if 'CPU' in value
                                            else tuple_attr + tuple([f'Attributes.Channel_{i+1}'
                                                                     for i in range(len(value[2]))])),
                                           ([(f'Мастер-модуля ({", ".join(sl_cpu_res.get(cpu))}) ({value[1]})' if cpu in sl_cpu_res
                                              else f'Мастер-модуль {alg_par} ({value[1]})'
                                             if 'CPU' in value else f'Модуль {alg_par} ({value[1]})'),
                                            alg_par, name_cpu_main, name_cpu_res] if 'CPU' in value
                                            else [f'Модуль {alg_par} ({value[1]})', alg_par] + value[2])))
                         for alg_par, value in sl_modules_cpu[cpu].items()}
                add_xml_par_plc(name_group='HW', sl_par=sl_modules_cpu[cpu],
                                parent_node=child_diag,
                                sl_attr_par=sl_tt)

            # Добавляем диагностику МКО
            # sl_obj_cpu_mko_index = {алг.имя объекта:
            # {cpu: ((полное имя переменной диагностики МКО, индекс, с кем обмен(кортеж, в случае указания объекта), )}}
            if cpu in sl_obj_cpu_mko_index.get(objects[0], 'бла'):
                # print(cpu, sl_obj_cpu_mko_index[objects[0]][cpu])
                # sl_par_add = {f"{cpu}_{signal[0].replace('|', '_')}": ('NoPing.NoPing_PLC_View', signal[1])
                #               for signal in sl_obj_cpu_mko_index[objects[0]][cpu]}
                sl_par_add = {}
                sl_par_msg = {}
                for signal in sl_obj_cpu_mko_index[objects[0]][cpu]:
                    second_cpu = signal[-1][1] if isinstance(signal[-1], tuple) else signal[-1]
                    first_obj_rus = objects[1]
                    second_obj_rus = sl_object_rus_name.get(signal[-1][0], '_') \
                        if isinstance(signal[-1], tuple) else objects[1]

                    # Проверка на существование обмена между контроллерами в конфигураторе
                    if not((cpu, second_cpu) in set_mko_par_cpu or (second_cpu, cpu) in set_mko_par_cpu):
                        continue
                    signal_name = f"{cpu}_{signal[0].replace('|', '_')}"
                    text_line_link = 'основной' if 'noping1' in signal[0].lower() or 'noping3' in signal[0].lower() \
                        else 'резервной'
                    if second_cpu in sl_cpu_res and ('noping3' in signal[0].lower() or 'noping4' in signal[0].lower()):
                        second_cpu_rus = sl_cpu_res.get(second_cpu, '_!')[1]
                    else:
                        second_cpu_rus = sl_cpu_rus_name.get(signal[-1][1]) \
                            if isinstance(signal[-1], tuple) else sl_cpu_rus_name.get(signal[-1])
                    text_msg = f'Нет связи между контроллерами ' \
                               f'{sl_cpu_rus_name.get(cpu)} ({first_obj_rus}) и ' \
                               f'{second_cpu_rus} ({second_obj_rus}) по {text_line_link} линии связи'
                    sl_par_add[signal_name] = ('NoPing.NoPing_PLC_View', signal[1])
                    sl_par_msg[signal_name] = (text_msg,)
                    # Если второй контроллер не резервируемый, то удаляем noping3 и noping4
                    if ('noping3' in signal[0].lower() or 'noping4' in signal[0].lower()) \
                            and not sl_cpu_res.get(second_cpu) \
                            and sl_par_add.get(signal_name) \
                            and sl_par_msg.get(signal_name):
                        del sl_par_add[signal_name]
                        del sl_par_msg[signal_name]
                    # print(f'{cpu}-{second_cpu}',(cpu, second_cpu) in set_mko_par_cpu,
                    # (second_cpu, cpu) in set_mko_par_cpu)
                    # print(first_obj_rus, second_obj_rus)
                    # print(signal, 'signal')
                    # print(text_msg, 'text_msg')

                if objects not in sl_add_obj_cpu_mko:
                    sl_add_obj_cpu_mko[objects] = {}
                sl_add_obj_cpu_mko[objects].update({cpu: sl_par_add})
                if cpu not in sl_add_cpu_mko:
                    sl_add_cpu_mko[cpu] = {
                        f'Diag.Connect.{par}.Value': (val[1], 'B') for par, val in sl_par_add.items()
                    }

                # print('sl_par_add', sl_par_add)
                # print('sl_par_msg', sl_par_msg)
                tuple_attr = ('unit.System.Attributes.Description', )
                add_xml_par_plc(name_group='Connect', sl_par=sl_par_add,
                                parent_node=child_diag,
                                sl_attr_par={alg_par: dict(zip(tuple_attr, value))
                                             for alg_par, value in sl_par_msg.items()})

            # Создаём узел System
            child_system = ET.SubElement(child_app, 'ct_object', name="System", access_level="public")

            # Пробегаемся по словарю ключами
            for node in ('SET', 'BTN', 'PZ', 'CNT', 'TS', 'PPU', 'ALR', 'ALG', 'WRN', 'MODES', 'GRH', 'CDO', 'PRU', 'Pars'):
                # если у текущего контроллера есть анпары или...
                if cpu in sl_w[node]['dict']:
                    tuple_attr = sl_w[node]['tuple_attr']
                    sl_w_node_dict_cpu_copy = copy(sl_w[node]['dict'][cpu])
                    if sl_w[node].get('dict_diff'):
                        # print(objects[1], sl_w[node]['dict_diff'][cpu])
                        for alg_par, value_diff in sl_w[node]['dict_diff'][cpu].items():
                            if f'{objects[1]}(нет)' in value_diff:
                                # print(objects[1], alg_par)
                                sl_w_node_dict_cpu_copy[alg_par] = \
                                    sl_w_node_dict_cpu_copy[alg_par][:1] + (sl_w_node_dict_cpu_copy[alg_par][1] +
                                                                            '(РЕЗЕРВ)',) + \
                                    sl_w_node_dict_cpu_copy[alg_par][2:]
                    add_xml_par_plc(name_group=node,
                                    sl_par=sl_w[node]['dict'][cpu],
                                    parent_node=child_system,
                                    sl_attr_par={alg_par: dict(zip(tuple_attr, value[1:]))
                                                 for alg_par, value in sl_w_node_dict_cpu_copy.items()})

            if cpu in return_sl_cpu_drv:
                # return_sl_cpu_drv = {cpu: {(Драйвер, рус имя драйвера):
                # {алг.пар: (Тип переменной, рус имя, тип сообщения, цвет отключения,
                # цвет включения, ед. изм., кол-во знаков, Сохраняемый - Да/Нет, кол-во знаков) }}}
                child_drv_node = ET.SubElement(child_system, 'ct_object', name="DRV", access_level="public")
                tuple_attr = ('unit.System.Attributes.Description', 'Attributes.Type_wrn_DRV',
                              'Attributes.sColorOff', 'Attributes.sColorOn',
                              'Attributes.EUnit', 'Attributes.FracDigits')
                for drv, sl_sig_drv in return_sl_cpu_drv[cpu].items():
                    add_xml_par_plc(name_group=drv,
                                    sl_par=sl_sig_drv,
                                    parent_node=child_drv_node,
                                    sl_attr_par={alg_par: dict(zip(tuple_attr, value[1:]))
                                                 for alg_par, value in sl_sig_drv.items()})

            # Если есть ТР в данном контроллере...
            if 'ТР' in sl_CPU_spec[cpu] and sl_TR:
                # Создаём узел TR, если в настроечном файле есть выбранный топливник
                if choice_tr in sl_TR:
                    if len(sl_TR[choice_tr]) == 1:
                        sub_node_tr = ''.join(sl_TR[choice_tr])
                        ET.SubElement(child_system, 'ct_object', name="TR",
                                      base_type=f'Types.TR.{choice_tr}.{sub_node_tr}.{sub_node_tr}_PLC_View',
                                      aspect="Types.PLC_Aspect", access_level="public")
                    else:
                        child_TR = ET.SubElement(child_system, 'ct_object', name="TR", access_level="public")
                        for sub_node_tr in sl_TR[choice_tr]:
                            ET.SubElement(child_TR, 'ct_object', name=f"{sub_node_tr.replace('TR_', '')}",
                                          base_type=f'Types.TR.{choice_tr}.{sub_node_tr}.{sub_node_tr}_PLC_View',
                                          aspect="Types.PLC_Aspect", access_level="public")

            # Нормируем и записываем PLC-аспект
            temp = ET.tostring(root_plc_aspect).decode('UTF-8')
            check_diff_file(check_path=os.path.join('File_for_Import', 'PLC_Aspect_importDomain'),
                            file_name_check=f'file_out_plc_{cpu}_{objects[2]}.omx-export',
                            new_data=multiple_replace_xml(lxml.etree.tostring(lxml.etree.fromstring(temp),
                                                                              pretty_print=True, encoding='unicode')),
                            message_print=f'Требуется заменить ПЛК-аспект контроллера {cpu}_{objects[2]}')
            # print(datetime.datetime.now(), f' - Создали аспект {cpu}_{objects[2]}')
            progress_percent += 1
            cur_percent = progress_percent * 100 / progress_ios_plc
            # rprint(f"Процент создания/проверки IOS и PLC Аспектов:{round(cur_percent, 2)}%")
            # print(f"[{progress_bar.replace(' ', '=', round(cur_percent))}]")

        # Если у текущего объекта есть контроллеры с САР на борту
        if set(sl_object_all[objects].keys()) & set([i for i in sl_CPU_spec if 'САР' in sl_CPU_spec[i]]):
            set_cpu_sar = set([i for i in sl_CPU_spec if 'САР' in sl_CPU_spec[i]])
            # На случай, если в объекте будут несколько контроллеров с САР, то в IOS аспекте будут созданы несколько
            for cpu_sar in set_cpu_sar:
                if cpu_sar in set(sl_object_all[objects].keys()):
                    child_ios_sar = ET.SubElement(
                        child_object, 'ct_object',
                        name=f"SAR" if tuple(sl_object_all[objects].keys()).count(cpu_sar) == 1 else f'SAR_{cpu_sar}',
                        access_level="public")
                    ET.SubElement(child_ios_sar, 'attribute', type='unit.System.Attributes.Description', value=f'САР')
                    for agreg, type_agreg in sl_agreg.items():
                        ET.SubElement(child_ios_sar, 'ct_object', name=f'{agreg}', base_type=f"{type_agreg}",
                                      aspect="Types.IOS_Aspect", access_level="public")
                    child_ios_sar_im = ET.SubElement(child_ios_sar, 'ct_object', name=f"IM", access_level="public")
                    for agreg, type_agreg in sl_agreg.items():
                        ET.SubElement(child_ios_sar_im, 'ct_object', name=f'{agreg}', base_type=f"{type_agreg}",
                                      aspect="Types.IOS_Aspect", access_level="public")
                    child_ios_sar_reload = ET.SubElement(child_ios_sar, 'ct_object', name=f"Reload",
                                                         access_level="public")
                    flag_reload_standart = False
                    if os.path.exists(os.path.join('Template_Alpha', 'SAR', 'Reload_checkbox')):
                        with open(os.path.join('Template_Alpha', 'SAR', 'Reload_checkbox'), 'r',
                                  encoding='UTF-8') as f_:
                            for line in f_:
                                if '#' in line:
                                    continue
                                if not line.strip():
                                    break
                                lst_sar_reload = [i.strip() for i in line.strip().split(';')]
                                if len(lst_sar_reload) == 3:
                                    ios_reload = ET.SubElement(
                                        child_ios_sar_reload, 'ct_object', name=f"{lst_sar_reload[1]}",
                                        base_type=f"Types.SAR_Switch_Reload.SAR_Switch_Reload_IOS_View",
                                        original=f"PLC_{cpu_sar}_{objects[2]}.CPU.Tree.SAR.Reload.{lst_sar_reload[1]}",
                                        aspect="Types.IOS_Aspect", access_level="public"
                                    )
                                    ET.SubElement(
                                        ios_reload, 'ct_init-ref',
                                        ref="_PLC_View",
                                        target=f"PLC_{cpu_sar}_{objects[2]}.CPU.Tree.SAR.Reload.{lst_sar_reload[1]}"
                                    )
                                else:
                                    flag_reload_standart = True
                    else:
                        flag_reload_standart = True

                    if flag_reload_standart:
                        for i in range(1, 10):
                            ios_km = ET.SubElement(child_ios_sar_reload, 'ct_object', name=f"GPA_KM{i}",
                                                   base_type=f"Types.SAR_Switch_Reload.SAR_Switch_Reload_IOS_View",
                                                   original=f"PLC_{cpu_sar}_{objects[2]}.CPU.Tree.SAR.Reload.GPA_KM{i}",
                                                   aspect="Types.IOS_Aspect", access_level="public")
                            ET.SubElement(ios_km, 'ct_init-ref',
                                          ref="_PLC_View",
                                          target=f"PLC_{cpu_sar}_{objects[2]}.CPU.Tree.SAR.Reload.GPA_KM{i}")
                            ios_mk = ET.SubElement(child_ios_sar_reload, 'ct_object', name=f"GPA_MK{i}",
                                                   base_type=f"Types.SAR_Switch_Reload.SAR_Switch_Reload_IOS_View",
                                                   original=f"PLC_{cpu_sar}_{objects[2]}.CPU.Tree.SAR.Reload.GPA_MK{i}",
                                                   aspect="Types.IOS_Aspect", access_level="public")
                            ET.SubElement(ios_mk, 'ct_init-ref',
                                          ref="_PLC_View",
                                          target=f"PLC_{cpu_sar}_{objects[2]}.CPU.Tree.SAR.Reload.GPA_MK{i}")

                    child_ios_khr = ET.SubElement(child_ios_sar_im, 'ct_object', name=f"KHR",
                                                  base_type='Types.SAR_Classic.KHR_IOS_View',
                                                  aspect="Types.IOS_Aspect",
                                                  original=f"PLC_{cpu_sar}_{objects[2]}.CPU.Tree.SAR.IM.KHR",
                                                  access_level="public")
                    ET.SubElement(child_ios_khr, 'ct_init-ref',
                                  ref="_PLC_View", target=f"PLC_{cpu_sar}_{objects[2]}.CPU.Tree.SAR.IM.KHR")
                    ET.SubElement(child_ios_khr, 'attribute', type='unit.System.Attributes.Description', value=f'КХР')
                    child_ios_sar_regul = ET.SubElement(child_ios_sar, 'ct_object', name=f"REGUL",
                                                        base_type='Types.SAR_Classic.REGUL_IOS_View',
                                                        aspect="Types.IOS_Aspect",
                                                        original=f"PLC_{cpu_sar}_{objects[2]}.CPU.Tree.SAR.REGUL",
                                                        access_level="public")
                    ET.SubElement(child_ios_sar_regul, 'ct_init-ref',
                                  ref="_PLC_View", target=f"PLC_{cpu_sar}_{objects[2]}.CPU.Tree.SAR.REGUL")
                    child_ios_sar_wrn = ET.SubElement(child_ios_sar, 'ct_object', name=f"WRN",
                                                      base_type='Types.SAR_Classic.WRN_IOS_View',
                                                      aspect="Types.IOS_Aspect",
                                                      original=f"PLC_{cpu_sar}_{objects[2]}.CPU.Tree.SAR.WRN",
                                                      access_level="public")
                    ET.SubElement(child_ios_sar_wrn, 'ct_init-ref',
                                  ref="_PLC_View", target=f"PLC_{cpu_sar}_{objects[2]}.CPU.Tree.SAR.WRN")
                    # os.path.dirname(sys.argv[0])
                    if os.path.exists(os.path.join(os.path.abspath(os.curdir),
                                                   'Template_Alpha', 'SAR', 'Tun_SAR.txt')):
                        child_ios_sar_tun = ET.SubElement(child_ios_sar, 'ct_object', name=f"Tuning",
                                                          access_level="public")
                        with open(os.path.join(os.path.abspath(os.curdir), 'Template_Alpha', 'SAR', 'Tun_SAR.txt'),
                                  'r', encoding='UTF-8') as f_signal:
                            for line in f_signal:
                                if '#' in line:
                                    continue
                                if not line.strip():
                                    break
                                lst_tun = line.strip().split(';')
                                if lst_tun[0] == 'sSetWord':
                                    types_tun = f'SAR_tuning_sSetWord'
                                else:
                                    types_tun = f'SAR_tuning' if lst_tun[1] == 'REAL' else f'SAR_tuning_int'
                                object_ios_sar_tuning = ET.SubElement(
                                    child_ios_sar_tun, 'ct_object', name=f"{lst_tun[0]}",
                                    base_type=f"Types.{types_tun}.{types_tun}_IOS_View",
                                    original=f"PLC_{cpu_sar}_{objects[2]}.CPU.Tree.SAR.Tuning.{lst_tun[0]}",
                                    aspect="Types.IOS_Aspect", access_level="public")
                                ET.SubElement(object_ios_sar_tuning, 'ct_init-ref',
                                              ref="_PLC_View",
                                              target=f"PLC_{cpu_sar}_{objects[2]}.CPU.Tree.SAR.Tuning.{lst_tun[0]}")

        # Пробегаемся по словарю ключами анпаров, расчётными и дискретными
        for node in ('AI', 'AE', 'DI', 'IM'):
            # Если у текущего объекта есть контроллеры с анпарами...
            if set(sl_object_all[objects].keys()) & set(sl_w[node]['dict'].keys()):
                # print(node, sl_w[node]['dict'])
                add_xml_par_ios(set_cpu_object=set(sl_object_all[objects].keys()),
                                objects=objects, name_group=node,
                                sl_par=sl_w[node]['dict'],
                                parent_node=child_object, sl_agreg=sl_w[node]['dict_agreg_IOS'], plc_node_tree=node)

        # Если у текущего объекта есть контроллеры с АПР на борту
        if set(sl_object_all[objects].keys()) & set([i for i in sl_CPU_spec if 'АПР' in sl_CPU_spec[i]]):
            set_cpu_apr = set([i for i in sl_CPU_spec if 'АПР' in sl_CPU_spec[i]])
            # На случай, если в объекте будут несколько контроллеров с АПР, то в IOS аспекте будут созданы несколько
            for cpu_apr in set_cpu_apr:
                if cpu_apr in set(sl_object_all[objects].keys()):
                    child_apr = ET.SubElement(
                        child_object, 'ct_object',
                        name=('APR' if tuple(sl_object_all[objects].keys()).count(cpu_apr) == 1 else f'APR_{cpu_apr}'),
                        aspect="Types.IOS_Aspect", access_level="public")
                    # ...добавляем агрегаторы
                    for agreg, type_agreg in sl_agreg.items():
                        ET.SubElement(child_apr, 'ct_object', name=f'{agreg}', base_type=f"{type_agreg}",
                                      aspect="Types.IOS_Aspect", access_level="public")
                    child_apr_im = ET.SubElement(child_apr, 'ct_object', name="IM",
                                                 base_type="Types.APR_IM.APR_IM_IOS_View",
                                                 original=f"PLC_{cpu_apr}_{objects[2]}.CPU.Tree.APR.IM",
                                                 aspect="Types.IOS_Aspect", access_level="public")
                    ET.SubElement(child_apr_im, 'ct_init-ref',
                                  ref="_PLC_View", target=f"PLC_{cpu_apr}_{objects[2]}.CPU.Tree.APR.IM")
                    if sl_tun_apr:
                        add_xml_par_ios(set_cpu_object=set(sl_object_all[objects].keys()),
                                        objects=objects, name_group='Tuning',
                                        sl_par=sl_tun_apr, parent_node=child_apr, plc_node_tree='APR.Tuning',
                                        sl_agreg={})

        # Создаём узел Diag
        child_diag = ET.SubElement(child_object, 'ct_object', name='Diag', access_level="public")
        ET.SubElement(child_diag, 'attribute', type='unit.System.Attributes.Description', value=f'Диагностика')
        # ...добавляем агрегаторы
        for agreg, type_agreg in sl_agreg.items():
            ET.SubElement(child_diag, 'ct_object', name=f'{agreg}', base_type=f"{type_agreg}",
                          aspect="Types.IOS_Aspect", access_level="public")

        # Если у текущего объекта есть контроллеры с модулями в HW
        if set(sl_object_all[objects].keys()) & set(sl_modules_cpu.keys()):
            add_xml_par_ios(set_cpu_object=set(sl_object_all[objects].keys()),
                            objects=objects, name_group='HW', sl_par=sl_modules_cpu,
                            parent_node=child_diag, sl_agreg=sl_agreg, plc_node_tree='Diag.HW')

        # Формируем узел Connect - диагностика связи с ПЛК
        if sl_object_all[objects]:
            child_connect = ET.SubElement(child_diag, 'ct_object', name='Connect', access_level="public")
            # ...добавляем агрегаторы
            for agreg, type_agreg in sl_agreg.items():
                ET.SubElement(child_connect, 'ct_object', name=f'{agreg}', base_type=f"{type_agreg}",
                              aspect="Types.IOS_Aspect", access_level="public")

            for cpu_connect in sl_object_all[objects]:
                name_cpu = ''.join([key for key, value in sl_modules_cpu.get(cpu_connect, {}).items()
                                    if 'CPU' in value])

                num_unet = ''.join([f"{num_u}" for num_u, i in all_cpu_select.items()
                                    if f'{cpu_connect}!{objects[2]}' in i])

                # Если имя непустое, то есть ПЛК объявлен на листе модулей, то добавляем по нему коннект в объект
                # также проверяем существование определённого номера модуля юнет
                if name_cpu and num_unet and int(num_unet) in all_cpu_select:
                    for num_port in range(1, 3):
                        child_sig_con = ET.SubElement(child_connect, 'ct_parameter',
                                                      name=f'Connect_{cpu_connect}{objects[2]}_port_{num_port}',
                                                      type='bool', direction='out', access_level="public")
                        ET.SubElement(child_sig_con, 'attribute', type='unit.Server.Attributes.Alarm',
                                      value=f'{{"Condition":{{"IsEnabled":"true",'
                                            f'"Subconditions":[{{"AckStrategy":2,"IsDeactivation":true,'
                                            f'"Message":". Нет связи с {name_cpu}. Порт {num_port}",'
                                            f'"Severity":40,"Type":2}},'
                                            f'{{"AckStrategy":2,"IsEnabled":true,'
                                            f'"Message":". Нет связи с {name_cpu}. Порт {num_port}",'
                                            f'"Severity":40,"Type":3}}],'
                                            f'"Type":2}}}}')
                        ET.SubElement(child_sig_con, 'attribute', type='unit.System.Attributes.InitialValue',
                                      value='true')
                        ET.SubElement(child_connect, 'ct_subject-ref',
                                      name=f'_{cpu_connect}_{objects[2]}_Eth{num_port}',
                                      object=f"Service.Modules."
                                             f"UNET Client{num_unet}.PLC_{cpu_connect}_{objects[2]}.CPU_Eth{num_port}",
                                      const_access="false", aspected="false", access_level="public")
                        ET.SubElement(child_connect, 'ct_bind',
                                      source=f"_{cpu_connect}_{objects[2]}_Eth{num_port}.IsConnected",
                                      target=f'Connect_{cpu_connect}{objects[2]}_port_{num_port}', action="set_all")
                    # Добавляем диагностику по резервным контроллерам, если такие есть
                    if name_cpu in sl_cpu_res.get(cpu_connect, ''):
                        name_res = sl_cpu_res[cpu_connect][1]
                        for num_port in range(1, 3):
                            child_sig_con = ET.SubElement(child_connect, 'ct_parameter',
                                                          name=f'Connect_{cpu_connect}{objects[2]}_port_{num_port}_res',
                                                          type='bool', direction='out', access_level="public")
                            ET.SubElement(child_sig_con, 'attribute', type='unit.Server.Attributes.Alarm',
                                          value=f'{{"Condition":{{"IsEnabled":"true",'
                                                f'"Subconditions":[{{"AckStrategy":2,"IsDeactivation":true,'
                                                f'"Message":". Нет связи с {name_res}. Порт {num_port}",'
                                                f'"Severity":40,"Type":2}},'
                                                f'{{"AckStrategy":2,"IsEnabled":true,'
                                                f'"Message":". Нет связи с {name_res}. Порт {num_port}",'
                                                f'"Severity":40,"Type":3}}],'
                                                f'"Type":2}}}}')
                            ET.SubElement(child_sig_con, 'attribute', type='unit.System.Attributes.InitialValue',
                                          value='true')
                            ET.SubElement(child_connect, 'ct_subject-ref',
                                          name=f'_{cpu_connect}_{objects[2]}_Eth{num_port}_res',
                                          object=f"Service.Modules."
                                                 f"UNET Client{num_unet}.PLC_{cpu_connect}_{objects[2]}.R-CPU_Eth{num_port}",
                                          const_access="false", aspected="false", access_level="public")
                            ET.SubElement(child_connect, 'ct_bind',
                                          source=f"_{cpu_connect}_{objects[2]}_Eth{num_port}_res.IsConnected",
                                          target=f'Connect_{cpu_connect}{objects[2]}_port_{num_port}_res',
                                          action="set_all")

            # Добавляем диагностику МКО, если такая есть
            if sl_obj_cpu_mko_index.get(objects[0]) and sl_add_obj_cpu_mko.get(objects):
                # print(sl_add_obj_cpu_mko)
                for cpu, sl_par_mko in sl_add_obj_cpu_mko[objects].items():
                    for par_mko_noping, tuple_par in sl_par_mko.items():
                        child_par = ET.SubElement(child_connect, 'ct_object', attrib={
                            'name': f"{par_mko_noping}",
                            'base-type': f"Types.{tuple_par[0].replace('PLC_View', 'IOS_View')}",
                            'aspect':  "Types.IOS_Aspect",
                            'original': f"PLC_{cpu}_{objects[2]}.CPU.Tree.Diag.Connect.{par_mko_noping}",
                            'access-level': 'public'
                        })
                        ET.SubElement(child_par, 'ct_init-ref',
                                      ref="_PLC_View",
                                      target=f"PLC_{cpu}_{objects[2]}.CPU.Tree.Diag.Connect.{par_mko_noping}")

        # Формируем узел NET, при условии, что в данном объекте что-то такое есть
        if objects[0] in return_sl_net:
            child_net = ET.SubElement(child_diag, 'ct_object', name='NET', access_level="public")
            # ...добавляем агрегаторы
            for agreg, type_agreg in sl_agreg.items():
                ET.SubElement(child_net, 'ct_object', name=f'{agreg}', base_type=f"{type_agreg}",
                              aspect="Types.IOS_Aspect", access_level="public")
            for alg, sl_value in return_sl_net[objects[0]].items():
                if "checkip" in sl_value.get('Type', 'бла').lower():
                    # pass
                    # print(alg, sl_value)
                    child_checkip = ET.SubElement(child_net, 'ct_object', name=f'{alg}',  # {objects[0]}_
                                                  access_level="public")
                    ET.SubElement(child_checkip, 'attribute', type=f"unit.System.Attributes.Description",
                                  value=f"{sl_value.get('Unit', 'юнит не определён')}")
                    # ...добавляем агрегаторы
                    for agreg, type_agreg in sl_agreg.items():
                        ET.SubElement(child_checkip, 'ct_object', name=f'{agreg}', base_type=f"{type_agreg}",
                                      aspect="Types.IOS_Aspect", access_level="public")
                    ET.SubElement(child_checkip, 'ct_subject-ref', name=f'_State', object=f"Service.State",
                                  const_access="false",
                                  aspected="false")
                    # Если устройство не определено как составное или сервер, то добавляем обработку двух портов
                    if sl_value.get('SType') and sl_value.get('SType') not in ('<SRV>', '<PART>'):
                        # ...добавляем обработку двух портов
                        for ip_key, port in {'IP': "Порт 1", 'IP_res': "Порт 2"}.items():
                            if sl_value.get(ip_key):
                                # print(ip_key, sl_value.get(ip_key))
                                ET.SubElement(child_checkip, 'ct_object', name=f'Ping_{ip_key}',
                                              access_level="public", base_type="Types.Ping")

                                child_ping_status = ET.SubElement(child_checkip, 'ct_parameter',
                                                                  name=f'Ping_status_{ip_key}',
                                                                  access_level="public", direction='out', type='bool')
                                # ET.SubElement(child_ping_status, 'attribute', type="unit.Server.Attributes.Replicate",
                                #               value=f"false")
                                ET.SubElement(child_ping_status, 'attribute', type='unit.Server.Attributes.Alarm',
                                              value=f'{{"Condition":{{"IsEnabled":"true",'
                                                    f'"Subconditions":[{{"AckStrategy":2,"IsDeactivation":true,'
                                                    f'"Message":". {port}. Нет связи",'
                                                    f'"Severity":40,"Type":2}},'
                                                    f'{{"AckStrategy":2,"IsEnabled":true,'
                                                    f'"Message":". {port}. Нет связи",'
                                                    f'"Severity":40,"Type":3}}],'
                                                    f'"Type":2}}}}')
                                #

                                object_handler = ET.SubElement(
                                    child_checkip, 'ct_handler', name=f'On_Ping_Status_{ip_key}',
                                    source_code=f'if (_State.Server.Value && Ping_{ip_key}.Status.Value '
                                                f'&& !Ping_status_{ip_key}.Value) {{\n'
                                                f'\tcommit Ping_status_{ip_key} = true;\n'
                                                f'\t}} else if (_State.Server.Value && !Ping_{ip_key}.Status.Value '
                                                f'&& Ping_status_{ip_key}.Value '
                                                f'|| Ping_status_{ip_key}.Quality< 192) {{\n'
                                                f'\tcommit Ping_status_{ip_key} = false;\n'
                                                f'}}'
                                )
                                ET.SubElement(object_handler, 'ct_trigger',
                                              on=f"Ping_{ip_key}.Status", cause="update")
                    else:
                        # Если устройство определено как составное или сервер, то добавляем обработку
                        # всех найденных портов
                        if sl_value.get('IPs'):
                            for i, ip in enumerate(sl_value.get('IPs'), start=1):
                                ET.SubElement(child_checkip, 'ct_object', name=f'Ping_IP{i}',
                                              access_level="public", base_type="Types.Ping")
                                child_ping_status = ET.SubElement(child_checkip, 'ct_parameter',
                                                                  name=f'Ping_status_IP{i}',
                                                                  access_level="public", direction='out', type='bool')
                                rus_name_unit = sl_value.get('Unit', '')  # Раньше добавлялось в ПС, хз зачем
                                ET.SubElement(child_ping_status, 'attribute', type='unit.Server.Attributes.Alarm',
                                              value=f'{{"Condition":{{"IsEnabled":"true",'
                                                    f'"Subconditions":[{{"AckStrategy":2,"IsEnabled":true,'
                                                    f'"Message":". Порт{i}. Нет связи",'
                                                    f'"Severity":40,"Type":2}},'
                                                    f'{{"AckStrategy":2,"IsDeactivation":true,'
                                                    f'"Message":". Порт{i}. Нет связи",'
                                                    f'"Severity":40,"Type":3}}],'
                                                    f'"Type":2}}}}')

                                object_handler = ET.SubElement(
                                    child_checkip, 'ct_handler', name=f'On_Ping_Status_IP{i}',
                                    source_code=f'if (_State.Server.Value && Ping_IP{i}.Status.Value '
                                                f'&& !Ping_status_IP{i}.Value) {{\n'
                                                f'\tcommit Ping_status_IP{i} = true;\n'
                                                f'\t}} else if (_State.Server.Value && !Ping_IP{i}.Status.Value '
                                                f'&& Ping_status_IP{i}.Value '
                                                f'|| Ping_status_IP{i}.Quality< 192) {{\n'
                                                f'\tcommit Ping_status_IP{i} = false;\n'
                                                f'}}'
                                )
                                ET.SubElement(object_handler, 'ct_trigger',
                                              on=f"Ping_IP{i}.Status", cause="update")
                else:
                    if sl_value.get('SType') and sl_value.get('SType') not in ('<SRV>', '<PART>'):
                        child_alg = ET.SubElement(child_net, 'ct_object', name=f'{alg}',  # {objects[0]}_
                                                  base_type=f"Types.SNMP_Switch.{sl_value['Type']}_IOS_View",
                                                  aspect="Types.IOS_Aspect",
                                                  original=f"Domain.{objects[0]}_{alg}.Runtime.Application.Data.Data",
                                                  access_level="public")
                        ET.SubElement(child_alg, 'attribute', type=f"unit.System.Attributes.Description",
                                      value=f"{sl_value['Unit']}")
                        ET.SubElement(child_alg, 'ct_init-ref', ref="_PLC_View",
                                      target=f"Domain.{objects[0]}_{alg}.Runtime.Application.Data.Data")
                        ET.SubElement(child_alg, 'ct_subject-ref', name=f'_State', object=f"Service.State",
                                      const_access="false",
                                      aspected="false")
                        object_handler = ET.SubElement(
                            child_alg, 'ct_handler', name=f'On_Ping_Status',
                            source_code=f'if (_State.Server.Value && Ping_IP.Status.Value && !Ping_status.Value) {{\n'
                                        f'\tcommit Ping_status = true;\n'
                                        f'\t}} else if (_State.Server.Value && !Ping_IP.Status.Value && Ping_status.Value '
                                        f'|| Ping_status.Quality< 192) {{\n'
                                        f'\tcommit Ping_status = false;\n'
                                        f'}}'
                        )
                        ET.SubElement(object_handler, 'ct_trigger',
                                      on=f"Ping_IP.Status", cause="update")
                        if sl_value['Option']:
                            lst_ai = sl_value['Option'].split(';')[:12]
                            lst_di = sl_value['Option'].split(';')[12:]
                            # print(lst_ai)
                            for i, ai in enumerate([lst_ai[:3], lst_ai[3:6], lst_ai[6:9], lst_ai[9:12]]):
                                if ''.join(ai) and 'Не используется' not in ai:
                                    # print(i, ai)
                                    diap = float(len([_ for _ in range(int(ai[1]), int(ai[2]))]))
                                    # print(diap)
                                    object_sig_ai = ET.SubElement(child_alg, 'ct_parameter', name=f'aiValue{i}',
                                                                  type='float32', direction='out', access_level="public")
                                    ET.SubElement(object_sig_ai, 'attribute', type="unit.System.Attributes.Description",
                                                  value=f"{ai[0]}")
                                    # <attribute type="unit.Server.Attributes.Replicate" value="false" />
                                    ET.SubElement(object_sig_ai, 'attribute', type="unit.Server.Attributes.Replicate",
                                                  value=f"false")
                                    ET.SubElement(child_alg, 'ct_formula',
                                                  source_code=f"{float(ai[1])} + (TypeConvert.ToFloat(_PLC_View.Signals_"
                                                              f"{sl_value['Type']}.aiValue{i})/65535)*{diap}",
                                                  target=f"aiValue{i}")
                            for i, di in enumerate(lst_di):
                                if di and 'Не используется' not in di:
                                    # print(i, di)
                                    description_ps = di.replace('(Инверсия)', '')
                                    object_sig_di = ET.SubElement(child_alg, 'ct_parameter', name=f'dioStatus{i}',
                                                                  type='bool', direction='out', access_level="public")
                                    ET.SubElement(object_sig_di, 'attribute', type="unit.System.Attributes.Description",
                                                  value=f"{description_ps}")
                                    ET.SubElement(object_sig_di, 'attribute', type="unit.Server.Attributes.Replicate",
                                                  value=f"false")
                                    ET.SubElement(child_alg, 'ct_formula',
                                                  source_code=f"TypeConvert.ToBool"
                                                              f"(_PLC_View.Signals_{sl_value['Type']}.dioStatus{i}) "
                                                              f"&& _State.Server",
                                                  target=f"dioStatus{i}")
                                    if '(Инверсия)' in di:
                                        ET.SubElement(object_sig_di, 'attribute', type='unit.Server.Attributes.Alarm',
                                                      value=f'{{"Condition":{{"IsEnabled":"true",'
                                                            f'"Subconditions":[{{"AckStrategy":2,"IsDeactivation":true,'
                                                            f'"Message":". {description_ps}",'
                                                            f'"Severity":40,"Type":2}},'
                                                            f'{{"AckStrategy":2,"IsEnabled":true,'
                                                            f'"Message":". {description_ps}",'
                                                            f'"Severity":40,"Type":3}}],'
                                                            f'"Type":2}}}}')
                                    else:
                                        ET.SubElement(object_sig_di, 'attribute', type='unit.Server.Attributes.Alarm',
                                                      value=f'{{"Condition":{{"IsEnabled":"true",'
                                                            f'"Subconditions":[{{"AckStrategy":2,"IsEnabled":true,'
                                                            f'"Message":". {description_ps}",'
                                                            f'"Severity":40,"Type":2}},'
                                                            f'{{"AckStrategy":2,"IsDeactivation":true,'
                                                            f'"Message":". {description_ps}",'
                                                            f'"Severity":40,"Type":3}}],'
                                                            f'"Type":2}}}}')

        # Создаём узел System
        child_system = ET.SubElement(child_object, 'ct_object', name='System', access_level="public")
        # ...добавляем агрегаторы
        for agreg, type_agreg in sl_agreg.items():
            ET.SubElement(child_system, 'ct_object', name=f'{agreg}', base_type=f"{type_agreg}",
                          aspect="Types.IOS_Aspect", access_level="public")

        # Пробегаемся по словарю ключами
        for node in ('SET', 'BTN', 'PZ', 'CNT', 'TS', 'PPU', 'ALR', 'ALG', 'WRN', 'MODES', 'GRH', 'CDO', 'PRU', 'Pars'):
            # Если у текущего объекта есть контроллеры с анпарами...
            if set(sl_object_all[objects].keys()) & set(sl_w[node]['dict'].keys()):
                add_xml_par_ios(set_cpu_object=set(sl_object_all[objects].keys()),
                                objects=objects, name_group=node,
                                sl_par=sl_w[node]['dict'],
                                parent_node=child_system, sl_agreg=sl_w[node]['dict_agreg_IOS'],
                                plc_node_tree=f'System.{node}')

        # Формируем драйверные переменные в IOS-аспекте
        # Если в текущем объекте есть контроллеры с драйверами...
        if set(sl_object_all[objects].keys()) & set(return_sl_cpu_drv.keys()):
            # ...то создаём узел DRV
            child_drv_node = ET.SubElement(child_system, 'ct_object', name='DRV', access_level="public")
            # ...добавляем агрегаторы
            for agreg, type_agreg in sl_agreg.items():
                ET.SubElement(child_drv_node, 'ct_object', name=f'{agreg}', base_type=f"{type_agreg}",
                              aspect="Types.IOS_Aspect", access_level="public")
            # для каждого драйвера...
            for driver in return_ios_drv:
                # print(objects[0], driver, set(return_ios_drv[driver].keys()) & set(sl_object_all[objects].keys()))
                # создаём узлы драйверов, если в объекте есть контроллеры данного драйвреа
                if not (set(return_ios_drv[driver].keys()) & set(sl_object_all[objects].keys())):
                    continue
                add_xml_par_ios(set_cpu_object=set(sl_object_all[objects].keys()),
                                objects=objects, name_group=driver,
                                sl_par=return_ios_drv[driver],
                                parent_node=child_drv_node,
                                sl_agreg=sl_agreg,
                                plc_node_tree=f'System.DRV.{driver[0]}')

            # for cpu_with_drv in sorted(tuple(set(sl_object_all[objects].keys()) & set(return_sl_cpu_drv.keys()))):
            #     for drv_tuple, sl_par in return_sl_cpu_drv[cpu_with_drv].items():
            #         add_xml_par_ios(set_cpu_object=set(sl_object_all[objects].keys()),
            #                         objects=objects, name_group=drv_tuple[0],
            #                         sl_par={cpu_with_drv: sl_par},
            #                         parent_node=child_drv_node,
            #                         sl_agreg=sl_agreg,
            #                         plc_node_tree=f'System.DRV.{drv_tuple[0]}')

        # Если в объекте есть контроллеры, помеченные с ТР
        if set(sl_object_all[objects].keys()) & set(cpu for cpu in sl_CPU_spec if 'ТР' in sl_CPU_spec[cpu]) and sl_TR:
            for cpu_tr in (cpu for cpu in sl_CPU_spec if 'ТР' in sl_CPU_spec[cpu]):
                # Создаём узел ТР, при условии, что выбранный топливник есть в словаре
                if choice_tr in sl_TR:
                    if len(sl_TR[choice_tr]) == 1:
                        sub_node_tr = ''.join(sl_TR[choice_tr])
                        child_sub = ET.SubElement(child_system, 'ct_object', name=f"TR",
                                                  base_type=f'Types.TR.{choice_tr}.'
                                                            f'{sub_node_tr}.{sub_node_tr}_IOS_View',
                                                  original=f"PLC_{cpu_tr}_{objects[2]}.CPU.Tree.System.TR",
                                                  aspect="Types.IOS_Aspect", access_level="public")
                        ET.SubElement(child_sub, 'ct_init-ref', ref="_PLC_View",
                                      target=f"PLC_{cpu_tr}_{objects[2]}.CPU.Tree.System.TR")
                    else:
                        child_TR = ET.SubElement(child_system, 'ct_object', name='TR', access_level="public")
                        # ...добавляем агрегаторы
                        for agreg, type_agreg in sl_agreg.items():
                            ET.SubElement(child_TR, 'ct_object', name=f'{agreg}', base_type=f"{type_agreg}",
                                          aspect="Types.IOS_Aspect", access_level="public")
                        for sub_node_tr in sl_TR[choice_tr]:
                            child_sub = ET.SubElement(child_TR, 'ct_object', name=f"{sub_node_tr.replace('TR_', '')}",
                                                      base_type=f"Types.TR.{choice_tr}."
                                                                f"{sub_node_tr}.{sub_node_tr}_IOS_View",
                                                      original=f"PLC_{cpu_tr}_{objects[2]}.CPU.Tree.System."
                                                               f"TR.{sub_node_tr.replace('TR_', '')}",
                                                      aspect="Types.IOS_Aspect", access_level="public")
                            ET.SubElement(child_sub, 'ct_init-ref',
                                          ref="_PLC_View", target=f"PLC_{cpu_tr}_{objects[2]}.CPU.Tree.System."
                                                                  f"TR.{sub_node_tr.replace('TR_', '')}")
        # Нормируем и записываем IOS-аспект
        temp = ET.tostring(root_ios_aspect).decode('UTF-8')
        check_diff_file(check_path=os.path.join('File_for_Import', 'IOS_Aspect_in_ApplicationServer'),
                        file_name_check=f'file_out_IOS_inApp_{objects[0]}.omx-export',
                        new_data=multiple_replace_xml(lxml.etree.tostring(lxml.etree.fromstring(temp),
                                                                          pretty_print=True, encoding='unicode')),
                        message_print=f'Требуется заменить IOS-аспект объекта {objects[0]}')
        # print(datetime.datetime.now(), f' - Создали аспект IOS-аспект {objects[0]}')
        progress_percent += 1
        cur_percent = progress_percent * 100 / progress_ios_plc
        # rprint(f"Процент создания/проверки IOS и PLC Аспектов:{round(cur_percent, 2)}%")
        # print(f"[{progress_bar.replace(' ', '=', round(cur_percent))}]")

    # print()
    # Создание сервисных сигналов
    is_create_service_signal(sl_object_all=sl_object_all, sl_cpu_res=sl_cpu_res, architecture=architecture,
                             server_name_osn=server_name_osn, server_name_rez=server_name_rez,
                             sl_cpu_archive=sl_cpu_archive,
                             all_cpu_select=all_cpu_select)

    # Создаём папку SYS
    is_create_sys(sl_object_all=sl_object_all, name_prj=name_prj, return_sl_net=return_sl_net,
                  architecture=architecture, sl_agreg=sl_agreg,
                  server_name_osn=server_name_osn, server_name_rez=server_name_rez)
    # # В случае клиент-серверной архитектуры создаём сообщения о резервном переходе
    # if architecture == 'клиент-сервер':
    #     is_create_service_snmp(server_name_rez=server_name_rez)

    # Создаём тренды
    is_create_trends(book=book, sl_object_all=sl_object_all, sl_cpu_spec=sl_CPU_spec, sl_all_drv=sl_all_drv,
                     sl_for_diag=sl_for_diag,
                     sl_need_add=is_update_dict(main_dict=return_sl_all_add_pars, sub_dict=return_sl_ai_add_pars))
    #
    # is_create_trends_jtr(book=book, sl_object_all=sl_object_all, sl_cpu_spec=sl_CPU_spec, sl_all_drv=sl_all_drv,
    #                      sl_for_diag=sl_for_diag,
    #                      sl_need_add=is_update_dict(main_dict=return_sl_all_add_pars, sub_dict=return_sl_ai_add_pars))

    book.close()

    # Создаём карту индексов
    '''
    В СЛУЧАЕ, ЕСЛИ У CPU НЕТ ПЕРЕМЕННЫХ ДАННОГО ТИПА, ТО В СООТВЕТСТВУЮЩЕМ СЛОВАРЕ ЕГО(CPU) НЕ БУДЕТ
    sl_sig_alg - словарь ALG переменных, ключ - cpu, значение - кортеж переменных ALG_
    sl_sig_mod - словарь режимов, ключ - cpu, значение - кортеж режимов (в том числе regNum)
    sl_sig_ppu - словарь ППУ, ключ - cpu, значение - кортеж ППУ
    sl_sig_ts - словарь ТС, ключ - cpu, значение - кортеж ТС
    sl_sig_wrn - словарь ПС, ключ - cpu, значение - кортеж ПС(в том числе DI_)
    sl_pz - словарь защит,  ключ - cpu, значение - кортеж (первая авария в ПЛК, последняя авария в ПЛК)
    sl_CPU_spec - словарь спец. добавок, ключ - cpu, значение - кортеж ('ТР', 'АПР') при наличии таковых в cpu
    sl_for_diag - словарь диагностики, ключ - cpu, значение - словарь: ключ - имя модуля, значение - тип модуля
                                                   в случае CPU - ключ - 'CPU', значение - кортеж (имя cpu,тип cpu)
    sl_cpu_drv_signal - словарь драйверов, ключ - cpu, значение - словарь - ключ - драйвер, значение - кортеж переменных
    sl_grh = словарь алг. переменных, ключ - cpu, значение - кортеж переменных GRH|
    tuple_all_cpu - кортеж всех контроллеров
    
    sl_sig_alr - словарь ALRов, ключ - cpu, значение - кортеж переменных аварий(без ALR|), в том числе АС - 
                 в индексах использую его для вычленения АС
    choice_tr - переменная str, в которой содержится выбранный в конфигураторе ТР
    sl_cpu_drv_iec - словарь переменных IEC, ключ - cpu, значение - словарь: ключ - алг.имя переменной IEC, 
                                                                             значение - тип переменной (R или B)
    sl_cpu_path = Словарь {ПЛК: путь к проекту ПЛК}
    '''
    # поддержать вытягивание индексов АС - сделано, но индексы хорошо бы переписать
    # Возможно, подробней рассмотреть аварии(ALR)!!!

    # return_ts, return_ppu, return_alr, return_alg, return_wrn, return_modes
    sl_sig_alg = {cpu: tuple(value.keys()) for cpu, value in return_alg.items()}
    sl_sig_mod = {cpu: tuple(value.keys()) for cpu, value in return_modes.items()}
    sl_sig_ppu = {cpu: tuple(value.keys()) for cpu, value in return_ppu.items()}
    sl_sig_ts = {cpu: tuple(value.keys()) for cpu, value in return_ts.items()}
    sl_sig_wrn = {cpu: tuple(value.keys()) for cpu, value in return_wrn.items()}
    sl_sig_alr = {cpu: tuple(value.keys()) for cpu, value in return_alr.items()}

    sl_cpu_drv_iec = {plc: {par: value[0].replace('DRV_AI.DRV_AI_PLC_View', 'R').replace('DRV_DI.DRV_DI_PLC_View', 'B')
                            for par, value in sl_driver.get(('IEC', 'Прочие архивируемые параметры'), {}).items()}
                      for plc, sl_driver in return_sl_cpu_drv.items()}
    # print(return_sl_cpu_drv)
    sl_cpu_drv_signal_with_imit = {cpu: {} for cpu in return_sl_cpu_drv}
    for cpu, sl_drv in return_sl_cpu_drv.items():
        for driver, par_sl in sl_drv.items():
            sl_cpu_drv_signal_with_imit[cpu].update({driver[0]: tuple()})
            for param, param_tuple in par_sl.items():
                if 'IMIT' in param_tuple[0]:
                    sl_cpu_drv_signal_with_imit[cpu][driver[0]] += (param,)
    # print(sl_cpu_drv_signal_with_imit)
    # print('sl_add_obj_cpu_mko', sl_add_obj_cpu_mko)
    # print('sl_obj_cpu_mko_index', sl_obj_cpu_mko_index)
    # print('sl_add_cpu_mko', sl_add_cpu_mko)
    if unet_version == 1 or unet_version == "1":
        create_index(tuple_all_cpu=tuple([cpu for obj in sl_object_all for cpu in sl_object_all[obj]]),
                     sl_sig_alg=sl_sig_alg,
                     sl_sig_mod=sl_sig_mod,
                     sl_sig_ppu=sl_sig_ppu,
                     sl_sig_ts=sl_sig_ts,
                     sl_sig_wrn=sl_sig_wrn,
                     sl_pz=sl_pz,
                     sl_cpu_spec=sl_CPU_spec,
                     sl_for_diag=sl_for_diag,
                     sl_cpu_drv_signal=sl_cpu_drv_signal,
                     sl_grh=sl_grh,
                     sl_sig_alr=sl_sig_alr,
                     choice_tr=choice_tr,
                     sl_cpu_drv_iec=sl_cpu_drv_iec,
                     # Передаём словари вида {cpu: кортеж алг. переменных},
                     # чтобы отсечь мусор, который может быть в ПЛК
                     sl_ai_config={cpu: tuple(sl_par.keys()) for cpu, sl_par in return_sl_ai.items()},
                     sl_ae_config={cpu: tuple(sl_par.keys()) for cpu, sl_par in return_sl_ae.items()},
                     sl_di_config={cpu: tuple(sl_par.keys()) for cpu, sl_par in return_sl_di.items()},
                     sl_set_config={cpu: tuple(sl_par.keys()) for cpu, sl_par in return_sl_set.items()},
                     sl_btn_config={cpu: tuple(sl_par.keys()) for cpu, sl_par in return_sl_btn.items()},
                     sl_im_config={cpu: tuple(sl_par.keys()) for cpu, sl_par in return_sl_im.items()},
                     sl_cpu_type_im={cpu: {sl_par[_][0].split('.')[0] for _ in sl_par} for cpu, sl_par in
                                     return_sl_im.items()},
                     sl_cpu_path=sl_cpu_path,
                     buff_size=buff_size,
                     sl_cpu_drv_signal_with_imit=sl_cpu_drv_signal_with_imit,
                     sl_cpu_cdo={cpu: tuple(sl_par.keys()) for cpu, sl_par in return_sl_cdo.items()},
                     sl_cpu_res=sl_cpu_res,
                     sl_add_cpu_mko=sl_add_cpu_mko,
                     sl_pru_config={cpu: tuple(sl_par.keys()) for cpu, sl_par in return_sl_pru.items()},
                     sl_need_add_pars=is_update_dict(main_dict=return_sl_all_add_pars, sub_dict=return_sl_ai_add_pars)
                     )
    elif unet_version == 2 or unet_version == "2":
        create_index_u2(tuple_all_cpu=tuple([cpu for obj in sl_object_all for cpu in sl_object_all[obj]]),
                        sl_sig_alg=sl_sig_alg,
                        sl_sig_mod=sl_sig_mod,
                        sl_sig_ppu=sl_sig_ppu,
                        sl_sig_ts=sl_sig_ts,
                        sl_sig_wrn=sl_sig_wrn,
                        sl_pz=sl_pz,
                        sl_cpu_spec=sl_CPU_spec,
                        sl_for_diag=sl_for_diag,
                        sl_cpu_drv_signal=sl_cpu_drv_signal,
                        sl_grh=sl_grh,
                        sl_sig_alr=sl_sig_alr,
                        choice_tr=choice_tr,
                        sl_cpu_drv_iec=sl_cpu_drv_iec,
                        # Передаём словари вида {cpu: кортеж алг. переменных},
                        # чтобы отсечь мусор, который может быть в ПЛК
                        sl_ai_config={cpu: tuple(sl_par.keys()) for cpu, sl_par in return_sl_ai.items()},
                        sl_ae_config={cpu: tuple(sl_par.keys()) for cpu, sl_par in return_sl_ae.items()},
                        sl_di_config={cpu: tuple(sl_par.keys()) for cpu, sl_par in return_sl_di.items()},
                        sl_set_config={cpu: tuple(sl_par.keys()) for cpu, sl_par in return_sl_set.items()},
                        sl_btn_config={cpu: tuple(sl_par.keys()) for cpu, sl_par in return_sl_btn.items()},
                        sl_im_config={cpu: tuple(sl_par.keys()) for cpu, sl_par in return_sl_im.items()},
                        sl_cpu_type_im={cpu: {sl_par[_][0].split('.')[0] for _ in sl_par} for cpu, sl_par in
                                        return_sl_im.items()},
                        sl_cpu_path=sl_cpu_path,
                        buff_size=buff_size,
                        sl_cpu_drv_signal_with_imit=sl_cpu_drv_signal_with_imit,
                        sl_cpu_cdo={cpu: tuple(sl_par.keys()) for cpu, sl_par in return_sl_cdo.items()},
                        sl_cpu_res=sl_cpu_res,
                        sl_add_cpu_mko={},  # sl_add_cpu_mko, !!!
                        sl_pru_config={cpu: tuple(sl_par.keys()) for cpu, sl_par in return_sl_pru.items()},
                        sl_need_add_pars=is_update_dict(main_dict=return_sl_all_add_pars,
                                                        sub_dict=return_sl_ai_add_pars)
                        )

    # print(return_alr)
    # print(sl_CPU_spec)
    # # добавление отсечки в файл изменений, чтобы разные сборки не сливались
    # if os.path.exists('Required_change.txt'):
    #     with open('Required_change.txt', 'r') as f_test:
    #         check_test = f_test.readlines()[-1]
    #     if check_test != '-' * 70 + '\n':
    #         with open('Required_change.txt', 'a') as f_test:
    #             f_test.write('-' * 70 + '\n')
    # Если есть аналоговые параметры, то делаем шаблон Reports
    if return_sl_ai:
        # Пока заблокирована функция для новых репортов
        # create_reports(sl_object_all=sl_object_all, node_param_rus='Измеряемые параметры', node_alg_name='AI',
        #                sl_param=return_sl_ai)
        create_reports_v80(sl_object_all=sl_object_all, node_param_rus='Измеряемые параметры', node_alg_name='AI',
                           sl_param=return_sl_ai)
    # Если есть расчётные параметры, то делаем шаблон Reports
    if return_sl_ae:
        # Пока заблокирована функция для новых репортов
        # create_reports(sl_object_all=sl_object_all, node_param_rus='Расчетные параметры', node_alg_name='AE',
        #                sl_param=return_sl_ae)
        create_reports_v80(sl_object_all=sl_object_all, node_param_rus='Расчетные параметры', node_alg_name='AE',
                           sl_param=return_sl_ae)
    # Если есть наработки, то делаем шаблон Reports
    if sl_cnt_xml:
        # Пока заблокирована функция для новых репортов
        # create_reports(sl_object_all=sl_object_all, node_param_rus='Наработки', node_alg_name='System.CNT',
        #                sl_param=sl_cnt_xml)
        create_reports_v80(sl_object_all=sl_object_all, node_param_rus='Наработки', node_alg_name='System.CNT',
                           sl_param=sl_cnt_xml)
    # Если есть защиты, то делаем шаблон Reports
    if sl_pz_xml:
        # Пока заблокирована функция для новых репортов
        # create_reports_pz(sl_object_all=sl_object_all, node_param_rus='Протокол проверки защит',
        #                   node_alg_name='System.PZ', sl_param=sl_pz_xml)
        create_reports_pz_v80(sl_object_all=sl_object_all, node_param_rus='Протокол проверки защит',
                              node_alg_name='System.PZ', sl_param=sl_pz_xml)

    if return_sl_sday:
        create_reports_sday(sl_object_all=sl_object_all, node_param_rus='Сменная ведомость', sl_param=return_sl_sday)
    else:
        try:
            create_reports_sday_v80(sl_object_all=sl_object_all, node_param_rus='Сменная ведомость')
            create_reports_sday_v10(sl_object_all=sl_object_all, node_param_rus='Сменная ведомость')
            create_reports_sut_v80(sl_object_all=sl_object_all, node_param_rus='Суточная ведомость')
        except (Exception, KeyError, IndexError):
            print('Файлы DailyList заполнены некорректно, обновите сменную ведомость')

    # Создаём мнемосхемы (их изменение не контролируется по причине смены uuid, позже будет доработано)
    # Проработать вопрос создания мнемосхем для разных объектов!!!
    # return_sl_mnemo = {узел: список параметров узла}
    # Передаём такой словарь, но упорядоченный по узлам, согласно объявлению
    # print(sl_mnemo_cnt)
    if sl_mnemo_ai:
        sl_mnemo_ai = {_: {node: sl_mnemo_ai[_].get(node) for node in tuple_mnemo if node in sl_mnemo_ai[_]}
                       for _ in sl_mnemo_ai}
        create_mnemo_param(name_list='Измеряемые', name_group='AI', name_page='AINP', base_type_param='S_A_INP_Param',
                           size_shirina=size_shirina,
                           size_vysota=size_vysota,
                           sl_param=sl_mnemo_ai,
                           # {node: sl_mnemo_ai.get(node) for node in tuple_mnemo if node in sl_mnemo_ai},
                           sl_object_all=sl_object_all,
                           tuple_mnemo=tuple_mnemo
                           )

    if sl_mnemo_ae:
        sl_mnemo_ae = {_: {node: sl_mnemo_ae[_].get(node) for node in tuple_mnemo if node in sl_mnemo_ae[_]}
                       for _ in sl_mnemo_ae}
        # for _ in sl_mnemo_ae:
        #     for node in sl_mnemo_ae[_]:
        #         print(node, len(sl_mnemo_ae[_][node]), sl_mnemo_ae[_][node])
        create_mnemo_param(name_list='Расчетные', name_group='AE', name_page='AEVL', base_type_param='S_A_INP_Param',
                           size_shirina=size_shirina,
                           size_vysota=size_vysota,
                           sl_param=sl_mnemo_ae,
                           sl_object_all=sl_object_all,
                           tuple_mnemo=tuple_mnemo
                           )

    if sl_mnemo_di:
        sl_mnemo_di = {_: {node: sl_mnemo_di[_].get(node) for node in tuple_mnemo if node in sl_mnemo_di[_]}
                       for _ in sl_mnemo_di}
        create_mnemo_param(name_list='Дискретные', name_group='DI', name_page='DINP', base_type_param='S_D_INP_Param',
                           size_shirina=size_shirina,
                           size_vysota=size_vysota,
                           sl_param=sl_mnemo_di,
                           sl_object_all=sl_object_all,
                           tuple_mnemo=tuple_mnemo
                           )

    if sl_cnt:
        create_mnemo_param(name_list='Наработка', name_group='System.CNT', name_page='SCNT', base_type_param='S_CNT',
                           size_shirina=size_shirina,
                           size_vysota=size_vysota,
                           sl_param=sl_mnemo_cnt,
                           sl_object_all=sl_object_all
                           )
    if sl_pz_xml:
        create_mnemo_pz(name_group='System.PZ', name_page='ПЗ', base_type_param='S_ALR',
                        size_shirina=size_shirina,
                        size_vysota=size_vysota,
                        # param_pz=[one_pz for tuple_pz in [sl_pz[cpu] for cpu in sl_pz] for one_pz in tuple_pz],
                        sl_pz_xml=sl_pz_xml,    # sl_with_check_pz=dict(ChainMap(*sl_pz_xml.values())),
                        sl_object_all=sl_object_all)
    # print()
    # for i in sl_condition_in_cpu:
    #     for mm in sl_condition_in_cpu[i]:
    #         print(mm, sl_condition_in_cpu[i][mm])
    if sl_command_in_cpu and sl_condition_in_cpu and return_modes:
        create_mnemo_visual(sl_object_all=sl_object_all, sl_command_in_cpu=sl_command_in_cpu,
                            sl_condition_in_cpu=sl_condition_in_cpu, sl_mod_cpu=sl_mod_cpu)
    if return_sl_mnemo_cdo:
        create_mnemo_param(name_list='Проверка выходов', name_group='System.CDO', name_page='CDO',
                           base_type_param='S_CDO_Param',
                           size_shirina=size_shirina,
                           size_vysota=size_vysota,
                           sl_param=return_sl_mnemo_cdo,
                           sl_object_all=sl_object_all
                           )
    # print(return_ios_drv)

    # return_ios_drv = {(Драйвер, рус имя драйвера): {cpu:
    # {алг.пар: (Тип переменной в студии, рус имя, тип сообщения, цвет отключения, цвет включения,
    # ед. изм., кол-во знаков) }}}
    if return_ios_drv:
        # sl_mnemo_drv = {driver[1]: list() for driver in return_ios_drv}
        sl_type_drv = {driver[1]: {} for driver in return_ios_drv}
        for driver in return_ios_drv:
            for cpu in return_ios_drv[driver]:
                for drv_par in return_ios_drv[driver][cpu]:
                    # sl_mnemo_drv[driver[1]].append(drv_par)
                    sl_type_drv[driver[1]].update({drv_par: return_ios_drv[driver][cpu][drv_par][0]})
        # for i, j in sl_mnemo_drv.items():
        #     print(i, j)
        # print(sl_type_drv)
        # print(dict(ChainMap(*return_ios_drv.values())).keys())

        sl_mnemo_drv = {drv: {cpu: [i for i in sl_p] for cpu, sl_p in sl_d.items()} for drv, sl_d in return_ios_drv.items()}
        # print(sl_mnemo_drv)
        # {cpu: [i for i in sl_p] for cpu, sl_p in sl_d.items}

        if not os.path.exists(os.path.join('File_for_Import', 'Mnemo', f'DRV')):
            os.mkdir(os.path.join('File_for_Import', 'Mnemo', f'DRV'))
        if not os.path.exists(os.path.join('File_for_Import', 'Mnemo', 'DRV', 'Systemach')):
            os.mkdir(os.path.join('File_for_Import', 'Mnemo', f'DRV', 'Systemach'))
        # Чистим старые мнемосхемы ПЗ
        # os.path.dirname(sys.argv[0])

        # for file in os.listdir(os.path.join(os.path.abspath(os.curdir), 'File_for_Import', 'Mnemo', 'DRV')):
        #     if file.endswith('.omobj'):
        #         os.remove(os.path.join(os.path.abspath(os.curdir), 'File_for_Import', 'Mnemo', 'DRV', file))

        for driver, sl_one_driver_cpu in return_ios_drv.items():
            create_mnemo_drv(name_group=f'System.DRV.{driver[0]}',
                             name_page=f'{driver[0]}',
                             name_pag_rus=f'{driver[1]}',
                             size_shirina=size_shirina,
                             size_vysota=size_vysota,
                             sl_one_driver_cpu=sl_one_driver_cpu,
                             sl_object_all=sl_object_all)

        create_mnemo_drv_general(sl_object_all=sl_object_all,
                                 name_group="System.DRV",
                                 size_shirina=size_shirina,
                                 size_vysota=size_vysota,
                                 sl_mnemo_drv=sl_mnemo_drv,
                                 sl_type_drv=sl_type_drv,  name_page=f'DRV',
                                 sl_all_drv=sl_all_drv)

    # добавление отсечки в файл изменений, чтобы разные сборки не сливались
    if os.path.exists('Required_change.txt'):
        with open('Required_change.txt', 'r') as f_test:
            check_test = f_test.readlines()[-1]
        if check_test != '-' * 70 + '\n':
            with open('Required_change.txt', 'a') as f_test:
                f_test.write('-' * 70 + '\n')

    print(datetime.datetime.now(), '- Окончание сборки всех файлов')
    input(f'{datetime.datetime.now()} - Сборка файлов завершена успешно. Нажмите Enter для выхода...')
except (Exception, KeyError):
    # print('Произошла ошибка выполнения')
    import sys
    logging.basicConfig(filename='error.log', filemode='a', datefmt='%d.%m.%y %H:%M:%S',
                        format='%(levelname)s - %(message)s - %(asctime)s')
    logging.exception("Ошибка выполнения")
    # os.path.dirname(sys.argv[0])
    for file in os.listdir(os.path.abspath(os.curdir)):
        if file.endswith('.omx-export') or file.endswith('.json'):
            os.remove(file)
    input('Во время сборки произошла ошибка, сформирован файл error.log. Нажмите Enter для выхода...')
