import copy
import openpyxl
# import datetime
import logging
# from my_func import *
from alpha_index import *
import warnings
from json import dumps as json_dumps
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

try:
    # path_config = input('Укажите путь до конфигуратора\n')
    with open('Source_list_config.txt', 'r', encoding='UTF-8') as f:
        path_config = f.readline().strip()

    print(datetime.datetime.now(), '- Начало сборки файлов')

    file_config = 'UnimodCreate.xlsm'

    # Ищем файл конфигуратора в указанном каталоге
    for file in os.listdir(path_config):
        if file.endswith('.xlsm') or file.endswith('.xls'):
            file_config = file
            break

    all_CPU = []
    sl_CPU_spec = {}
    pref_IP = []
    sl_object_rus = {}
    sl_object_all = {}
    num_pz = 0
    lst_all_wrn = []
    lst_all_ts = []
    lst_all_ppu = []
    lst_all_mod = []
    lst_all_alg = []
    sl_all_pz = {}
    sl_for_diag = {}
    sl_cpu_drv_signal = {}
    '''Словарь модулей'''
    sl_modules = {
        'M547A': ['Резерв'] * 16,
        'M537V': ['Резерв'] * 8,
        'M557D': ['Резерв'] * 32,
        'M557O': ['Резерв'] * 32,
        'M932C_2N': ['Резерв'] * 8,
        'M903E': 'CPU',
        'M991E': 'CPU'
    }
    # Считываем файл-шаблон для AI  AE SET
    with open(os.path.join('Template', 'Temp_AIAESET'), 'r', encoding='UTF-8') as f:
        tmp_object_AIAESET = f.read()
    # Считываем файл-шаблон для DI
    with open(os.path.join('Template', 'Temp_DI'), 'r', encoding='UTF-8') as f:
        tmp_object_DI = f.read()
    # Считываем файл-шаблон для IM
    with open(os.path.join('Template', 'Temp_IM'), 'r', encoding='UTF-8') as f:
        tmp_object_IM = f.read()
    # Считываем файл-шаблон для BTN CNT
    with open(os.path.join('Template', 'Temp_BTN_CNT_sig'), 'r', encoding='UTF-8') as f:
        tmp_object_BTN_CNT_sig = f.read()  # его же используем для диагностики CPU потому что подходит
    # Считываем файл-шаблон для PZ
    with open(os.path.join('Template', 'Temp_PZ'), 'r', encoding='UTF-8') as f:
        tmp_object_PZ = f.read()
    # Считываем файл-шаблон для группы
    with open(os.path.join('Template', 'Temp_group'), 'r', encoding='UTF-8') as f:
        tmp_group = f.read()
    # Считываем файл-шаблон для app
    with open(os.path.join('Template', 'Temp_app'), 'r', encoding='UTF-8') as f:
        tmp_app = f.read()
    # Считываем файл-шаблон для контроллера
    with open(os.path.join('Template', 'Temp_TREI'), 'r', encoding='UTF-8') as f:
        tmp_trei = f.read()
    # Считываем файл-шаблон для global
    with open(os.path.join('Template', 'Temp_global'), 'r', encoding='UTF-8') as f:
        tmp_global = f.read()
    # Считываем файл-шаблон для топливного регулятора
    with open(os.path.join('Template', 'Temp_TR_ps90'), 'r', encoding='UTF-8') as f:
        tmp_tr_ps90 = f.read()
    # Считываем файл-шаблон для АПР
    with open(os.path.join('Template', 'Temp_APR'), 'r', encoding='UTF-8') as f:
        tmp_apr = f.read()
    # Считываем файл-шаблон для драйверных параметров
    with open(os.path.join('Template', 'Temp_drv_par'), 'r', encoding='UTF-8') as f:
        tmp_drv_par = f.read()

    book = openpyxl.open(os.path.join(path_config, file_config))  # , read_only=True
    # читаем список всех контроллеров
    sheet = book['Настройки']  # worksheets[1]
    cells = sheet['B2': 'B22']
    for p in cells:
        if p[0].value is not None:
            all_CPU.append(p[0].value)
    # Читаем префиксы IP адреса ПЛК(нужно продумать про новые конфигураторы)
    cells = sheet['A1': 'B' + str(sheet.max_row)]
    for p in cells:
        if p[0].value == 'Cетевая часть адреса основной сети (связь с CPU)':
            pref_IP.append(p[1].value + '.')
        if p[0].value == 'Cетевая часть адреса резервной сети (связь с CPU)':
            pref_IP.append(p[1].value + '.')
            break

    # Читаем состав объектов
    cells = sheet['B24': 'R38']
    for p in cells:
        if p[0].value is not None:
            sl_object_rus[p[0].value] = p[1].value
            tmp0 = []
            tmp1 = []
            for i in range(12, 17):
                if p[i].value == 'ON':
                    tmp0.append(p[i - 10].value)
                    tmp1.append(p[i - 5].value)
            sl_object_all[p[0].value] = [tmp0, tmp1]
    # Мониторинг ТР и АПР в контроллере
    cells = sheet['B1':'L21']
    for p in cells:
        if p[0].value is None:
            break
        else:
            sl_CPU_spec[p[0].value] = []
            if p[is_f_ind(cells[0], 'FLR')].value == 'ON' and p[is_f_ind(cells[0], 'Тип ТР')].value == 'ПС90':
                sl_CPU_spec[p[0].value].append('ТР')
            if p[is_f_ind(cells[0], 'APR')].value == 'ON':
                sl_CPU_spec[p[0].value].append('АПР')

    # Определение заведённых драйверов
    cells = sheet['A1': 'A' + str(sheet.max_row)]
    drv_eng, drv_rus = [], []
    for p in cells:
        if p[0].value == 'Наименование драйвера (Eng)':
            jj = 1
            while sheet[p[0].row][jj].value and sheet[p[0].row + 1][jj].value:
                drv_eng.append(sheet[p[0].row][jj].value)
                drv_rus.append(sheet[p[0].row + 1][jj].value)
                jj += 1
    sl_all_drv = dict(zip(drv_eng, drv_rus))
    ff = open('file_plc.txt', 'w', encoding='UTF-8')
    ff.close()
    # Далее для всех контроллеров, что нашли, делаем
    for i in all_CPU:
        sl_CPU_one = {}
        # Диагностика модулей (DIAG)
        sheet = book['Модули']
        cells = sheet['A1': 'G' + str(sheet.max_row)]
        sl_modules_cpu = {}
        for p in cells:
            if p[is_f_ind(cells[0], 'CPU')].value == i:
                aa = copy.copy(sl_modules[p[is_f_ind(cells[0], 'Шифр модуля')].value])
                sl_modules_cpu[p[is_f_ind(cells[0], 'Имя модуля')].value] = [p[is_f_ind(cells[0], 'Шифр модуля')].value,
                                                                             aa]
        # пробегаемся по листам, где могут быть указаны каналы модулей
        for jj in ['Измеряемые', 'Входные', 'Выходные', 'ИМ(АО)']:
            sheet_run = book[jj]
            cells_run = sheet_run['A1': 'O' + str(sheet_run.max_row)]
            # пробегаемся по параметрам на листе
            for p in cells_run:
                # если параметр принадлежит текущему ПЛК и не указан НЕстандартный канал, то вносим в список
                if p[is_f_ind(cells_run[0], 'CPU')].value == i and \
                        p[is_f_ind(cells_run[0], 'Нестандартный канал')].value == 'Нет':
                    tmp_ind = int(p[is_f_ind(cells_run[0], 'Номер канала')].value) - 1
                    sl_modules_cpu[p[is_f_ind(cells_run[0], 'Номер модуля')].value][1][tmp_ind] = \
                        is_cor_chr(p[is_f_ind(cells_run[0], 'Наименование параметра')].value)
                # если выбран контроль цепи и контроль стандартный, то также добавляем в список
                if p[is_f_ind(cells_run[0], 'CPU')].value == i and \
                        p[is_f_ind(cells_run[0], 'Контроль цепи')].value == 'Да' and \
                        p[is_f_ind(cells_run[0], 'Нестандартный канал КЦ')].value == 'Нет':
                    ind_tmp = int(p[is_f_ind(cells_run[0], 'Номер канала контроля')].value) - 1
                    sl_modules_cpu[p[is_f_ind(cells_run[0], 'Номер модуля контроля')].value][1][ind_tmp] = \
                        f"КЦ: {is_cor_chr(p[is_f_ind(cells_run[0], 'Наименование параметра')].value)}"

        if len(sl_modules_cpu) != 0:
            tmp_line_ = is_create_objects_diag(sl_modules_cpu)
            tmp_line_ = (Template(tmp_group).substitute(name_group='HW', objects=tmp_line_))

            with open('file_out_group.txt', 'w', encoding='UTF-8') as f:
                f.write(Template(tmp_group).substitute(name_group='Diag', objects=tmp_line_.rstrip()))

            # для каждого контроллера сначала создаём пустой словарь в словаре
            sl_for_diag[i] = {}
            # далее проходим по локальному словарю и для каждого модуля грузим {алг.имя модуля: тип модуля}
            for jj in sl_modules_cpu:
                # в случае CPU - {CPU: алг.имя модуля}
                if sl_modules_cpu[jj][0] in ('M903E', 'M991E'):
                    sl_for_diag[i].update({'CPU': jj})
                else:
                    sl_for_diag[i].update({jj: sl_modules_cpu[jj][0]})

        # Измеряемые
        sheet = book['Измеряемые']  # .worksheets[3]
        cells = sheet['A1': 'AG' + str(sheet.max_row)]
        sl_CPU_one = is_load_ai_ae_set(i, cells, is_f_ind(cells[0], 'Алгоритмическое имя'),
                                       is_f_ind(cells[0], 'Наименование параметра'),
                                       is_f_ind(cells[0], 'Единицы измерения'),
                                       is_f_ind(cells[0], 'Короткое наименование'),
                                       is_f_ind(cells[0], 'Количество знаков'),
                                       is_f_ind(cells[0], 'CPU'))

        if len(sl_CPU_one) != 0:
            tmp_line_ = is_create_objects_ai_ae_set(sl_CPU_one, tmp_object_AIAESET, 'Types.AI.AI_PLC_View')

            with open('file_out_group.txt', 'a', encoding='UTF-8') as f:
                f.write(Template(tmp_group).substitute(name_group='AI', objects=tmp_line_))

        # Расчетные
        sheet = book['Расчетные']  # .worksheets[4]
        cells = sheet['A1': 'AE' + str(sheet.max_row)]
        sl_CPU_one = is_load_ai_ae_set(i, cells, is_f_ind(cells[0], 'Алгоритмическое имя'),
                                       is_f_ind(cells[0], 'Наименование параметра'),
                                       is_f_ind(cells[0], 'Единицы измерения'),
                                       is_f_ind(cells[0], 'Короткое наименование'),
                                       is_f_ind(cells[0], 'Количество знаков'),
                                       is_f_ind(cells[0], 'CPU'))

        if len(sl_CPU_one) != 0:
            tmp_line_ = is_create_objects_ai_ae_set(sl_CPU_one, tmp_object_AIAESET, 'Types.AE.AE_PLC_View')

            with open('file_out_group.txt', 'a', encoding='UTF-8') as f:
                f.write(Template(tmp_group).substitute(name_group='AE', objects=tmp_line_))

        # Дискретные
        sheet = book['Входные']  # .worksheets[6]
        cells = sheet['A1': 'AC' + str(sheet.max_row)]
        sl_CPU_one, sl_wrn = is_load_di(i, cells, is_f_ind(cells[0], 'Алгоритмическое имя'),
                                        is_f_ind(cells[0], 'ИМ'),
                                        is_f_ind(cells[0], 'Наименование параметра'),
                                        is_f_ind(cells[0], 'Цвет при наличии'),
                                        is_f_ind(cells[0], 'Цвет при отсутствии'),
                                        is_f_ind(cells[0], 'Предупреждение'),
                                        is_f_ind(cells[0], 'Текст предупреждения'),
                                        is_f_ind(cells[0], 'CPU'))

        if len(sl_CPU_one) != 0:
            tmp_line_ = is_create_objects_di(sl_CPU_one, tmp_object_DI, 'Types.DI.DI_PLC_View')

            with open('file_out_group.txt', 'a', encoding='UTF-8') as f:
                f.write(Template(tmp_group).substitute(name_group='DI', objects=tmp_line_))

        # ИМ
        sheet = book['ИМ']  # .worksheets[9]
        cells = sheet['A1': 'T' + str(sheet.max_row)]
        sl_CPU_one = is_load_im(i, cells, is_f_ind(cells[0], 'Алгоритмическое имя'),
                                is_f_ind(cells[0], 'Наименование параметра'),
                                is_f_ind(cells[0], 'Тип ИМ'),
                                is_f_ind(cells[0], 'Род'),
                                is_f_ind(cells[0], 'Считать наработку'),
                                is_f_ind(cells[0], 'Считать перестановки'),
                                is_f_ind(cells[0], 'CPU'))
        sl_cnt = {}
        for key, value in sl_CPU_one.items():
            if value[4] == 'Да':
                sl_cnt[key+'_WorkTime'] = [value[0]]
            if value[5] == 'Да':
                sl_cnt[key+'_Swap'] = [value[0]]

        # ИМ АО- объединяем словари с ИМами
        sheet = book['ИМ(АО)']  # .worksheets[8]
        cells = sheet['A1': 'AA' + str(sheet.max_row)]
        sl_CPU_one = {**sl_CPU_one, **is_load_im_ao(i, cells, is_f_ind(cells[0], 'Алгоритмическое имя'),
                                                    is_f_ind(cells[0], 'Наименование параметра'),
                                                    is_f_ind(cells[0], 'Род'),
                                                    is_f_ind(cells[0], 'ИМ'),
                                                    is_f_ind(cells[0], 'CPU'))}

        if len(sl_CPU_one) != 0:
            tmp_line_ = is_create_objects_im(sl_CPU_one, tmp_object_IM)

            with open('file_out_group.txt', 'a', encoding='UTF-8') as f:
                f.write(Template(tmp_group).substitute(name_group='IM', objects=tmp_line_))

        # Добавляем АПР, если для данного контроллера указан АПР в настройках
        if 'АПР' in sl_CPU_spec[i]:
            with open('file_out_group.txt', 'a', encoding='UTF-8') as f:
                f.write(tmp_apr.rstrip())

        # Кнопки(в составе System)
        sheet = book['Кнопки']  # .worksheets[10]
        cells = sheet['A1': 'C' + str(sheet.max_row)]
        sl_CPU_one = is_load_btn(i, cells, is_f_ind(cells[0], 'Алгоритмическое имя'),
                                 is_f_ind(cells[0], 'Наименование параметра'),
                                 is_f_ind(cells[0], 'CPU'))

        tmp_subgroup = ''
        if len(sl_CPU_one) != 0:
            tmp_line_ = is_create_objects_btn_cnt(sl_CPU_one, tmp_object_BTN_CNT_sig, 'Types.BTN.BTN_PLC_View')
            tmp_subgroup += Template(tmp_group).substitute(name_group='BTN', objects=tmp_line_)

        # Уставки(в составе System)
        sheet = book['Уставки']  # .worksheets[5]
        cells = sheet['A1': 'AG' + str(sheet.max_row)]
        sl_CPU_one = is_load_ai_ae_set(i, cells, is_f_ind(cells[0], 'Алгоритмическое имя'),
                                       is_f_ind(cells[0], 'Наименование параметра'),
                                       is_f_ind(cells[0], 'Единицы измерения'),
                                       is_f_ind(cells[0], 'Короткое наименование'),
                                       is_f_ind(cells[0], 'Количество знаков'),
                                       is_f_ind(cells[0], 'CPU'))

        if len(sl_CPU_one) != 0:
            tmp_line_ = is_create_objects_ai_ae_set(sl_CPU_one, tmp_object_AIAESET, 'Types.SET.SET_PLC_View')
            tmp_subgroup += Template(tmp_group).substitute(name_group='SET', objects=tmp_line_)

        # Наработки(CNT- прочитали при ИМ) в составе System)
        if len(sl_cnt) != 0:
            tmp_line_ = is_create_objects_btn_cnt(sl_cnt, tmp_object_BTN_CNT_sig, 'Types.CNT.CNT_PLC_View')
            tmp_subgroup += Template(tmp_group).substitute(name_group='CNT', objects=tmp_line_)

        # Защиты(PZ) в составе System
        sheet = book['Сигналы']  # .worksheets[11]
        cells = sheet['A1': 'N' + str(sheet.max_row)]
        sl_all_pz[i] = [num_pz]
        sl_CPU_one, num_pz = is_load_pz(i, cells, num_pz,
                                        is_f_ind(cells[0], 'Наименование параметра'),
                                        is_f_ind(cells[0], 'Тип защиты'),
                                        is_f_ind(cells[0], 'Единица измерения'),
                                        is_f_ind(cells[0], 'CPU'))
        if num_pz not in sl_all_pz[i]:
            sl_all_pz[i].append(num_pz)

        if len(sl_CPU_one) != 0:
            tmp_line_ = is_create_objects_pz(sl_CPU_one, tmp_object_PZ, 'Types.PZ.PZ_PLC_View')
            tmp_subgroup += Template(tmp_group).substitute(name_group='PZ', objects=tmp_line_)

        # ПС(WRN) в составе System, каждая ПС как отдельный объект, дополняем словарь, созданный при анализе DI
        tmp_wrn, sl_ts, sl_ppu, sl_alr, sl_modes, sl_alg = is_load_sig(i, cells,
                                                                       is_f_ind(cells[0], 'Алгоритмическое имя'),
                                                                       is_f_ind(cells[0], 'Наименование параметра'),
                                                                       is_f_ind(cells[0], 'Тип защиты'),
                                                                       is_f_ind(cells[0], 'CPU'))
        sl_wrn = {**sl_wrn, **tmp_wrn}

        if len(sl_wrn) != 0:
            tmp_line_ = is_create_objects_sig(sl_wrn, tmp_object_BTN_CNT_sig)
            tmp_subgroup += Template(tmp_group).substitute(name_group='WRN', objects=tmp_line_)
            lst_all_wrn += list(sl_wrn.keys())

        # ТС в составе System, сам словарь загружен ранее вместе с ПС
        if len(sl_ts) != 0:
            tmp_line_ = is_create_objects_sig(sl_ts, tmp_object_BTN_CNT_sig)
            tmp_subgroup += Template(tmp_group).substitute(name_group='TS', objects=tmp_line_)
            lst_all_ts += list(sl_ts.keys())

        # ППУ в составе System, сам словарь загружен ранее вместе с ПС
        if len(sl_ppu) != 0:
            tmp_line_ = is_create_objects_sig(sl_ppu, tmp_object_BTN_CNT_sig)
            tmp_subgroup += Template(tmp_group).substitute(name_group='PPU', objects=tmp_line_)
            lst_all_ppu += list(sl_ppu.keys())

        # ALR(Защиты и АС) в составе System, сам словарь загружен ранее вместе с ПС
        if len(sl_alr) != 0:
            tmp_line_ = is_create_objects_sig(sl_alr, tmp_object_BTN_CNT_sig)
            tmp_subgroup += Template(tmp_group).substitute(name_group='ALR', objects=tmp_line_)

        # MODES(Режимы) в составе System, сам словарь загружен ранее вместе с ПС
        if len(sl_modes) != 0:
            tmp_line_ = is_create_objects_sig(sl_modes, tmp_object_BTN_CNT_sig)
            tmp_subgroup += Template(tmp_group).substitute(name_group='MODES', objects=tmp_line_)
            lst_all_mod += list(sl_modes.keys())

        # ALG в составе System, сам словарь загружен ранее вместе с ПС(позже поддержать новый конфигуратор)
        if len(sl_alg) != 0:
            tmp_line_ = is_create_objects_sig(sl_alg, tmp_object_BTN_CNT_sig)
            tmp_subgroup += Template(tmp_group).substitute(name_group='ALG', objects=tmp_line_)
            lst_all_alg += list(sl_alg.keys())

        # Добавляем топливный регулятор, если для данного контроллера указан ТР ПС90 в настройках
        if 'ТР' in sl_CPU_spec[i]:
            tmp_subgroup += tmp_tr_ps90.rstrip()

        # Драйвера в составе System
        sheet = book['Драйвера']
        cells = sheet['A1': 'N' + str(sheet.max_row)]
        '''
        for jj in range(len(cells[0])):
            print(jj, multiple_replace(cells[0][jj].value))  #
        '''
        sl_CPU_one = is_load_drv(controller=i, cell=cells, alg_name=is_f_ind(cells[0], 'Алгоритмическое имя'),
                                 name_par=is_f_ind(cells[0], 'Наименование параметра'),
                                 eunit=is_f_ind(cells[0], 'Единица измерения'),
                                 type_sig=is_f_ind(cells[0], 'Тип'),
                                 type_msg=is_f_ind(cells[0], 'Тип сообщения'),
                                 c_on=is_f_ind(cells[0], 'Цвет при наличии'),
                                 c_off=is_f_ind(cells[0], 'Цвет при отсутствии'),
                                 f_dig=is_f_ind(cells[0], 'Число знаков'),
                                 cpu=is_f_ind(cells[0], 'CPU'))
        if sl_CPU_one:
            sl_drv_cpu = {}  # {(алг.имя драйвера, русское наименование драйвера): [сигналы драйвера с обвесами]}
            for key in sl_CPU_one:
                if key[0] in sl_all_drv:
                    v = (key[0], sl_all_drv[key[0]])
                    if v not in sl_drv_cpu:
                        sl_drv_cpu[v] = [(key[1], *sl_CPU_one[key])]
                    else:
                        sl_drv_cpu[v].append((key[1], *sl_CPU_one[key]))
            # print(sl_drv)
            # print(sl_drv_cpu[('DRV_EF_Rtg', 'Расходомер ТГ')])
            # в общем словаре сигналов драйверов создаём пустой словарь с текущим контроллером
            sl_cpu_drv_signal[i] = {}
            for key in sl_drv_cpu:
                # для каждого драйвера создаём пустой кортеж
                tmp_sl_plus = ()
                for value in sl_drv_cpu[key]:  # пробегам по сигналам драйвера
                    # в созданный ранее кортеж добавляем алг. имя переменной
                    tmp_sl_plus += (value[0],)
                # после создания кортежа со всеми переменными драйвера
                # обновляем общий словарь с ключом текущего контроллера(i) словарём
                # {алг. имя драйвера: (кортеж его переменных)}
                sl_cpu_drv_signal[i].update({key[0]: tmp_sl_plus})

            tmp_sub_drv = ''
            for drv in sl_drv_cpu:
                tmp_line_ = is_create_objects_drv(sl_drv_cpu=sl_drv_cpu, tuple_name_drv=drv,
                                                  template_text=tmp_drv_par)
                tmp_sub_drv += Template(tmp_group).substitute(name_group=drv[0], objects=tmp_line_)

            tmp_subgroup += Template(tmp_group).substitute(name_group='DRV', objects=tmp_sub_drv.rstrip())

        # Формируем подгруппу
        if tmp_subgroup != '':
            with open('file_out_group.txt', 'a', encoding='UTF-8') as f:
                f.write(Template(tmp_group).substitute(name_group='System', objects=tmp_subgroup.rstrip()))

        # Формирование выходного файла app
        with open('file_out_group.txt', 'r', encoding='UTF-8') as f:
            tmp_line_ = f.read().rstrip()
            tmp_line_ += '\n\t<trei:unet-address-map name="UnetAddressMap" />\n'
        with open('file_app_out.txt', 'w', encoding='UTF-8') as f:
            f.write(Template(tmp_app).substitute(name_app='Tree', ct_object=tmp_line_.rstrip()))

        with open('file_app_out.txt', 'r', encoding='UTF-8') as f:
            tmp_line_ = f.read().rstrip()

        # Если нет папки File_out, то создадим её
        if not os.path.exists('File_out'):
            os.mkdir('File_out')
        # Если нет папки File_out/PLC_Aspect_importDomain, то создадим её
        if not os.path.exists(os.path.join('File_out', 'PLC_Aspect_importDomain')):
            os.mkdir(os.path.join('File_out', 'PLC_Aspect_importDomain'))
        # Для каждого объекта создаём контроллер
        num_obj_plc = 1
        for obj in sl_object_all:
            if i in sl_object_all[obj][0]:
                index_tmp = sl_object_all[obj][0].index(i)
                with open('file_plc.txt', 'a', encoding='UTF-8') as f:
                    last_dig_ip = sl_object_all[obj][1][index_tmp]
                    tmp_dig_ip1 = (last_dig_ip if '(' not in last_dig_ip else last_dig_ip[:last_dig_ip.find('(')])
                    tmp_dig_ip2 = (last_dig_ip if '(' not in last_dig_ip else last_dig_ip[last_dig_ip.find('(') + 1:-1])
                    f.write(Template(tmp_trei).substitute(plc_name='PLC_' + i + str(num_obj_plc), plc_name_type='CPU',
                                                          ip_eth1=pref_IP[0] + tmp_dig_ip1,
                                                          ip_eth2=pref_IP[1] + tmp_dig_ip2,
                                                          dp_app=tmp_line_))
                # Для каждого контроллера создадим отдельный файл для импорта только его одного
                tmp_plc = Template(tmp_trei).substitute(plc_name='PLC_' + i + str(num_obj_plc), plc_name_type='CPU',
                                                        ip_eth1=pref_IP[0] + tmp_dig_ip1,
                                                        ip_eth2=pref_IP[1] + tmp_dig_ip2,
                                                        dp_app=tmp_line_)
                # Проверка изменений, и если есть изменения, то запись
                check_diff_file(check_path=os.path.join('File_out', 'PLC_Aspect_importDomain'),
                                file_name=f'file_out_plc_{i}_{num_obj_plc}.omx-export',
                                new_data=Template(tmp_global).substitute(dp_node=tmp_plc),
                                message_print=f'Требуется заменить ПЛК-аспект {i}_{num_obj_plc}')
                # прибавляем номер объекта для формирования следующего файла
                num_obj_plc += 1
            else:
                num_obj_plc += 1
                continue

        # чистим файл групп для корректной обработки отсуствия
        file_clear = open('file_out_group.txt', 'w', encoding='UTF-8')
        file_clear.close()

    # Формирование выходного файла для каждого контроллера
    with open('file_plc.txt', 'r', encoding='UTF-8') as f:
        tmp_line_ = f.read().rstrip()

    with open('file_out_plc_aspect.omx-export', 'w', encoding='UTF-8') as f:
        f.write(Template(tmp_global).substitute(dp_node=tmp_line_))

    '''ТРЕНДЫ- JSON'''
    # Если нет папки File_out, то создадим её
    if not os.path.exists('File_out'):
        os.mkdir('File_out')
    # Если нет папки File_out/Trends, то создадим её
    if not os.path.exists(os.path.join('File_out', 'Trends')):
        os.mkdir(os.path.join('File_out', 'Trends'))

    # Считываем файл-шаблон для сигнала тренда
    with open(os.path.join('Template', 'Temp_signal_trends'), 'r', encoding='UTF-8') as f_trends:
        tmp_signal_trends = f_trends.read()

    # Определение объявленных мнемосхем с листа настроек
    sheet = book['Настройки']
    cells = sheet['A1': 'A' + str(sheet.max_row)]
    tuple_node_trends = tuple()
    for p in cells:
        if p[0].value == 'Наименование мнемосхемы':
            jj = 1
            while sheet[p[0].row][jj].value:
                tuple_node_trends += (sheet[p[0].row][jj].value,)
                jj += 1
    # print(tuple_node_trends)
    # Словарь соответствия листов конфигуратора и имени группы на трендах
    sl_group_trends = {
        'Измеряемые': ('Аналоговые входные', 'AI'),
        'Расчетные': ('Расчетные параметры', 'AE'),
        'Входные': ('Дискретные входные', 'DI'),
        'ИМ': ('Исполнительные механизмы', 'IM'),
        'ИМ(АО)': ('Исполнительные механизмы', 'IM'),
        'Драйвера': ('Сигналы от драйверов', 'System.DRV'),
        'Сигналы': ()
    }
    sl_brk_group_trends = {
        'Измеряемые': ('Отказы аналоговых входных', 'AI'),
        'Расчетные': ('Отказы расчетных параметров', 'AE'),
        'Входные': ('Отказы дискретных входных', 'DI')
    }
    sl_state_im_gender = {
        'Включить': {'М': 'Включен', 'Ж': 'Включена', 'С': 'Включено'},
        'Открыть': {'М': 'Открыт', 'Ж': 'Открыта', 'С': 'Открыто'},
        'Включить_off': {'М': 'Отключен', 'Ж': 'Отключена', 'С': 'Отключено'},
        'Открыть_off': {'М': 'Закрыт', 'Ж': 'Закрыта', 'С': 'Закрыто'}
    }
    sl_antonym = {
        'Открыть': 'Закрыть',
        'Включить': 'Отключить'
    }

    # Для каждого объекта, прочитанного ранее
    for obj in sl_object_all:
        lst_json = []
        # Для каждого листа конфигуратора в словаре групп для трендов(ключи словаря опрделили
        # по объявленным мнемосхемам)
        for list_config in sl_group_trends:
            # для каждой группы создаём словарь с пустыми словарями для каждого узла
            sl_node_trends = {node: {} for node in tuple_node_trends}
            # Создаём словарь возможных типов защит - используем при парсинге листа Сигналы и сборке аварий
            sl_node_alr = {}
            sl_node_modes = {}  # Словарь возможных режимов
            sl_node_drv = {}  # Словарь возможных драйверов
            # Выбираем  текущий лист в качестве активного
            sheet = book[list_config]
            # Устанавливаем Диапазон считывания для первой строки (узнать индексы столбцов)
            cells_name = sheet['A1': 'AG1']
            rus_par_ind = is_f_ind(cells_name[0], 'Наименование параметра')
            alg_name_ind = is_f_ind(cells_name[0], 'Алгоритмическое имя')
            eunit_ind = is_f_ind(cells_name[0], 'Единицы измерения')
            node_name_ind = is_f_ind(cells_name[0], 'Узел')
            eunit_drv_ind = is_f_ind(cells_name[0], 'Единица измерения')
            t_sig_drv_ind = is_f_ind(cells_name[0], 'Тип')

            # Устанавливаем диапазон для чтения параметров
            cells_read = sheet['A2': 'AG' + str(sheet.max_row)]
            # пробегаемся по параметрам листа
            for par in cells_read:
                # Если конец строки, то заканчиваем обработку ячеек
                if par[rus_par_ind].value is None:
                    break
                # при условии, что находимся на листе 'Измеряемые', 'Расчетные' или на Входные и сигнал не привязан к ИМ
                if list_config in ('Измеряемые', 'Расчетные') or \
                        list_config == 'Входные' and par[is_f_ind(cells_name[0], 'ИМ')].value == 'Нет':
                    # создаём промежуточный словарь {рус.имя: (алг.имя, единицы измерения)}
                    sl_par_trends = {f_ind_json(par[rus_par_ind].value): (par[alg_name_ind].value.replace('|', '_') +
                                                                          '.Value', par[eunit_ind].value)}
                    # добавляем словарь параметра в словарь узла
                    sl_node_trends[par[node_name_ind].value].update(sl_par_trends)
                # при условии, что парсим лист ИМов
                elif list_config in ('ИМ',):
                    # если 'ИМ1Х0', 'ИМ1Х0и'
                    if par[is_f_ind(cells_name[0], 'Тип ИМ')].value in ('ИМ1Х0', 'ИМ1Х0и'):
                        move_ = par[is_f_ind(cells_name[0], 'Вкл./откр.')].value  # определяем тип открытия
                        # создаём промежуточный словарь {рус.имя: (алг.имя, единицы измерения)}
                        sl_par_trends = \
                            {f'{f_ind_json(par[rus_par_ind].value)}. {move_}': (par[alg_name_ind].value + '.oOn', '-')}
                        # добавляем словарь параметра в словарь узла
                        sl_node_trends[par[node_name_ind].value].update(sl_par_trends)

                    # если 'ИМ1Х1', 'ИМ1Х1и'
                    elif par[is_f_ind(cells_name[0], 'Тип ИМ')].value in ('ИМ1Х1', 'ИМ1Х1и'):
                        move_ = par[is_f_ind(cells_name[0], 'Вкл./откр.')].value  # определяем тип открытия
                        gender_ = par[is_f_ind(cells_name[0], 'Род')].value  # определяем род ИМ
                        # создаём промежуточный словарь с возможными префиксами сигналов
                        sl_tmp = {'.oOn': move_, '.stOn': sl_state_im_gender[move_][gender_]}
                        for par_pref in sl_tmp:
                            # создаём промежуточный словарь {рус.имя: (алг.имя, единицы измерения)}
                            sl_par_trends = \
                                {f"{f_ind_json(par[rus_par_ind].value)}. {sl_tmp[par_pref]}": (par[alg_name_ind].value +
                                                                                               par_pref, '-')}
                            # добавляем словарь параметра в словарь узла
                            sl_node_trends[par[node_name_ind].value].update(sl_par_trends)

                    # если 'ИМ1Х2', 'ИМ1Х2и'
                    elif par[is_f_ind(cells_name[0], 'Тип ИМ')].value in ('ИМ1Х2', 'ИМ1Х2и'):
                        move_ = par[is_f_ind(cells_name[0], 'Вкл./откр.')].value  # определяем тип открытия
                        gender_ = par[is_f_ind(cells_name[0], 'Род')].value  # определяем род ИМ
                        # создаём промежуточный словарь с возможными префиксами сигналов
                        sl_tmp = {'.oOn': move_, '.stOn': sl_state_im_gender[move_][gender_],
                                  '.stOff': sl_state_im_gender[move_ + '_off'][gender_]}
                        for par_pref in sl_tmp:
                            # создаём промежуточный словарь {рус.имя: (алг.имя, единицы измерения)}
                            sl_par_trends = \
                                {f'{f_ind_json(par[rus_par_ind].value)}. {sl_tmp[par_pref]}': (par[alg_name_ind].value +
                                                                                               par_pref, '-')}
                            # добавляем словарь параметра в словарь узла
                            sl_node_trends[par[node_name_ind].value].update(sl_par_trends)

                    # если 'ИМ2Х2', 'ИМ2Х2с', 'ИМ2Х4'
                    elif par[is_f_ind(cells_name[0], 'Тип ИМ')].value in ('ИМ2Х2', 'ИМ2Х2с', 'ИМ2Х4'):
                        move_ = par[is_f_ind(cells_name[0], 'Вкл./откр.')].value  # определяем тип открытия
                        gender_ = par[is_f_ind(cells_name[0], 'Род')].value  # определяем род ИМ
                        # создаём промежуточный словарь с возможными префиксами сигналов
                        sl_tmp = {'.oOn': move_, '.oOff': sl_antonym[move_],
                                  '.stOn': sl_state_im_gender[move_][gender_],
                                  '.stOff': sl_state_im_gender[move_ + '_off'][gender_]}
                        for par_pref in sl_tmp:
                            # создаём промежуточный словарь {рус.имя: (алг.имя, единицы измерения)}
                            sl_par_trends = \
                                {f'{f_ind_json(par[rus_par_ind].value)}. {sl_tmp[par_pref]}': (par[alg_name_ind].value +
                                                                                               par_pref, '-')}
                            # добавляем словарь параметра в словарь узла
                            sl_node_trends[par[node_name_ind].value].update(sl_par_trends)
                # при условии, что парсим лист ИМ(АО) и выделена как ИМ
                elif list_config in ('ИМ(АО)',) and par[is_f_ind(cells_name[0], 'ИМ')].value == 'Да':
                    # sl_tmp  промежуточный словарь с возможными префиксами сигналов
                    sl_tmp = {'.Set': 'Задание', '.iPos': 'Положение'}
                    for par_pref in sl_tmp:
                        # создаём промежуточный словарь {рус.имя: (алг.имя, единицы измерения)}
                        sl_par_trends = \
                            {f'{f_ind_json(par[rus_par_ind].value)}. {sl_tmp[par_pref]}': (par[alg_name_ind].value +
                                                                                           par_pref,
                                                                                           par[eunit_ind].value)}
                        # добавляем словарь параметра в словарь узла
                        sl_node_trends[par[node_name_ind].value].update(sl_par_trends)

                # при условии, что парсим лист Сигналы
                elif list_config in ('Сигналы', ):
                    type_prot = par[is_f_ind(cells_name[0], 'Тип защиты')].value
                    # Если увидели аварию
                    if type_prot in 'АОссАОбсВОссВОбсАОНО':
                        # создаём промежуточный словарь аварии {рус.имя: (алг.имя, единицы измерения - '-')}
                        sl_alr_trends = \
                            {f'{f_ind_json(par[rus_par_ind].value)}': (par[alg_name_ind].value.replace('|', '.') +
                                                                       '.Value', '-')}
                        # если в словаре аварий отсутвует такая авария, то создаём
                        if type_prot not in sl_node_alr:
                            sl_node_alr[type_prot] = sl_alr_trends
                        else:  # иначе обновляем словарь, который есть
                            sl_node_alr[type_prot].update(sl_alr_trends)
                    # Если увидели режим
                    elif type_prot in ('Режим',):
                        # то добавляем в словарь режимов с ключом рус имени аварии : (алг.имя, единицы измерения - '-')
                        sl_node_modes[f_ind_json(par[rus_par_ind].value)] = \
                            (par[alg_name_ind].value.replace('MOD|', '') + '.Value', '-')
                # при условии, что парсим лист Драйверов
                elif list_config in ('Драйвера',):
                    sl_type_unit = {'BOOL': '-', 'INT': par[eunit_drv_ind].value, 'FLOAT': par[eunit_drv_ind].value,
                                    'IEC': '-', 'Daily': '-'}
                    drv_ = par[is_f_ind(cells_name[0], 'Драйвер')].value
                    # создаём промежуточный словарь сигнала драйвера {рус.имя: (алг.имя, единицы измерения - '-')}
                    sl_drv_trends = \
                        {f'{f_ind_json(par[rus_par_ind].value)}': (par[alg_name_ind].value + '.Value',
                                                                   sl_type_unit[par[t_sig_drv_ind].value])}
                    # если в словаре драйверов отсутвует такой драйвер, то создаём
                    if drv_ not in sl_node_drv:
                        sl_node_drv[drv_] = sl_drv_trends
                    else:  # иначе обновляем словарь, который есть
                        sl_node_drv[drv_].update(sl_drv_trends)
            # на этапе парсинга листа ИМАО добавляем в тренды АПР
            if list_config == 'ИМ(АО)':
                if 'АПР' in [item for sublist in [i for i in sl_CPU_spec.values() if i] for item in sublist]:
                    sl_tmp = {'Set': 'Задание', 'Pos': 'Положение'}
                    for par_pref in sl_tmp:
                        lst_json.append(
                            {"Signal": {"UserTree": f'Исполнительные механизмы/Главная/АПРК. {sl_tmp[par_pref]}',
                                        "OpcTag": f'{obj}.APR.IM.{par_pref}',
                                        "EUnit": '%',
                                        "Description": f'АПРК. {sl_tmp[par_pref]}'}})
            # для каждого узла(мнемосхемы)...
            for node in sl_node_trends:
                # ...для каждого параметра по отсортированному словарю параметров в узле...
                for param in sorted(sl_node_trends[node]):
                    # ...собираем json
                    # print(node, param, sl_node_trends[node][param])
                    lst_json.append(
                        {"Signal": {"UserTree": f'{sl_group_trends[list_config][0]}/{node}/{param}',
                                    "OpcTag":
                                        f'{obj}.{sl_group_trends[list_config][1]}.{sl_node_trends[node][param][0]}',
                                    "EUnit": sl_node_trends[node][param][1],
                                    "Description": param}})

            # для каждого узла-драйвера в отсортированном словаре драйверов...
            for drv in sorted(sl_node_drv):
                # ... для каждого сигнала драйвера по отсортированному словарю сигналов в узле-драйвере
                for sig_drv in sorted(sl_node_drv[drv]):
                    # ...собираем json
                    # дополнительная проверка наличия драйвера в объявленных
                    if drv in sl_all_drv:
                        lst_json.append(
                            {"Signal": {"UserTree": f"{sl_group_trends['Драйвера'][0]}/{sl_all_drv[drv]}/{sig_drv}",
                                        "OpcTag":
                                            f"{obj}.{sl_group_trends['Драйвера'][1]}.{sl_node_drv[drv][sig_drv][0]}",
                                        "EUnit": sl_node_drv[drv][sig_drv][1],
                                        "Description": sig_drv}})

            # для каждого типа аварии в остортированном словаре типов аварий...
            for node in sorted(sl_node_alr):
                # ...для каждой аварии по отсортированному словарю аварий в узле(типе)...
                for alr in sorted(sl_node_alr[node]):
                    # ...собираем json
                    lst_json.append(
                        {"Signal": {"UserTree": f"Аварии/{node}/{node}. {alr}",
                                    "OpcTag":
                                        f"{obj}.System.{sl_node_alr[node][alr][0]}",
                                    "EUnit": sl_node_alr[node][alr][1],
                                    "Description": f'{node}. {alr}'}})
            # для каждого режима в отсортированном словаре режимов
            for mode in sorted(sl_node_modes):
                # ...собираем json
                lst_json.append(
                    {"Signal": {"UserTree": f"Режимы/Режим {mode}",
                                "OpcTag":
                                    f"{obj}.System.MODES.{sl_node_modes[mode][0]}",
                                "EUnit": sl_node_modes[mode][1],
                                "Description": mode}})

        # Дополнительный перебор для сбора отказов
        for list_config in sl_brk_group_trends:
            # для каждой группы создаём словарь с пустыми словарями для каждого узла
            sl_node_trends = {node: {} for node in tuple_node_trends}
            # Выбираем  текущий лист в качестве активного
            sheet = book[list_config]
            # Устанавливаем Диапазон считывания для первой строки (узнать индексы столбцов)
            cells_name = sheet['A1': 'AG1']
            rus_par_ind = is_f_ind(cells_name[0], 'Наименование параметра')
            alg_name_ind = is_f_ind(cells_name[0], 'Алгоритмическое имя')
            eunit_ind = is_f_ind(cells_name[0], 'Единицы измерения')
            node_name_ind = is_f_ind(cells_name[0], 'Узел')

            # Устанавливаем диапазон для чтения параметров
            cells_read = sheet['A2': 'AG' + str(sheet.max_row)]
            # пробегаемся по параметрам листа
            for par in cells_read:
                # Если конец строки, то заканчиваем обработку ячеек
                if par[rus_par_ind].value is None:
                    break
                # при условии, что находимся на листе 'Измеряемые', 'Расчетные' или на Входные и сигнал не привязан к ИМ
                if list_config in ('Измеряемые', 'Расчетные') or \
                        list_config == 'Входные' and par[is_f_ind(cells_name[0], 'ИМ')].value == 'Нет':
                    # создаём промежуточный словарь {рус.имя: (алг.имя, единицы измерения)}
                    sl_par_trends = \
                        {f_ind_json(par[rus_par_ind].value): (par[alg_name_ind].value.replace('|', '_') + '.fValue',
                                                              par[eunit_ind].value)}
                    # добавляем словарь параметра в словарь узла
                    sl_node_trends[par[node_name_ind].value].update(sl_par_trends)

            # для каждого узла(мнемосхемы)...
            for node in sl_node_trends:
                # ...для каждого параметра по отсортированному словарю параметров в узле...
                for param in sorted(sl_node_trends[node]):
                    # ...собираем json
                    lst_json.append(
                        {"Signal": {"UserTree": f'{sl_brk_group_trends[list_config][0]}/{node}/Отказ - {param}',
                                    "OpcTag":
                                        f'{obj}.{sl_brk_group_trends[list_config][1]}.{sl_node_trends[node][param][0]}',
                                    "EUnit": sl_node_trends[node][param][1],
                                    "Description": f'Отказ - {param}'}})

        # Проверка изменений, и если есть изменения, то запись файла json
        check_diff_file(check_path=os.path.join('File_out', 'Trends'),
                        file_name=f'Tree{obj}.json',
                        new_data=json_dumps({"UserTree": lst_json}, indent=1, ensure_ascii=False),
                        message_print=f'Требуется заменить файл Tree{obj}.json')

    book.close()

    os.remove('file_plc.txt')
    os.remove('file_out_group.txt')
    # os.remove('file_out_objects.txt')
    os.remove('file_app_out.txt')
    print(datetime.datetime.now(), 'Окончание сборки фалов ПЛК-Аспектов и файлов трендов')

    path_nku = ''
    # Если есть файл-источник конфигуратора НКУ
    if os.path.exists('Source_NKU.txt'):
        with open('Source_NKU.txt', 'r', encoding='UTF-8') as f_nku:
            path_nku = f_nku.readline().strip()
    # Если считан путь к конфигуратору НКУ
    if path_nku:
        # Даём по умолчанию название конфигуратору
        config_nku = 'UnimodCreate v.4.4.xlsm'
        # Ищем файл конфигуратора NKU в указанном каталоге
        for file in os.listdir(path_nku):
            if file.endswith('.xlsm') or file.endswith('.xls'):
                config_nku = file
                break
        # Считываем файл-шаблон для DI_NKU
        with open(os.path.join('Template', 'Temp_DI_NKU'), 'r', encoding='UTF-8') as f:
            tmp_object_DI_NKU = f.read()

        book_nku = openpyxl.open(os.path.join(path_nku, config_nku))  # , read_only=True
        # Дискретные НКУ
        sheet_nku = book_nku['Входные']  # .worksheets[6]
        cells_npu = sheet_nku['A1': 'AC' + str(sheet_nku.max_row)]

        sl_CPU_nku, sl_wrn_nku = is_load_di_nku('GPA', cells_npu, is_f_ind(cells_npu[0], 'Алгоритмическое имя'),
                                                is_f_ind(cells_npu[0], 'ИМ'),
                                                is_f_ind(cells_npu[0], 'Наименование параметра'),
                                                is_f_ind(cells_npu[0], 'Цвет при наличии'),
                                                is_f_ind(cells_npu[0], 'Цвет при отсутствии'),
                                                is_f_ind(cells_npu[0], 'Предупреждение'),
                                                is_f_ind(cells_npu[0], 'Текст предупреждения'),
                                                is_f_ind(cells_npu[0], 'CPU'))
        book_nku.close()

        # Если словарь НКУ сигналов не пустой
        if sl_CPU_nku:
            tmp_line_ = is_create_objects_di_nku(sl_CPU_nku, tmp_object_DI_NKU, 'Types.DI_NKU.DI_NKU_PLC_View',
                                                 sl_wrn_nku)
            tmp_group_nku = Template(tmp_group).substitute(name_group='NKU', objects=tmp_line_)

            check_diff_file(check_path=os.path.join('File_out', 'PLC_Aspect_importDomain'),
                            file_name=f'file_out_plc_NKU_inGPA.omx-export',
                            new_data=Template(tmp_global).substitute(dp_node=tmp_group_nku),
                            message_print=f'Требуется заменить Сигналы НКУ в аспекте ПЛК')
            if os.path.exists('Source_list_plc.txt'):
                create_index_nku(name_plc_nku='GPA', sl_signal_nku=sl_CPU_nku)
    if os.path.exists('Source_list_plc.txt'):
        create_index(lst_alg=lst_all_alg, lst_mod=lst_all_mod, lst_ppu=lst_all_ppu, lst_ts=lst_all_ts,
                     lst_wrn=lst_all_wrn, sl_pz_anum=sl_all_pz, sl_cpu_spec=sl_CPU_spec, sl_diag=sl_for_diag,
                     sl_cpu_drv_signal=sl_cpu_drv_signal)
    print(datetime.datetime.now(), 'Окончание сборки карт индексов')
    # добавление отсечки в файл изменений, чтобы разные сборки не сливались
    if os.path.exists('Required_change.txt'):
        with open('Required_change.txt', 'r') as f_test:
            check_test = f_test.readlines()[-1]
        if check_test != '-' * 70 + '\n':
            with open('Required_change.txt', 'a') as f_test:
                f_test.write('-' * 70 + '\n')
    input(f'{datetime.datetime.now()} - Сборка файлов завершена успешно. Нажмите Enter для выхода...')

except (Exception, KeyError):
    # в случае возникновения какой-либо ошибки, чистим возможные промежуточные файлы
    for file_error_clear in ('file_plc.txt', 'file_out_group.txt', 'file_app_out.txt', 'file_app_out.txt'):
        if os.path.exists(file_error_clear):
            os.remove(file_error_clear)

    logging.basicConfig(filename='error.log', filemode='a', datefmt='%d.%m.%y %H:%M:%S',
                        format='%(levelname)s - %(message)s - %(asctime)s')
    logging.exception("Ошибка выполнения")
    # print('Произошла ошибка выполнения')
    input('Во время сборки произошла ошибка сформирован файл error.log. Нажмите Enter для выхода...')
