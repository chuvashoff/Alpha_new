# from my_func import is_f_ind
from string import Template
from copy import copy
import os
import re
from algroritm import is_load_algoritm
import datetime
from time import sleep
# для тестирования новых возможностей
import xml.etree.ElementTree as ET
import lxml.etree
from openpyxl.utils import get_column_letter


def multiple_replace_xml(target_str):
    replace_values = {'address_map': 'address-map', 'access_level': 'access-level', 'ct_object': 'ct:object',
                      'base_type': 'base-type', 'xmlns_dp': 'xmlns:dp', 'xmlns_trei': 'xmlns:trei',
                      'xmlns_ct': 'xmlns:ct', 'trei_': 'trei:', 'dp_application': 'dp:application',
                      'ct_init-ref': 'ct:init-ref', 'ct_parameter': 'ct:parameter', 'ct_subject-ref': 'ct:subject-ref',
                      'const_access': 'const-access', 'ct_bind': 'ct:bind', 'xmlns_r': 'xmlns:r',
                      'xmlns_snmp': 'xmlns:snmp', 'dp_computer': 'dp:computer',
                      'dp_ethernet-adapter': 'dp:ethernet-adapter', 'dp_external-runtime': 'dp:external-runtime',
                      'snmp_': 'snmp:', 'poll_port': 'poll-port', 'poll_password': 'poll-password',
                      'notification_port': 'notification-port', 'notification_password': 'notification-password',
                      'protocol_version': 'protocol-version', 'security_level': 'security-level',
                      'auth_protocol': 'auth-protocol', 'priv_protocol': 'priv-protocol',
                      'source_code': 'source-code', 'ct_trigger': 'ct:trigger', 'ct_handler': 'ct:handler',
                      'ct_formula': 'ct:formula', 'format_version': 'format-version', 'ct_timer': 'ct:timer'}
    # получаем заменяемое: подставляемое из словаря в цикле
    for i, j in replace_values.items():
        # меняем все target_str на подставляемое
        target_str = target_str.replace(i, j)
    return target_str


def add_xml_par_plc(name_group, sl_par, parent_node, sl_attr_par):
    # sl_par = {алг_пар: (тип параметра в студии, русское имя, ед. изм., короткое имя, количество знаков)}
    # sl_attr_par - словарь атрибутов параметра {алг_пар: {тип атрибута: значение атрибута}}
    child_group = ET.SubElement(parent_node, 'ct_object',
                                name=(f'{name_group[0]}' if isinstance(name_group, tuple) else f"{name_group}"),
                                access_level="public")
    if isinstance(name_group, tuple):
        ET.SubElement(child_group, 'attribute', type='unit.System.Attributes.Description', value=f'{name_group[1]}')
    # Для каждого параметра в текущем контроллере
    # создаём конструкцию параметра
    for par, tuple_par in sl_par.items():
        child_par = ET.SubElement(child_group, 'ct_object', name=f"{par}",
                                  base_type=f"Types.{tuple_par[0]}",
                                  aspect="Types.PLC_Aspect", access_level="public")
        # Если 'AI', 'AE', 'SET', то добавляем Value с гистерезисом
        if name_group in ('AI', 'AE', 'SET'):
            deadband_hist = 1 / (10 ** int(tuple_par[5])) if tuple_par[5] != '' else 0.01
            ET.SubElement(child_par, 'attribute', type=f"Attributes.Enable_History", value=f"True")
            ET.SubElement(child_par, 'attribute', type=f"Attributes.DeadBand_History", value=f"{deadband_hist}")
        # Если ИМ АО, то добавляем Set и iPos с гистерезисом
        elif name_group in ('IM',) and 'IM_AO' in tuple_par[0]:
            deadband_hist = 1 / (10 ** int(tuple_par[6])) if tuple_par[6] != '' else 0.01
            ET.SubElement(child_par, 'attribute', type=f"Attributes.Enable_History", value=f"True")
            ET.SubElement(child_par, 'attribute', type=f"Attributes.DeadBand_History", value=f"{deadband_hist}")
        # Если драйверный параметр, AI и сохраняем историю
        elif 'DRV' in tuple(parent_node.attrib.values()):
            if 'Сохраняемый - Да' in tuple_par:
                if 'AI' in tuple_par[0]:
                    deadband_hist = 1 / (10 ** int(tuple_par[8])) if tuple_par[8] != '' else 0.01
                    ET.SubElement(child_par, 'attribute', type=f"Attributes.Enable_History", value=f"True")
                    ET.SubElement(child_par, 'attribute', type=f"Attributes.DeadBand_History", value=f"{deadband_hist}")
                else:
                    ET.SubElement(child_par, 'attribute', type=f"Attributes.Enable_History", value=f"True")
            else:
                # Строчки по deadband при отсутствии сохранения добавил в отпуске, нужно собрать сборку
                deadband_hist = 1
                ET.SubElement(child_par, 'attribute', type=f"Attributes.Enable_History", value=f"False")
                ET.SubElement(child_par, 'attribute', type=f"Attributes.DeadBand_History", value=f"{deadband_hist}")
        # for type_attr, value in sl_attr_par[par].items():
        for type_attr, value in sl_attr_par.get(par, {}).items():
            ET.SubElement(child_par, 'attribute', type=f"{type_attr}", value=f"{value}")
    return


def add_xml_par_ios(set_cpu_object, objects, name_group, sl_par, parent_node, plc_node_tree, sl_agreg):
    # ...то создаём узел AI в IOS-аспекте
    child_group = ET.SubElement(parent_node, 'ct_object',
                                name=f'{name_group[0]}' if isinstance(name_group, tuple) else f"{name_group}",
                                access_level="public")
    # ...добавляем агрегаторы
    if isinstance(name_group, tuple):
        ET.SubElement(child_group, 'attribute', type='unit.System.Attributes.Description', value=f'{name_group[1]}')
    for agreg, type_agreg in sl_agreg.items():
        ET.SubElement(child_group, 'ct_object', name=f'{agreg}', base_type=f"{type_agreg}",
                      aspect="Types.IOS_Aspect", access_level="public")
    # Для каждого контроллера в структуре анпаров...
    for cpu in sl_par:
        # Для каждого параметра...
        # Если перебираемый контроллер принадлежит объекту
        if cpu in set_cpu_object:
            # ...для каждого параметра контроллера...
            for par, tuple_par in sl_par[cpu].items():
                # ...создаём структуру параметра
                # base_type = (f"Types.{tuple_par[0].replace('PLC_View', 'IOS_NotHistory_View')}"  #
                #              if 'Сохраняемый - Нет' in tuple_par
                #              else f"Types.{tuple_par[0].replace('PLC_View', 'IOS_View')}")
                # aspect = ("Types.IOS_NotHistory_Aspect" if 'Сохраняемый - Нет' in tuple_par else "Types.IOS_Aspect")
                child_par = ET.SubElement(child_group, 'ct_object', name=f"{par}",
                                          base_type=f"Types.{tuple_par[0].replace('PLC_View', 'IOS_View')}",
                                          aspect="Types.IOS_Aspect",
                                          original=f"PLC_{cpu}_{objects[2]}.CPU.Tree.{plc_node_tree}.{par}",
                                          access_level="public")
                ET.SubElement(child_par, 'ct_init-ref',
                              ref="_PLC_View", target=f"PLC_{cpu}_{objects[2]}.CPU.Tree.{plc_node_tree}.{par}")
                # # Если 'AI', 'AE', 'SET', то добавляем Value с гистерезисом
                # if name_group in ('AI', 'AE', 'SET'):
                #     deadband_hist = 1/(10**int(tuple_par[5])) if tuple_par[5] != '' else 0.01
                #     object_value = ET.SubElement(child_par, 'ct_parameter', name='Value', type='float32',
                #                                  direction='out', access_level="public")
                #     ET.SubElement(object_value, 'attribute', type="unit.Server.Attributes.History",
                #                   value=f"Enable=\"True\" Deadband=\"{deadband_hist}\" ServerTime=\"False\"")
                #     ET.SubElement(object_value, 'attribute', type="unit.System.Attributes.Description",
                #                   value=f"@(object:unit.System.Attributes.Description)")
                #     ET.SubElement(object_value, 'attribute', type="unit.Server.Attributes.Unit",
                #                   value=f"@(object:Attributes.EUnit)")
                #     ET.SubElement(child_par, 'ct_formula', source_code="_PLC_View.States.Value", target="Value")
                # # Если ИМ АО, то добавляем Set и iPos с гистерезисом
                # elif name_group in ('IM',) and 'IM_AO' in tuple_par[0]:
                #     deadband_hist = 1/(10**int(tuple_par[6])) if tuple_par[6] != '' else 0.01
                #     for sig, attr_sig in {'Set': ('Задание',), 'iPos': ('Положение',)}.items():
                #         object_sig = ET.SubElement(child_par, 'ct_parameter', name=f'{sig}', type='float32',
                #                                    direction='out', access_level="public")
                #         ET.SubElement(object_sig, 'attribute', type="unit.System.Attributes.Description",
                #                       value=f"{attr_sig[0]}")
                #         ET.SubElement(object_sig, 'attribute', type="unit.Server.Attributes.Unit",
                #                       value=f"@(object:Attributes.EUnit)")
                #         ET.SubElement(object_sig, 'attribute', type="unit.Server.Attributes.History",
                #                       value=f"Enable=\"True\" Deadband=\"{deadband_hist}\" ServerTime=\"False\"")
                #         ET.SubElement(child_par, 'ct_formula', source_code=f"_PLC_View.States.{sig}", target=f"{sig}")
                # # Если драйверный параметр, AI и сохраняем историю
                # elif 'System.DRV' in plc_node_tree and 'DRV_AI_IOS_View' in base_type:
                #     deadband_hist = 1/(10**int(tuple_par[8])) if tuple_par[8] != '' else 0.01
                #     object_value = ET.SubElement(child_par, 'ct_parameter', name='Value', type='float32',
                #                                  direction='out', access_level="public")
                #     ET.SubElement(object_value, 'attribute', type="unit.Server.Attributes.History",
                #                   value=f"Enable=\"True\" Deadband=\"{deadband_hist}\" ServerTime=\"False\"")
                #     ET.SubElement(object_value, 'attribute', type="unit.System.Attributes.Description",
                #                   value=f"@(object:unit.System.Attributes.Description)")
                #     ET.SubElement(object_value, 'attribute', type="unit.Server.Attributes.Unit",
                #                   value=f"@(object:Attributes.EUnit)")
                #     ET.SubElement(child_par, 'ct_formula', source_code="_PLC_View.States.Value", target="Value")

    return


def is_read_ai_ae_set(sheet, type_signal):
    # return_sl = {cpu: {алг_пар: (тип параметра в студии, русское имя, ед. изм., короткое имя, количество знаков,
    # количество знаков для истории)}}
    return_sl = {}
    return_sl_diff = {}
    sl_cpu_fast = {}
    # return_sl_sday = {cpu: {Префикс папки.alg_par: русское имя}}
    return_sl_sday = {}
    # return_sl_mnemo = {cpu: {узел: список параметров узла}}
    return_sl_mnemo = {}
    sl_plc_aspect = {'AI': 'AI.AI_PLC_View', 'AE': 'AE.AE_PLC_View', 'SET': 'SET.SET_PLC_View'}
    # print('Зашли в функцию, максимальное число колонок - ', sheet.max_column)
    cells = sheet['A1': get_column_letter(100) + str(100)]  # sheet.max_column
    index_alg_name = is_f_ind(cells[0], 'Алгоритмическое имя')
    index_rus_name = is_f_ind(cells[0], 'Наименование параметра')
    index_unit = is_f_ind(cells[0], 'Единицы измерения')
    index_short_name = is_f_ind(cells[0], 'Короткое наименование')
    index_frag_dig = is_f_ind(cells[0], 'Количество знаков')
    index_cpu_name = is_f_ind(cells[0], 'CPU')
    index_res = is_f_ind(cells[0], 'Резервный')
    index_node_mnemo = is_f_ind(cells[0], 'Узел')
    index_sday = is_f_ind(cells[0], 'Используется в ведомости')
    index_fast = is_f_ind(cells[0], 'Передача по МЭК')

    cells = sheet['A2': get_column_letter(100) + str(sheet.max_row)]
    # Составляем множество контроллеров, у которых есть данные параметры
    # set_par_cpu = set()
    for par in cells:
        if par[0].value is None:
            break
        # set_par_cpu.add(par[index_cpu_name].value)
        # Проверка комментария
        # if par[index_cpu_name].comment:
        #     print(str(par[index_cpu_name].comment).split()[1].split(';'))
        if par[index_res].value == 'Нет':
            if par[index_cpu_name].value not in return_sl:
                # return_sl = {cpu: {алг_пар: (русское имя, ед изм., короткое имя, количество знаков,
                # количество знаков для истории)}}
                return_sl[par[index_cpu_name].value] = {}

            # Обрабатываем и составляем словарь отличий между параметрами по объектам
            if par[index_cpu_name].value not in return_sl_diff:
                # return_sl = {cpu: {алг_пар: (русское имя, ед изм., короткое имя, количество знаков,
                # количество знаков для истории)}}
                return_sl_diff[par[index_cpu_name].value] = {}
            return_sl_diff[par[index_cpu_name].value].update({
                par[index_alg_name].value.replace('|', '_'):
                    str(par[index_cpu_name].comment).split()[1].split(';') if par[index_cpu_name].comment else ''})

            if par[index_sday].value == 'Да':
                if par[index_cpu_name].value not in return_sl_sday:
                    return_sl_sday[par[index_cpu_name].value] = {}
                return_sl_sday[par[index_cpu_name].value].update(
                    {f"{type_signal}.{par[index_alg_name].value.replace('|', '_')}": par[index_rus_name].value})

            frag_dig_hist = ''.join([i for i
                                     in str(par[index_frag_dig].comment)[:str(par[index_frag_dig].comment).find('by')]
                                     if i.isdigit()]) if par[index_frag_dig].comment else par[index_frag_dig].value

            return_sl[par[index_cpu_name].value].update({par[index_alg_name].value.replace('|', '_'): (
                sl_plc_aspect.get(type_signal, 'Пустой тип'),
                par[index_rus_name].value,
                par[index_unit].value,
                par[index_short_name].value,
                par[index_frag_dig].value,
                frag_dig_hist
            )})
            # Если не парсим уставки, то заполняем словарь для мнемосхемы
            if 'SP|' not in par[index_alg_name].value and par[index_cpu_name].value not in return_sl_mnemo:
                return_sl_mnemo[par[index_cpu_name].value] = {}
            if 'SP|' not in par[index_alg_name].value and \
                    par[index_node_mnemo].value not in return_sl_mnemo[par[index_cpu_name].value]:
                return_sl_mnemo[par[index_cpu_name].value][par[index_node_mnemo].value] = list()
            if 'SP|' not in par[index_alg_name].value:
                return_sl_mnemo[par[index_cpu_name].value][par[index_node_mnemo].value].append(
                    par[index_alg_name].value.replace('|', '_')
                )
            if par[index_fast].value == 'Да':
                if par[index_cpu_name].value not in sl_cpu_fast:
                    sl_cpu_fast[par[index_cpu_name].value] = (par[index_alg_name].value,)
                else:
                    sl_cpu_fast[par[index_cpu_name].value] += (par[index_alg_name].value,)

    return return_sl, return_sl_mnemo, return_sl_sday, return_sl_diff, sl_cpu_fast


def is_read_di(sheet):
    # return_sl_di = {cpu: {алг_пар: (тип параметра в студии, русское имя, sColorOff, sColorOn)}}
    return_sl_di = {}
    return_sl_diff = {}
    # return_sl_mnemo = {узел: список параметров узла}
    return_sl_mnemo = {}

    # Словарь соответствия цветов и его идентификатора в Альфе
    sl_color_di = {'FF969696': '0', 'FF00B050': '1', 'FFFFFF00': '2', 'FFFF0000': '3'}
    # Словарь соответствия типа сигнала(DI или DI_AI) и его ПЛК-Аспекта
    sl_plc_aspect = {'Да': 'DI.DI_PLC_View', 'Нет': 'DI.DI_PLC_View', 'AI': 'DI_AI.DI_AI_PLC_View'}
    # Словарь предупреждений {CPU : {алг.имя : (рус.имя, тип наличия)}}
    sl_wrn_di = {}

    cells = sheet['A1': get_column_letter(100) + str(100)]
    index_alg_name = is_f_ind(cells[0], 'Алгоритмическое имя')
    index_rus_name = is_f_ind(cells[0], 'Наименование параметра')
    index_im = is_f_ind(cells[0], 'ИМ')
    index_color_on = is_f_ind(cells[0], 'Цвет при наличии')
    index_color_off = is_f_ind(cells[0], 'Цвет при отсутствии')
    index_cpu_name = is_f_ind(cells[0], 'CPU')
    index_res = is_f_ind(cells[0], 'Резервный')
    index_control_cel = is_f_ind(cells[0], 'Контроль цепи')
    index_wrn = is_f_ind(cells[0], 'Предупреждение')
    index_wrn_text = is_f_ind(cells[0], 'Текст предупреждения')
    index_node_mnemo = is_f_ind(cells[0], 'Узел')

    cells = sheet['A2': get_column_letter(100) + str(sheet.max_row)]
    # Составляем множество контроллеров, у которых есть данные параметры
    # set_par_cpu = set()
    for par in cells:
        if par[0].value is None:
            break
        # set_par_cpu.add(par[index_cpu_name].value)
        if par[index_cpu_name].value not in sl_wrn_di:
            sl_wrn_di[par[index_cpu_name].value] = {}
        if par[index_res].value == 'Нет' and par[index_im].value == 'Нет':
            if par[index_cpu_name].value not in return_sl_di:
                # return_sl_di = {cpu: {алг_пар: (тип параметра в студии, русское имя, sColorOff, sColorOn)}}
                return_sl_di[par[index_cpu_name].value] = {}

            # Обрабатываем и составляем словарь отличий между параметрами по объектам
            if par[index_cpu_name].value not in return_sl_diff:
                # return_sl = {cpu: {алг_пар: (русское имя, ед изм., короткое имя, количество знаков,
                # количество знаков для истории)}}
                return_sl_diff[par[index_cpu_name].value] = {}
            return_sl_diff[par[index_cpu_name].value].update({
                par[index_alg_name].value.replace('|', '_'):
                    str(par[index_cpu_name].comment).split()[1].split(';') if par[index_cpu_name].comment else ''})

            return_sl_di[par[index_cpu_name].value].update(
                {par[index_alg_name].value.replace('|', '_'): (
                    sl_plc_aspect.get(par[index_control_cel].value),
                    par[index_rus_name].value,
                    sl_color_di.get(par[index_color_off].fill.start_color.index, '404'),
                    sl_color_di.get(par[index_color_on].fill.start_color.index, '404'))})

            # Если есть предупреждение по дискрету и канал не переведён в резерв и не привязан к ИМ,
            # то добавляем в словарь предупреждений по дискретам
            if 'Да' in par[index_wrn].value:
                cpu_name_par = par[index_cpu_name].value
                alg_par = par[index_alg_name].value.replace('|', '_')
                sl_wrn_di[cpu_name_par][alg_par] = (par[index_wrn_text].value, par[index_wrn].value)

            if par[index_cpu_name].value not in return_sl_mnemo:
                return_sl_mnemo[par[index_cpu_name].value] = {}
            if par[index_node_mnemo].value not in return_sl_mnemo[par[index_cpu_name].value]:
                return_sl_mnemo[par[index_cpu_name].value][par[index_node_mnemo].value] = list()
            return_sl_mnemo[par[index_cpu_name].value][par[index_node_mnemo].value].append(
                par[index_alg_name].value.replace('|', '_')
            )

    return return_sl_di, sl_wrn_di, return_sl_mnemo, return_sl_diff


def is_read_im(sheet, sheet_imao):
    # return_sl_im = {cpu: {алг_пар: (русское имя, StartView, Gender, ед изм., количество знаков(для д.ИМ=0),
    # количество знаков для истории)}}
    return_sl_im = {}
    return_sl_diff = {}

    sl_im_plc = {}
    # Словарь соответствия типа ИМ и его ПЛК аспекта
    if os.path.exists(os.path.join('Template_Alpha', 'Systemach', f'dict_im')):
        with open(os.path.join('Template_Alpha', 'Systemach', f'dict_im'), 'r', encoding='UTF-8') as f_signal:
            for line in f_signal:
                if '#' in line:
                    continue
                if not line.strip():
                    break
                line = line.strip()
                lst_line = [i.strip() for i in line.split(':')]
                sl_im_plc[lst_line[0]] = lst_line[1]
    else:
        print(f'Не найден файл сигналов Template_Alpha/Systemach/dict_im, структуры ИМ добавлены не будут')
    # sl_im_plc = {'ИМ1Х0': 'IM1x0.IM1x0_PLC_View', 'ИМ1Х1': 'IM1x1.IM1x1_PLC_View', 'ИМ1Х2': 'IM1x2.IM1x2_PLC_View',
    #              'ИМ2Х2': 'IM2x2.IM2x2_PLC_View', 'ИМ2Х4': 'IM2x2.IM2x4_PLC_View',
    #              'ИМ1Х0и': 'IM1x0inv.IM1x0inv_PLC_View', 'ИМ1Х1и': 'IM1x1inv.IM1x1inv_PLC_View',
    #              'ИМ1Х2и': 'IM1x2inv.IM1x2inv_PLC_View',
    #              'ИМ2Х2с': 'IM2x2.IM2x2_PLC_View',
    #              'ИМАО': 'IM_AO.IM_AO_PLC_View', 'ИМ2Х2ПЧ': 'IM2x2PCH.IM2x2PCH_PLC_View'}

    # Словарь соответствия рода ИМ и его идентификатора в Альфе
    sl_gender = {'С': '0', 'М': '1', 'Ж': '2'}

    cells = sheet['A1': get_column_letter(200) + str(200)]
    index_alg_name = is_f_ind(cells[0], 'Алгоритмическое имя')
    index_rus_name = is_f_ind(cells[0], 'Наименование параметра')
    index_type_im = is_f_ind(cells[0], 'Тип ИМ')
    index_gender = is_f_ind(cells[0], 'Род')
    index_cpu_name = is_f_ind(cells[0], 'CPU')
    index_work_time = is_f_ind(cells[0], 'Считать наработку')
    index_swap = is_f_ind(cells[0], 'Считать перестановки')
    index_start_view = is_f_ind(cells[0], 'Тип ИМ', target_count=2)
    index_save_history = is_f_ind(cells[0], 'Сохранять в истории')

    cells = sheet['A2': get_column_letter(200) + str(sheet.max_row)]
    # Составляем множество контроллеров, у которых есть данные параметры
    # set_par_cpu = set()
    sl_cnt = {}
    for par in cells:
        if par[0].value is None:
            break
        # set_par_cpu.add(par[index_cpu_name].value)
        # Проверка комментария
        # if par[index_cpu_name].comment:
        #     print(str(par[index_cpu_name].comment).split()[1].split(';'))
        # Если считаем наработку, то добавляем в словарь sl_cnt = {CPU: {алг.имя : русское имя}}
        if par[index_work_time].value == 'Да':
            if par[index_cpu_name].value not in sl_cnt:
                sl_cnt[par[index_cpu_name].value] = {}
            sl_cnt[par[index_cpu_name].value].update(
                {par[index_alg_name].value + '_WorkTime': par[index_rus_name].value})
        # Если считаем перестановки, то добавляем в словарь sl_cnt = {CPU: {алг.имя : русское имя}}
        if par[index_swap].value == 'Да':
            if par[index_cpu_name].value not in sl_cnt:
                sl_cnt[par[index_cpu_name].value] = {}
            sl_cnt[par[index_cpu_name].value].update(
                {par[index_alg_name].value + '_Swap': par[index_rus_name].value})

        # Обрабатываем и составляем словарь отличий между параметрами по объектам
        if par[index_cpu_name].value not in return_sl_diff:
            # return_sl = {cpu: {алг_пар: (русское имя, ед изм., короткое имя, количество знаков,
            # количество знаков для истории)}}
            return_sl_diff[par[index_cpu_name].value] = {}

        if par[index_cpu_name].value not in return_sl_im:
            # return_sl_im = {cpu: {алг_пар: (тип ИМа, русское имя, StartView, Gender, ед изм. (для д.ИМ='-'),
            # количество знаков(для д.ИМ=0), количество знаков для истории)}}
            return_sl_im[par[index_cpu_name].value] = {}
        if par[index_type_im].value in sl_im_plc:
            # print(par[index_save_history].value) !!!
            return_sl_im[par[index_cpu_name].value].update(
                {par[index_alg_name].value: (sl_im_plc.get(par[index_type_im].value),
                                             par[index_rus_name].value,
                                             ''.join([i for i in par[index_start_view].value if i.isdigit()]),
                                             sl_gender.get(par[index_gender].value),
                                             '-',
                                             '0',
                                             '0')})

            return_sl_diff[par[index_cpu_name].value].update({
                par[index_alg_name].value:
                    str(par[index_cpu_name].comment).split()[1].split(';') if par[index_cpu_name].comment else ''})

    # Обрабатываем ИМ АО
    cells = sheet_imao['A1': get_column_letter(200) + str(200)]

    index_alg_name = is_f_ind(cells[0], 'Алгоритмическое имя')
    index_rus_name = is_f_ind(cells[0], 'Наименование параметра')
    index_start_view = is_f_ind(cells[0], 'Тип ИМ')
    index_yes_im = is_f_ind(cells[0], 'ИМ')
    index_gender = is_f_ind(cells[0], 'Род')
    index_res = is_f_ind(cells[0], 'Резервный')
    index_cpu_name = is_f_ind(cells[0], 'CPU')
    index_frag_dig = is_f_ind(cells[0], 'Количество знаков')
    index_unit = is_f_ind(cells[0], 'Единицы измерения')

    cells = sheet_imao['A2': get_column_letter(200) + str(sheet.max_row)]
    # Составляем множество контроллеров, у которых есть данные параметры
    # set_par_cpu_imao = set()
    for par in cells:
        if par[0].value is None:
            break
        # set_par_cpu_imao.add(par[index_cpu_name].value)
        # Проверка комментария
        # if par[index_cpu_name].comment:
        #     print(str(par[index_cpu_name].comment).split()[1].split(';'))

        if par[index_yes_im].value == 'Да' and par[index_res].value == 'Нет':
            if par[index_cpu_name].value not in return_sl_im:
                # return_sl_im = {cpu: {алг_пар: (тип ИМа, русское имя, StartView, Gender, ед изм., количество знаков,
                # количество знаков для истории)}}
                return_sl_im[par[index_cpu_name].value] = {}
            frag_dig_hist = ''.join([i for i
                                     in str(par[index_frag_dig].comment)[:str(par[index_frag_dig].comment).find('by')]
                                     if i.isdigit()]) if par[index_frag_dig].comment else par[index_frag_dig].value

            # Обрабатываем и составляем словарь отличий между параметрами по объектам
            if par[index_cpu_name].value not in return_sl_diff:
                # return_sl = {cpu: {алг_пар: (русское имя, ед изм., короткое имя, количество знаков,
                # количество знаков для истории)}}
                return_sl_diff[par[index_cpu_name].value] = {}

            if 'ИМАО' in sl_im_plc:
                return_sl_im[par[index_cpu_name].value].update(
                    {par[index_alg_name].value: (sl_im_plc.get('ИМАО'),
                                                 par[index_rus_name].value,
                                                 ''.join([i for i in par[index_start_view].value if i.isdigit()]),
                                                 sl_gender.get(par[index_gender].value),
                                                 par[index_unit].value,
                                                 par[index_frag_dig].value,
                                                 frag_dig_hist)})

                return_sl_diff[par[index_cpu_name].value].update({
                    par[index_alg_name].value:
                        str(par[index_cpu_name].comment).split()[1].split(';') if par[index_cpu_name].comment else ''})

    return return_sl_im, sl_cnt, return_sl_diff


def is_read_create_diag(book, name_prj, *sheets_signal):
    sheet_module = book['Модули']
    sl_cpu_res = {}
    # sl_cpu_rus_name = {alg.cpu: русское имя}
    sl_cpu_rus_name = {}
    # Словарь возможных модулей со стартовым описанием каналов
    sl_modules_channel = {}
    if os.path.exists(os.path.join('Template_Alpha', 'Systemach', f'dict_module_channel')):
        with open(os.path.join('Template_Alpha', 'Systemach', f'dict_module_channel'), 'r', encoding='UTF-8') as f_sig:
            for line in f_sig:
                if '#' in line:
                    continue
                if not line.strip():
                    break
                line = line.strip()
                lst_line = [i.strip() for i in line.split(':')]
                if '*' in lst_line[1]:
                    tm = [i.strip() for i in lst_line[1].split('*')]
                    sl_modules_channel[lst_line[0]] = [tm[0]] * int(tm[1])
                else:
                    sl_modules_channel[lst_line[0]] = lst_line[1]
    else:
        print(f'Не найден файл сигналов Template_Alpha/Systemach/dict_module_channel, '
              f'структуры модулей добавлены не будут')
    # sl_modules = {
    #     'M547A': ['Резерв'] * 16,
    #     'M537V': ['Резерв'] * 8,
    #     'M557D': ['Резерв'] * 32,
    #     'M557O': ['Резерв'] * 32,
    #     'M932C_2N': ['Резерв'] * 8,
    #     'M903E': 'CPU', 'M991E': 'CPU', 'M915E': 'CPU', 'M501E': 'CPU', 'M991S': 'CPU',
    #     'M548A': ['Резерв'] * 16,
    #     'M538V': ['Резерв'] * 8,
    #     'M558D': ['Резерв'] * 32,
    #     'M558O': ['Резерв'] * 32,
    #     'M531I': ['Резерв'] * 8,
    #     'M543G': ['Резерв'] * 16,
    #     'M5571': ['Резерв'] * 32,
    #     'M532U': ['Резерв'] * 8,  # Добавлено в тестовом режиме для Игринской!!!
    #     'M582IS': ['Резерв'] * 3,  # Добавлено в тестовом режиме для Игринской!!!
    # }
    sl_type_modules = {}
    tuple_cpu_name = tuple()
    if os.path.exists(os.path.join('Template_Alpha', 'Systemach', f'dict_module')):
        with open(os.path.join('Template_Alpha', 'Systemach', f'dict_module'), 'r', encoding='UTF-8') as f_signal:
            for line in f_signal:
                if '#' in line:
                    continue
                if not line.strip():
                    break
                line = line.strip()
                lst_line = [i.strip() for i in line.split(':')]
                sl_type_modules[lst_line[0]] = lst_line[1]
                if 'CPU' in lst_line[1]:
                    tuple_cpu_name += (lst_line[0],)
    else:
        print(f'Не найден файл сигналов Template_Alpha/Systemach/dict_module, структуры модулей добавлены не будут')

    # sl_type_modules = {
    #     'M903E': 'Types.DIAG_CPU.DIAG_CPU_M903E_PLC_View',
    #     'M991E': 'Types.DIAG_CPU.DIAG_CPU_M991E_PLC_View',
    #     'M991S': 'Types.DIAG_CPU.DIAG_CPU_M991E_PLC_View',
    #     'M547A': 'Types.DIAG_M547A.DIAG_M547A_PLC_View',
    #     'M548A': 'Types.DIAG_M548A.DIAG_M548A_PLC_View',
    #     'M537V': 'Types.DIAG_M537V.DIAG_M537V_PLC_View',
    #     'M538V': 'Types.DIAG_M538V.DIAG_M538V_PLC_View',
    #     'M932C_2N': 'Types.DIAG_M932C_2N.DIAG_M932C_2N_PLC_View',
    #     'M557D': 'Types.DIAG_M557D.DIAG_M557D_PLC_View',
    #     'M558D': 'Types.DIAG_M558D.DIAG_M558D_PLC_View',
    #     'M557O': 'Types.DIAG_M557O.DIAG_M557O_PLC_View',
    #     'M558O': 'Types.DIAG_M558O.DIAG_M558O_PLC_View',
    #     'M915E': 'Types.DIAG_CPU.DIAG_CPU_M915E_PLC_View',
    #     'M501E': 'Types.DIAG_CPU.DIAG_CPU_M501E_PLC_View',
    #     'M531I': 'Types.DIAG_M531I.DIAG_M531I_PLC_View',
    #     'M543G': 'Types.DIAG_M543G.DIAG_M543G_PLC_View',
    #     'M532U': 'Types.DIAG_M532U_test.DIAG_M532U_test_PLC_View',  # Добавлено в тестовом режиме для Игринской!!!
    #     'M582IS': 'Types.DIAG_M582IS_test.DIAG_M582IS_test_PLC_View',  # Добавлено в тестовом режиме для Игринской!!!
    # }
    cells = sheet_module['A1': get_column_letter(50) + str(50)]
    type_module_index = is_f_ind(cells[0], 'Шифр модуля')
    name_module_index = is_f_ind(cells[0], 'Имя модуля')
    view_module_index = is_f_ind(cells[0], 'Тип модуля')
    cpu_index = is_f_ind(cells[0], 'CPU')
    cells = sheet_module['A2': get_column_letter(50) + str(sheet_module.max_row)]
    sl_modules_cpu = {}
    # словарь sl_modules_cpu {имя CPU: {имя модуля: (тип модуля, [каналы])}}
    for p in cells:
        if p[0].value is None:
            break
        # Собираем русские имена контроллеров
        if p[view_module_index].value == 'Контроллер':
            sl_cpu_rus_name[p[cpu_index].value] = p[name_module_index].value
        # Собираем резервируемые контроллеры
        if p[0].comment:
            name_res = str(p[0].comment)[str(p[0].comment).find(' ') + 1: str(p[0].comment).find('by')].strip()
            name_osn = p[0].value
            sl_cpu_res[p[cpu_index].value] = (name_osn, name_res)
        if sl_modules_channel.get(p[type_module_index].value):
            aa = copy(sl_modules_channel[p[type_module_index].value])
            if p[cpu_index].value not in sl_modules_cpu:
                sl_modules_cpu[p[cpu_index].value] = {}
            sl_modules_cpu[p[cpu_index].value].update({p[name_module_index].value: (p[type_module_index].value, aa)})
    # В тестовом режиме для игринской добавляем два новых модуля в GTU жёстко!!!
    if 'Игринская' in name_prj:
        if sl_modules_channel.get('M532U') and sl_modules_channel.get('M582IS'):
            aa = copy(sl_modules_channel['M532U'])
            sl_modules_cpu['GTU'].update({'AD102': ('M532U', aa)})
            aa = copy(sl_modules_channel['M582IS'])
            sl_modules_cpu['GTU'].update({'AD1': ('M582IS', aa)})

    # sl_for_diag - словарь для корректной передачи для создания индексов
    sl_for_diag = {}
    for name_cpu, value in sl_modules_cpu.items():
        keys_sl_for_diag = [i if value[i][0] not in tuple_cpu_name
                            else 'CPU' for i in value]
        value_sl_for_diag = [value[i][0] if value[i][0] not in tuple_cpu_name
                             else (i, value[i][0]) for i in value]
        sl_for_diag[name_cpu] = dict(zip(keys_sl_for_diag, value_sl_for_diag))

    # пробегаемся по листам, где могут быть указаны каналы модулей
    for jj in sheets_signal:
        sheet_run = book[jj]
        cells_run = sheet_run['A1': 'O' + str(sheet_run.max_column)]
        num_canal_index = is_f_ind(cells_run[0], 'Номер канала')
        no_stand_index = is_f_ind(cells_run[0], 'Нестандартный канал')
        cpu_par_index = is_f_ind(cells_run[0], 'CPU')
        name_module_par_index = is_f_ind(cells_run[0], 'Номер модуля')
        name_par_index = is_f_ind(cells_run[0], 'Наименование параметра')
        control_index = is_f_ind(cells_run[0], 'Контроль цепи')
        no_stand_kc_index = is_f_ind(cells_run[0], 'Нестандартный канал КЦ')
        name_module_par_kc_index = is_f_ind(cells_run[0], 'Номер модуля контроля')
        num_canal_kc_index = is_f_ind(cells_run[0], 'Номер канала контроля')
        reserve_par_index = is_f_ind(cells_run[0], 'Резервный')
        cells_run = sheet_run['A2': 'O' + str(sheet_run.max_row)]
        # пробегаемся по параметрам на листе
        for par in cells_run:
            # Если сигнал не резервный...
            if par[reserve_par_index].value == 'Нет':
                # если не указан НЕстандартный канал, то вносим в список
                if par[no_stand_index].value == 'Нет':
                    tmp_ind = int(par[num_canal_index].value) - 1
                    if sl_modules_cpu.get(par[cpu_par_index].value) and \
                            sl_modules_cpu[par[cpu_par_index].value].get(par[name_module_par_index].value):
                        sl_modules_cpu[par[cpu_par_index].value][par[name_module_par_index].value][1][tmp_ind] = \
                            par[name_par_index].value
                # если выбран контроль цепи и контроль стандартный, то также добавляем в список
                if par[control_index].value == 'Да' and par[no_stand_kc_index].value == 'Нет':
                    tmp_ind = int(par[num_canal_kc_index].value) - 1
                    if sl_modules_cpu.get(par[cpu_par_index].value) and \
                            sl_modules_cpu[par[cpu_par_index].value].get(par[name_module_par_kc_index].value):
                        sl_modules_cpu[par[cpu_par_index].value][par[name_module_par_kc_index].value][1][tmp_ind] = \
                            f"КЦ: {par[name_par_index].value}"

    # sl_modules_cpu {имя CPU: {имя модуля: (тип модуля в студии, тип модуля, [каналы])}}
    sl_modules_cpu = {cpu: {mod: (sl_type_modules.get(value[0], 'Types.').replace('Types.', ''), ) + value
                            for mod, value in sl_modules_cpu[cpu].items() if sl_type_modules.get(value[0])}
                      for cpu in sl_modules_cpu}
    for cpu in sl_modules_cpu:
        if cpu in sl_cpu_res:
            for j in [i for i in sl_modules_cpu[cpu] if i in sl_cpu_res[cpu]]:
                sl_modules_cpu[cpu][j] = (sl_modules_cpu[cpu][j][0].replace('_PLC_View', '_Res_PLC_View'), ) + \
                                         sl_modules_cpu[cpu][j][1:]
    return sl_modules_cpu, sl_for_diag, sl_cpu_res, sl_cpu_rus_name


def is_create_net(sl_object_all, sheet_net):
    # Проработать в студии IOlogic - настройка в конфигураторе !!!
    cells = sheet_net['A1': get_column_letter(60) + str(60)]

    index_rus_name = is_f_ind(cells[0], 'Наименование юнита')
    index_alg_name = is_f_ind(cells[0], 'Алгоритмическое имя')
    index_object_name = is_f_ind(cells[0], 'Объект')
    index_type = is_f_ind(cells[0], 'Тип')
    index_ip = is_f_ind(cells[0], 'IP-адрес')
    index_ip_res = is_f_ind(cells[0], 'Резервный IP-адрес')
    index_option = is_f_ind(cells[0], 'Настройка')

    cells = sheet_net['A2': get_column_letter(60) + str(sheet_net.max_row)]

    # tuple_net_res = ('SWITCH_TREI_S34X',)
    tuple_net_with_option = ('IOLOGIC_MOXA_E2242',)

    # return_sl_net = {Объект: {алг.имя: {параметры столбцов: значение}}}
    return_sl_net = {}
    for par in cells:
        if par[0].value is None:
            break
        if par[index_object_name].value not in return_sl_net:
            return_sl_net[par[index_object_name].value] = {}
        return_sl_net[par[index_object_name].value].update({par[index_alg_name].value: {
            'Unit': par[index_rus_name].value,
            'Type': par[index_type].value,
            'IP': '.'.join([a.lstrip('0') for a in par[index_ip].value.split('.')]),
            'IP_res': ('.'.join([a.lstrip('0') for a in par[index_ip_res].value.split('.')])
                       if par[index_ip_res].fill.start_color.index == '00000000' else ''),
            'Option': (par[index_option].value if par[index_type].value in tuple_net_with_option else '')}})

    # Для каждого объекта
    for objects in sl_object_all:
        # Если нашли объект в словаре
        if objects[0] in return_sl_net:
            root_plc_aspect = ET.Element('omx', xmlns="system", xmlns_dp="automation.deployment",
                                         xmlns_snmp="automation.snmp", xmlns_ct="automation.control")
            for alg, sl_value in return_sl_net[objects[0]].items():
                child_alg = ET.SubElement(root_plc_aspect, 'dp_computer', name=f"{objects[0]}_{alg}")
                ET.SubElement(child_alg, 'dp_ethernet-adapter', name="Eth1", address=sl_value['IP'])
                if sl_value['IP_res']:
                    ET.SubElement(child_alg, 'dp_ethernet-adapter', name="Eth2", address=sl_value['IP_res'])
                if "checkip" not in sl_value['Type'].lower():
                    child_runtime = ET.SubElement(child_alg, 'dp_external-runtime', name="Runtime")
                    ET.SubElement(child_runtime, 'snmp_snmp-agent', name="SnmpAgent",
                                  poll_port="161", poll_password="public", notification_port="162",
                                  notification_password="public", protocol_version="Snmp_v1",
                                  security_level="NoAuthNoPriv", auth_protocol="MD5", priv_protocol="AES",
                                  address_map="Application.SnmpLinkMap")
                    child_app = ET.SubElement(child_runtime, 'dp_application-object', name="Application",
                                              access_level="public")
                    ET.SubElement(child_app, 'snmp_snmp-link-map', name="SnmpLinkMap",
                                  file='SNMP\\' + f'{sl_value["Type"]}_map.xml')
                    child_data = ET.SubElement(child_app, 'ct_object', name="Data", access_level="public")
                    child_data_app = ET.SubElement(child_data, 'ct_object', name="Data",
                                                   base_type=f"Types.SNMP_Switch.{sl_value['Type']}_PLC_View",
                                                   aspect="Types.PLC_Aspect",
                                                   access_level="public")
                    ET.SubElement(child_data_app, 'attribute', type="unit.System.Attributes.Description",
                                  value=f"{sl_value['Unit']}")

            # Нормируем и записываем IOS-аспект
            temp = ET.tostring(root_plc_aspect).decode('UTF-8')

            check_diff_file(check_path=os.path.join('File_for_Import', 'PLC_Aspect_importDomain'),
                            file_name_check=f'file_out_NET_{objects[0]}.omx-export',
                            new_data=multiple_replace_xml(lxml.etree.tostring(lxml.etree.fromstring(temp),
                                                                              pretty_print=True, encoding='unicode')),
                            message_print=f'Требуется заменить ПЛК-аспект сети объекта {objects[0]} '
                                          f'(Файл NET_{objects[0]})')

    # Собираем карту пингушечки
    # return_sl_net = {Объект: {алг.имя: {параметры столбцов: значение}}}
    # print(return_sl_net)

    tmp_signal_netdiag = '  <item Binding="Introduced">\n' \
                         '    <node-path>$name_signal</node-path>\n' \
                         '    <node>$device</node>\n' \
                         '    <adapter>$adapter</adapter>\n' \
                         '    <request>Ping</request>\n' \
                         '    <function>$function</function>\n' \
                         '  </item>\n'

    s_all = ''
    # root_map_net = ET.Element('root', format_version="0")

    for obj, sl_device_net in return_sl_net.items():
        for device, sl_device_par in sl_device_net.items():
            if obj != 'Система':
                for ip_key, adapter in {'IP': "Eth1", 'IP_res': "Eth2"}.items():
                    if sl_device_par.get(ip_key):
                        name_node = f'{obj}.Diag.NET.{obj}_{device}.Ping_{ip_key}'
                        for signal_function in ('Enable', 'ResetStat', 'FailCount',
                                                'LastError', 'LastFailDuration', 'Status', 'SuccCount',
                                                'TimeOut', 'TotalFailDuration',
                                                'Filtered.FailedAttemptsCount', 'Filtered.Status',
                                                'Filtered.SuccAttemptsCount', 'IPaddress', 'RTTTime'):

                            s_all += Template(tmp_signal_netdiag).substitute(
                                name_signal=f'{name_node}.{signal_function}',
                                device=f'{obj}_{device}',
                                adapter=adapter,
                                function=signal_function
                            )

    # Проверка изменений, и если есть изменения, то запись
    # Если нет папки File_for_Import, то создадим её
    if not os.path.exists('File_for_Import'):
        os.mkdir('File_for_Import')
    # Если нет папки File_for_Import/Maps, то создадим её
    if not os.path.exists(os.path.join('File_for_Import', 'Maps')):
        os.mkdir(os.path.join('File_for_Import', 'Maps'))

    new_map = '<root format-version=\"0\">\n' + s_all.rstrip() + '\n</root>'

    check_diff_file(check_path=os.path.join('File_for_Import', 'Maps'),
                    file_name_check=f'NetDiagAddressMap.xml',
                    new_data=new_map,
                    message_print=f'Требуется заменить карту NetDiagAddressMap.xml')

    return return_sl_net


def is_create_service_signal(sl_object_all: dict, sl_cpu_res: dict, architecture: str,
                             server_name_osn: str, server_name_rez: str, sl_cpu_archive: dict):
    sl_num_eth = {1: 'Основная сеть', 2: 'Резервная сеть'}
    # Новая часть
    root_aspect = ET.Element('omx', xmlns="system", xmlns_ct="automation.control", xmlns_r="automation.reference")
    child_service = ET.SubElement(root_aspect, 'ct_object', name='Service', access_level="public")
    ET.SubElement(child_service, 'attribute', type="unit.Server.Attributes.Replicate", value="false")
    child_modules = ET.SubElement(child_service, 'ct_object', name='Modules', access_level="public")
    child_unet = ET.SubElement(child_modules, 'ct_object', name='UNET Client', access_level="public")
    ET.SubElement(child_unet, 'attribute', type="unit.System.Attributes.Comment", value="для диагностики связи с ПЛК")

    # Для каждого объекта...
    for objects in sl_object_all:
        # ...для каждого контроллера...
        for cpu in sl_object_all[objects]:
            child_plc = ET.SubElement(child_unet, 'ct_object', name=f'PLC_{cpu}_{objects[2]}', access_level="public")
            for numeth in range(1, 3):
                child_eth = ET.SubElement(child_plc, 'ct_object', name=f'CPU_Eth{numeth}', access_level="public")
                ET.SubElement(child_eth, 'ct_parameter', name='IsConnected',
                              type='bool', direction='out', access_level="public")
                ET.SubElement(child_eth, 'attribute', type='unit.System.Attributes.Comment',
                              value=f"{sl_num_eth.get(numeth)}")
            # Добавляем диагностику по резервным контроллерам, если такие есть
            if cpu in sl_cpu_res:
                for numeth in range(1, 3):
                    child_eth = ET.SubElement(child_plc, 'ct_object', name=f'R-CPU_Eth{numeth}', access_level="public")
                    ET.SubElement(child_eth, 'ct_parameter', name='IsConnected',
                                  type='bool', direction='out', access_level="public")
                    ET.SubElement(child_eth, 'attribute', type='unit.System.Attributes.Comment',
                                  value=f"{sl_num_eth.get(numeth)} резервного ПЛК")

    child_state = ET.SubElement(child_service, 'ct_object', name='State', access_level="public")
    ET.SubElement(child_state, 'ct_parameter', name='Server',
                  type='bool', direction='out', access_level="public")
    child_hist = ET.SubElement(child_modules, 'ct_object', name='HistoryModule', access_level="public")
    child_storage = ET.SubElement(child_hist, 'ct_object', name='Storages', access_level="public")
    child_historian = ET.SubElement(child_storage, 'ct_object', name='Historian', access_level="public")

    if architecture == 'клиент-сервер':
        # Добавляем структуру InervalLock в Historian (оставляем только папочки)
        for serv_name in (server_name_osn, server_name_rez):
            child_hist_serv = ET.SubElement(child_historian, 'ct_object',
                                            attrib={'name': f'Database_{serv_name}', 'access-level': 'public'})
            child_hist_serv_2 = ET.SubElement(child_hist_serv, 'ct_object',
                                              attrib={'name': f'Database_{serv_name}', 'access-level': 'public'})
            # child_interval_lock = ET.SubElement(child_hist_serv_2, 'ct_object', name="IntervalLock",
            #                                     access_level="public")
            # obj_high = ET.SubElement(child_interval_lock, 'ct_parameter', name='High', type='uint64',
            #                          direction='out', access_level="public")
            # ET.SubElement(obj_high, 'attribute', type='unit.Server.Attributes.History',
            #               value=f"Enable=\"True\" ServerTime=\"False\"")
            #
            # obj_low = ET.SubElement(child_interval_lock, 'ct_parameter', name='Low', type='uint64',
            #                         direction='out', access_level="public")
            # ET.SubElement(obj_low, 'attribute', type='unit.Server.Attributes.History',
            #               value=f"Enable=\"True\" ServerTime=\"False\"")
            # obj_write_lock = ET.SubElement(child_interval_lock, 'ct_parameter', name='WriteLock', type='bool',
            #                                direction='out', access_level="public")
            # ET.SubElement(obj_write_lock, 'attribute', type='unit.Server.Attributes.History',
            #               value=f"Enable=\"True\" ServerTime=\"False\"")
            # ET.SubElement(obj_write_lock, 'attribute', type='unit.Server.Attributes.Alarm',
            #               value=f'{{"Condition":{{"IsEnabled":"true",'
            #                     f'"Subconditions":[{{"AckStrategy":1,"IsEnabled":true,'
            #                     f'"Message":"Архивация выполнена",'
            #                     f'"Severity":500,"Type":2}},'
            #                     f'{{"AckStrategy":1,"IsEnabled":false,'
            #                     f'"Message":"___",'
            #                     f'"Severity":500,"Type":3}}],'
            #                     f'"Type":2}}}}')
            # ET.SubElement(child_interval_lock, 'ct_parameter', name='Description', type='string',
            #               direction='out', access_level="public")
            # ET.SubElement(child_interval_lock, 'ct_subject-ref', name=f'rlock', object=f"rlock", const_access="false",
            #               aspected="false")
            # ET.SubElement(child_interval_lock, 'ct_subject-ref',
            #               name=f'_State', object=f"State", const_access="false",
            #               aspected="false")
            #
            # object_handler = ET.SubElement(
            #     child_interval_lock, 'ct_handler', name=f'Handler',
            #     source_code=f'if (rlock.setAlarm.Value && _State.Server.Value) {{\n'
            #                 f'\tcommit Description = rlock.setDescription;\n'
            #                 f'\tcommit High = rlock.High;\n'
            #                 f'\tcommit Low = rlock.Low;\n'
            #                 f'\tcommit WriteLock = true;\n}}'
            # )
            # ET.SubElement(object_handler, 'ct_trigger',
            #               on=f"rlock.setAlarm", cause="update")

        # Добавляем Redundancy
        child_redundancy = ET.SubElement(child_service, 'ct_object', name='Redundancy', access_level="public")
        ET.SubElement(child_redundancy, 'ct_parameter', name='Switch',
                      type='bool', direction='out', access_level="public")
        ET.SubElement(child_redundancy, 'ct_formula', source_code="false", target="Switch")
        for ch in ('Channel1', 'Channel2'):
            child_ch = ET.SubElement(child_redundancy, 'ct_object', name=f'{ch}', access_level="public")
            ET.SubElement(child_ch, 'ct_parameter', name='ConnectionEstablished',
                          type='bool', direction='out', access_level="public")

        # # Добавляем rlock
        # child_obj_rlock = ET.SubElement(child_service, 'ct_object', name="rlock",
        #                                 access_level="public")
        # ET.SubElement(child_obj_rlock, 'ct_parameter', name='setAlarm', type='bool', direction='in',
        #               access_level="public")
        # ET.SubElement(child_obj_rlock, 'ct_parameter', name='setDescription', type='string',
        #               direction='in', access_level="public")
        # ET.SubElement(child_obj_rlock, 'ct_parameter', name='High', type='uint64', direction='in',
        #               access_level="public")
        # ET.SubElement(child_obj_rlock, 'ct_parameter',
        #               name='Low', type='uint64', direction='in', access_level="public")
        # ET.SubElement(child_obj_rlock, 'ct_parameter', name='setUserAlarm', type='bool',
        #               direction='in', access_level="public")
        # # Для каждого объекта
        # for objects, sl_plc in sl_object_all.items():
        #     for plc in sl_plc:
        #         if plc in sl_cpu_archive:
        #             for signal_arch, rus_name_arch in sl_cpu_archive[plc].items():
        #                 ET.SubElement(child_obj_rlock, 'ct_parameter', name=f'Low_{signal_arch}{objects[0]}_tmp',
        #                               type='uint64', direction='in',
        #                               access_level="public")
        #                 ET.SubElement(child_obj_rlock, 'ct_parameter', name=f'High_{signal_arch}{objects[0]}_tmp',
        #                               type='uint64', direction='in',
        #                               access_level="public")
        #                 ET.SubElement(child_obj_rlock, 'ct_subject-ref', name=f'{objects[0]}_{signal_arch}_START',
        #                               object=f"{objects[0]}.System.TS.{signal_arch}_START", const_access="false",
        #                               aspected="false")
        #                 ET.SubElement(child_obj_rlock, 'ct_subject-ref', name=f'{objects[0]}_{signal_arch}_STOP',
        #                               object=f"{objects[0]}.System.TS.{signal_arch}_STOP", const_access="false",
        #                               aspected="false")
        #                 object_handler_start = ET.SubElement(
        #                     child_obj_rlock, 'ct_handler', name=f'{signal_arch}_Handler_START_{objects[0]}',
        #                     source_code=f'if ({objects[0]}_{signal_arch}_START.PLCSignals.Value)\n{{\n'
        #                                 f'\tcommit Low_{signal_arch}{objects[0]}_tmp = '
        #                                 f'Variant.ToUint8(DateTime.UtcNow() - 18000000000,0);'
        #                                 f'//Время до начала события - по + 18000000000,0 - 30 минут\n}}'
        #                 )
        #                 ET.SubElement(object_handler_start, 'ct_trigger',
        #                               on=f"{objects[0]}_{signal_arch}_START.PLCSignals.Value", cause="change")
        #                 object_handler_stop = ET.SubElement(
        #                     child_obj_rlock, 'ct_handler', name=f'{signal_arch}_Handler_STOP_{objects[0]}',
        #                     source_code=f'if ({objects[0]}_{signal_arch}_STOP.PLCSignals.Value)\n{{\n'
        #                                 f'\tcommit setDescription.Value = \"{objects[1]} {rus_name_arch}\";\n'
        #                                 f'\tcommit High_{signal_arch}{objects[0]}_tmp = '
        #                                 f'Variant.ToUint8(DateTime.UtcNow() + 18000000000,0);'
        #                                 f'//Время после события - по + 18000000000,0 - 30 минут\n'
        #                                 f'\tcommit Low = Low_{signal_arch}{objects[0]}_tmp;\n'
        #                                 f'\tcommit High = High_{signal_arch}{objects[0]}_tmp;\n'
        #                                 f'\tcommit setAlarm.Value = true;\n'
        #                                 f'\tcommit setAlarm.Value = false;\n}}'
        #                 )
        #                 ET.SubElement(object_handler_stop, 'ct_trigger',
        #                               on=f"{objects[0]}_{signal_arch}_STOP.PLCSignals.Value", cause="change")

        # Добавляем SNMP
        child_snmp = ET.SubElement(child_service, 'ct_object', name='SNMP', access_level="public")
        ET.SubElement(child_snmp, 'attribute', type='unit.Server.Attributes.Replicate', value=f"false")
        for comp, par in {
            'Computer1': ('SNMP_Agent.Application.SNMP',),
            'Computer2': (f'{server_name_rez}.SNMP_Agent.Application.SNMP',),
            'ComputerLocal': ('SNMP_AgentLocal.Application.SNMP',)
        }.items():
            child_comp = ET.SubElement(child_snmp, 'ct_object', name=f'{comp}', base_type="Types.SNMP.Server_IOS_View",
                                       access_level="public", aspect="Types.IOS_Aspect",
                                       original=f"{par[0]}")
            ET.SubElement(child_comp, 'ct_init-ref', ref="_Server_PLC", target=f"{par[0]}")

        ET.SubElement(child_snmp, 'ct_subject-ref', name=f'ConnectionStatus', object=f"Redundancy.Channel1",
                      const_access="false",
                      aspected="false")
        ET.SubElement(child_snmp, 'ct_subject-ref', name=f'ConnectionStatus2', object=f"Redundancy.Channel2",
                      const_access="false", aspected="false")
        ET.SubElement(child_snmp, 'ct_parameter', name='prevConnect1',
                      type='bool', direction='out', access_level="public")
        ET.SubElement(child_snmp, 'ct_parameter', name='prevConnect2',
                      type='bool', direction='out', access_level="public")
        ET.SubElement(child_snmp, 'ct_parameter', name='tmpValue',
                      type='string', direction='out', access_level="public")
        ET.SubElement(child_snmp, 'ct_parameter', name='tmpValue2',
                      type='string', direction='out', access_level="public")
        child_counter = ET.SubElement(child_snmp, 'ct_parameter', name='Counter',
                                      type='uint64', direction='out', access_level="public")
        ET.SubElement(child_counter, 'attribute', type='unit.Server.Attributes.Replicate', value=f"true")
        ET.SubElement(child_counter, 'attribute', type='unit.System.Attributes.InitialValue', value=f"1000")
        ET.SubElement(child_counter, 'attribute', type='unit.Server.Attributes.InitialQuality', value=f"192")
        ET.SubElement(child_snmp, 'ct_parameter', name='Value',
                      type='string', direction='out', access_level="public")
        ET.SubElement(child_snmp, 'ct_parameter', name='Value2',
                      type='string', direction='out', access_level="public")
        ET.SubElement(child_snmp, 'ct_timer', name='Timer', period='100')
        ET.SubElement(child_snmp, 'ct_subject-ref', name=f'State', object=f"State", const_access="false",
                      aspected="false")

        object_handler_connectionlost = ET.SubElement(
            child_snmp, 'ct_handler', name=f'Handler_ConnectionLost',
            source_code=f'if (ConnectionStatus.ConnectionEstablished == false && '
                        f'ConnectionStatus2.ConnectionEstablished == false)\n'
                        f'{{\n'
                        f'	localName: string = ComputerLocal.computerName.Value;\n'
                        f'	remoteName: string = "";	\n'
                        f'	if (localName == Computer1.computerName.Value) \n'
                        f'	{{\n'
                        f'		remoteName = Computer2.computerName.Value;\n'
                        f'	}} else \n'
                        f'	{{\n'
                        f'		remoteName = Computer1.computerName.Value;\n'
                        f'	}}\n'
                        f'	tmpValue2 = "";\n'
                        f'	tmpValue = "<Subcondition Type=\\"Dynamic\\" Enabled=\\"1\\" SoundEnabled=\\"0\\" '
                        f'AckRequired =\\"1\\" Sound=\\"\\" Message=\\". На сервере \'" + localName +"\' '
                        'отсутствует связь с резервным сервером \'"+ remoteName +"\'.\\" Severity=\\"40\\"/>";\n'
                        f'}}\n'
                        f'else\n'
                        f'{{\n'
                        f'	tmpValue = "<Subcondition Type=\\"Dynamic\\" Enabled=\\"0\\" />";\n\n}}\n'
                        f'if (ConnectionStatus.ConnectionEstablished != prevConnect1 && '
                        f'ConnectionStatus2.ConnectionEstablished != prevConnect2) {{\n'
                        f'	Counter = 0;\n'
                        f'	prevConnect1 = ConnectionStatus.ConnectionEstablished;\n'
                        f'	prevConnect2 = ConnectionStatus2.ConnectionEstablished;\n'
                        f'}}\n'
        )
        ET.SubElement(object_handler_connectionlost, 'ct_trigger',
                      on=f"ConnectionStatus.ConnectionEstablished", cause="update")
        ET.SubElement(object_handler_connectionlost, 'ct_trigger',
                      on=f"ConnectionStatus2.ConnectionEstablished", cause="update")

        object_handler_switchactive = ET.SubElement(
            child_snmp, 'ct_handler', name=f'Handler_SwitchActive',
            source_code=f'if (State.Server && (ConnectionStatus.ConnectionEstablished || '
                        f'ConnectionStatus2.ConnectionEstablished))\n'
                        f'{{\n'
                        f'	localName: string = ComputerLocal.computerName.Value;\n'
                        f'  remoteName: string = "";\n'
                        f'  if (localName == Computer1.computerName.Value)\n'
                        f'{{\n'
                        f'		remoteName = Computer2.computerName.Value;\n'
                        f'	}} else \n'
                        f'	{{\n'
                        f'		remoteName = Computer1.computerName.Value;\n'
                        f'	}};\n'
                        f'	Counter = 0;\n\n'
                        f'	tmpValue2 = "<Subcondition Type=\\"Dynamic\\" Enabled=\\"1\\" SoundEnabled=\\"0\\"  '
                        f'AckRequired =\\"1\\" Sound=\\"\\" Message=\\". '
                        f'Произошел резервный переход. Активный сервер \'" + localName +"\'\\" Severity=\\"40\\"/>";\n'
                        f'}}\n'

        )
        ET.SubElement(object_handler_switchactive, 'ct_trigger',
                      on=f"State.Server", cause="update")

        object_handler_switchred = ET.SubElement(
            child_snmp, 'ct_handler', name=f'Handler_SwitchRed',
            source_code=f'if (State.Server && (ConnectionStatus.ConnectionEstablished || '
                        f'ConnectionStatus2.ConnectionEstablished))\n'
                        f'{{\n'
                        f'	localName: string = ComputerLocal.computerName.Value;\n'
                        f'  remoteName: string = "";\n'
                        f'  if (localName == Computer1.computerName.Value)\n'
                        f'{{\n'
                        f'		remoteName = Computer2.computerName.Value;\n'
                        f'	}} else \n'
                        f'	{{\n'
                        f'		remoteName = Computer1.computerName.Value;\n'
                        f'	}};\n'
                        f'	Counter = 0;\n\n'
                        f'	tmpValue = "<Subcondition Type=\\"Dynamic\\" Enabled=\\"1\\" SoundEnabled=\\"0\\"  '
                        f'AckRequired =\\"1\\" Sound=\\"\\" Message=\\". '
                        f'Произошел резервный переход. Резервный сервер \'" + remoteName +"\'\\" Severity=\\"40\\"/>";\n'
                        f'}}\n'

        )
        ET.SubElement(object_handler_switchred, 'ct_trigger',
                      on=f"State.Server", cause="update")

        object_handler_1 = ET.SubElement(
            child_snmp, 'ct_handler', name=f'Handler',
            source_code=f'Counter = Counter + 1;\n'
                        f'if (Counter == 20) \n'
                        f'{{\n'
                        f'	if (ConnectionStatus.ConnectionEstablished || ConnectionStatus2.ConnectionEstablished) {{\n'
                        f'		Value = "<Subcondition Type=\\"Dynamic\\" Enabled=\\"0\\" />";\n'
                        f'	Value2 = "<Subcondition Type=\\"Dynamic\\" Enabled=\\"0\\" />";\n'
                        f'	}}\n'
                        f'}};\n\n'
                        f'if (Counter == 21)\n'
                        f'{{\n'
                        f'	Value = tmpValue;\n'
                        f'	Value2 = tmpValue2;\n'
                        f'}};\n\n'
                        f'if (Counter == 70)\n'
                        f'{{\n'
                        f'	if (ConnectionStatus.ConnectionEstablished || ConnectionStatus2.ConnectionEstablished) {{\n'
                        f'		Value = "<Subcondition Type=\\"Dynamic\\" Enabled=\\"0\\" />";\n'
                        f'	Value2 = "<Subcondition Type=\\"Dynamic\\" Enabled=\\"0\\" />";	\n'
                        f'	}}\n}}\n'


        )
        ET.SubElement(object_handler_1, 'ct_trigger',
                      on=f"Timer", cause="update")
    elif architecture == 'сингл':
        child_hist_serv = ET.SubElement(child_historian, 'ct_object',
                                        attrib={'name': f'Database', 'access-level': 'public'})
        ET.SubElement(child_hist_serv, 'ct_object',
                      attrib={'name': f'Database', 'access-level': 'public'})

    # Нормируем и записываем IOS-аспект
    temp = ET.tostring(root_aspect).decode('UTF-8')

    check_diff_file(check_path=os.path.join('File_for_Import', 'IOS_Aspect_in_ApplicationServer'),
                    file_name_check=f'Service_object(импортируется в приложение сервера).omx-export',
                    new_data=multiple_replace_xml(lxml.etree.tostring(lxml.etree.fromstring(temp),
                                                                      pretty_print=True, encoding='unicode')),
                    message_print=f'Требуется заменить сервисные сигналы (файл Service_object.omx-export)')

    if architecture == 'клиент-сервер':
        root_aspect = ET.Element('omx')
        for a, b in {'xmlns': 'system', 'xmlns:srv': 'server', "xmlns:eth": 'automation.ethernet'}.items():
            root_aspect.set(a, b)
        snmp = ET.SubElement(root_aspect, 'srv:snmp-manager', name="SNMP ManagerLocal")
        for agent in ('SNMP_AgentLocal.SnmpAgent', f'{server_name_rez}.SNMP_AgentLocal.SnmpAgent'):
            ET.SubElement(snmp, 'srv:agent-config', attrib={'agent': agent, 'requests-interval': '30',
                                                                   'response-interval': "1000"})
        ET.SubElement(snmp, 'eth:ethernet-adapter-binding', adapter="LocalEthernet")

        # Нормируем и записываем IOS-аспект
        temp = ET.tostring(root_aspect).decode('UTF-8')

        check_diff_file(
            check_path=os.path.join('File_for_Import', 'IOS_Aspect_in_ApplicationServer'),
            file_name_check=f'SNMP ManagerLocal(импортируется в сервер.Server).omx-export',
            new_data=multiple_replace_xml(lxml.etree.tostring(lxml.etree.fromstring(temp),
                                                              pretty_print=True, encoding='unicode')),
            message_print=f'Требуется заменить ManagerLocal в основном сервере'
                          f'(файл SNMP ManagerLocal.omx-export)'
        )

    return


# Функция создания узла SYS
def is_create_sys(sl_object_all: dict, name_prj: str, return_sl_net: dict, architecture: str, sl_agreg: dict):
    root_aspect = ET.Element('omx', xmlns="system", xmlns_ct="automation.control", xmlns_r="automation.reference")
    child_sys = ET.SubElement(root_aspect, 'ct_object', name='SYS', access_level="public")
    ET.SubElement(child_sys, 'attribute', type="unit.System.Attributes.Description", value="Система")

    child_message = ET.SubElement(child_sys, 'ct_object', name='Message', access_level="public")
    ET.SubElement(child_message, 'attribute', type="unit.System.Attributes.Comment",
                  value="Для глобальных сообщений оператора")
    child_cmd = ET.SubElement(child_message, 'ct_parameter', name='cmd_operator', type="bool", direction="out",
                              access_level="public")
    ET.SubElement(child_cmd, 'attribute', type='unit.Server.Attributes.Alarm',
                  value=f'{{"Condition":{{"IsEnabled":"true",'
                        f'"Subconditions":[{{"AckStrategy":0,"IsEnabled":true,'
                        f'"Message":"___",'
                        f'"Severity":10,"Type":2}},'
                        f'{{"AckStrategy":0,"IsEnabled":true,'
                        f'"Message":"___",'
                        f'"Severity":10,"Type":3}}],'
                        f'"Type":2}}}}')
    ET.SubElement(child_cmd, 'attribute', type="unit.System.Attributes.InitialValue", value="false")
    ET.SubElement(child_cmd, 'attribute', type="unit.Server.Attributes.InitialQuality", value="216")

    child_msg = ET.SubElement(child_message, 'ct_parameter', name='msg_operator', type="string", direction="out",
                              access_level="public")
    ET.SubElement(child_msg, 'attribute', type="unit.System.Attributes.InitialValue", value="Действие оператора")
    ET.SubElement(child_msg, 'attribute', type="unit.Server.Attributes.InitialQuality", value="216")

    child_handler = ET.SubElement(child_message, 'ct_handler', name='Handler',
                                  source_code="cmd_operator.Messages.Selected.Message = msg_operator.Value;")
    ET.SubElement(child_handler, 'ct_trigger', on='cmd_operator', cause="message-prepare")

    child_objname = ET.SubElement(child_sys, 'ct_parameter', name='ObjNam', type="string", direction="out",
                                  access_level="public")
    ET.SubElement(child_objname, 'attribute', type="unit.System.Attributes.InitialValue",
                  value=f"{';'.join([obj[1] for obj in sl_object_all])}")
    ET.SubElement(child_objname, 'attribute', type="unit.Server.Attributes.InitialQuality", value="216")
    ET.SubElement(child_objname, 'attribute', type="unit.Server.Attributes.Replicate", value="false")

    child_objalg = ET.SubElement(child_sys, 'ct_parameter', name='ObjAlg', type="string", direction="out",
                                 access_level="public")
    ET.SubElement(child_objalg, 'attribute', type="unit.System.Attributes.InitialValue",
                  value=f"{';'.join([obj[0] for obj in sl_object_all])}")
    ET.SubElement(child_objalg, 'attribute', type="unit.Server.Attributes.InitialQuality", value="216")
    ET.SubElement(child_objalg, 'attribute', type="unit.Server.Attributes.Replicate", value="false")

    child_objnum = ET.SubElement(child_sys, 'ct_parameter', name='ObjNum', type="string", direction="out",
                                 access_level="public")
    ET.SubElement(child_objnum, 'attribute', type="unit.System.Attributes.InitialValue", value=f"{len(sl_object_all)}")
    ET.SubElement(child_objnum, 'attribute', type="unit.Server.Attributes.InitialQuality", value="216")
    ET.SubElement(child_objnum, 'attribute', type="unit.Server.Attributes.Replicate", value="false")

    child_systemname = ET.SubElement(child_sys, 'ct_parameter', name='SystemName', type="string", direction="out",
                                     access_level="public")
    ET.SubElement(child_systemname, 'attribute', type="unit.System.Attributes.InitialValue",
                  value=f"{name_prj}")
    ET.SubElement(child_systemname, 'attribute', type="unit.Server.Attributes.InitialQuality", value="216")
    ET.SubElement(child_systemname, 'attribute', type="unit.Server.Attributes.Replicate", value="false")

    child_ap_activity_switch = ET.SubElement(child_sys, 'ct_parameter', name='AP_activity_switch',
                                             type="bool", direction="out", access_level="public")
    ET.SubElement(child_ap_activity_switch, 'attribute', type="unit.System.Attributes.InitialValue",
                  value="false")
    ET.SubElement(child_ap_activity_switch, 'attribute', type="unit.Server.Attributes.InitialQuality", value="216")

    child_control = ET.SubElement(child_sys, 'ct_object', name='ZControlForAAP', access_level="public")
    child_ready = ET.SubElement(child_control, 'ct_parameter', name='Ready',
                                type="bool", direction="out", access_level="public")
    ET.SubElement(child_ready, 'attribute', type="unit.System.Attributes.InitialValue", value="true")
    ET.SubElement(child_ready, 'attribute', type="unit.Server.Attributes.InitialQuality", value="216")

    for agreg, type_agreg in sl_agreg.items():
        ET.SubElement(child_sys, 'ct_object', name=f'{agreg}', base_type=f"{type_agreg}",
                      aspect="Types.IOS_Aspect", access_level="public")

    if architecture == 'клиент-сервер':
        ET.SubElement(child_sys, 'ct_subject-ref', name=f'_SNMP', object=f"Service.SNMP",
                      const_access="false",
                      aspected="false")
        for par in ('Value', 'Value2'):
            child_value = ET.SubElement(child_sys, 'ct_parameter',
                                        attrib={
                                            'name': f'{par}', 'type': "string",
                                            'direction': "out", 'access-level': "public"
                                        })
            ET.SubElement(child_value, 'attribute', type='unit.Server.Attributes.Alarm',
                          value=f'{{"Condition":{{"IsEnabled":"true",'
                                f'"Subconditions":[{{"AckStrategy":1,"IsEnabled":true,'
                                f'"Message":"Проверка1",'
                                f'"Severity":1,"Type":1}}],"Type":1}}}}'
                          )
            child_handler = ET.SubElement(child_sys, 'ct_handler', name=f'Handler_{par}',
                                          source_code=f"commit {par} = _SNMP.{par};")
            ET.SubElement(child_handler, 'ct_trigger', on=f'_SNMP.{par}', cause="change")

    if 'Система' in return_sl_net:
        child_diagnostic = ET.SubElement(child_sys, 'ct_object', name=f'Diagnostic',
                                         access_level="public")
        # ...добавляем агрегаторы
        for agreg, type_agreg in sl_agreg.items():
            ET.SubElement(child_diagnostic, 'ct_object', name=f'{agreg}', base_type=f"{type_agreg}",
                          aspect="Types.IOS_Aspect", access_level="public")

        ET.SubElement(child_diagnostic, 'ct_subject-ref', name=f'_State', object=f"Service.State",
                      const_access="false",
                      aspected="false")
        for alg, sl_value in return_sl_net.get('Cистема', {}).items():
            for ip_key, port in {'IP': "Порт 1", 'IP_res': "Порт 2"}.items():
                if sl_value.get(ip_key):
                    ET.SubElement(child_diagnostic, 'ct_object', name=f'{alg}_Ping_{ip_key}',
                                  access_level="public", base_type="Types.Ping")
                    child_ping_status = ET.SubElement(child_diagnostic, 'ct_parameter',
                                                      name=f'{alg}_Ping_status_{ip_key}',
                                                      access_level="public", direction='out', type='bool')
                    ET.SubElement(child_ping_status, 'attribute', type="unit.Server.Attributes.Replicate",
                                  value=f"false")
                    rus_name_unit = sl_value.get('Unit', '')
                    ET.SubElement(child_ping_status, 'attribute', type='unit.Server.Attributes.Alarm',
                                  value=f'{{"Condition":{{"IsEnabled":"true",'
                                        f'"Subconditions":[{{"AckStrategy":2,"IsEnabled":true,'
                                        f'"Message":". {rus_name_unit}. {port}. Нет связи",'
                                        f'"Severity":40,"Type":2}},'
                                        f'{{"AckStrategy":2,"IsDeactivation":true,'
                                        f'"Message":". {rus_name_unit}. {port}. Нет связи",'
                                        f'"Severity":40,"Type":3}}],'
                                        f'"Type":2}}}}')

                    object_handler = ET.SubElement(
                        child_checkip, 'ct_handler', name=f'{alg}_On_Ping_Status_{ip_key}',
                        source_code=f'if (_State.Server.Value && {alg}_Ping_{ip_key}.Status.Value '
                                    f'&& !{alg}_Ping_status_{ip_key}.Value) {{\n'
                                    f'\tcommit {alg}_Ping_status_{ip_key} = true;\n'
                                    f'\t}} else if (_State.Server.Value && !{alg}_Ping_{ip_key}.Status.Value '
                                    f'&& {alg}_Ping_status_{ip_key}.Value '
                                    f'|| {alg}_Ping_status_{ip_key}.Quality< 192) {{\n'
                                    f'\tcommit {alg}_Ping_status_{ip_key} = false;\n'
                                    f'}}'
                    )
                    ET.SubElement(object_handler, 'ct_trigger',
                                  on=f"{alg}_Ping_{ip_key}.Status", cause="update")

    # Нормируем и записываем IOS-аспект
    temp = ET.tostring(root_aspect).decode('UTF-8')

    check_diff_file(check_path=os.path.join('File_for_Import', 'IOS_Aspect_in_ApplicationServer'),
                    file_name_check=f'SYS_object(импортируется в приложение сервера).omx-export',
                    new_data=multiple_replace_xml(lxml.etree.tostring(lxml.etree.fromstring(temp),
                                                                      pretty_print=True, encoding='unicode')),
                    message_print=f'Требуется заменить объект SYS (файл SYS_object.omx-export)')
    return


def is_read_btn(sheet):
    # return_sl = {cpu: {алг_пар: (Тип кнопки в студии, русское имя, )}}
    return_sl = {}
    return_sl_diff = {}
    cells = sheet['A1': get_column_letter(50) + str(50)]
    index_alg_name = is_f_ind(cells[0], 'Алгоритмическое имя')
    index_rus_name = is_f_ind(cells[0], 'Наименование параметра')
    index_cpu_name = is_f_ind(cells[0], 'CPU')

    cells = sheet['A2': get_column_letter(50) + str(sheet.max_row)]
    # Составляем множество контроллеров, у которых есть данные параметры
    # set_par_cpu = set()
    for par in cells:
        if par[0].value is None:
            break
        # set_par_cpu.add(par[index_cpu_name].value)
        if par[index_cpu_name].value not in return_sl:
            # return_sl = {cpu: {алг_пар: (Тип кнопки в студии, русское имя, )}}
            return_sl[par[index_cpu_name].value] = {}
        return_sl[par[index_cpu_name].value].update({
            'BTN_' + par[index_alg_name].value[par[index_alg_name].value.find('|')+1:]: (
                'BTN.BTN_PLC_View',
                par[index_rus_name].value)
        })
        # Проверка комментария
        # if par[index_cpu_name].comment:
        #     print(str(par[index_cpu_name].comment).split()[1].split(';'))
        # Обрабатываем и составляем словарь отличий между параметрами по объектам
        if par[index_cpu_name].value not in return_sl_diff:
            # return_sl = {cpu: {алг_пар: (русское имя, ед изм., короткое имя, количество знаков,
            # количество знаков для истории)}}
            return_sl_diff[par[index_cpu_name].value] = {}
        return_sl_diff[par[index_cpu_name].value].update({
            'BTN_' + par[index_alg_name].value[par[index_alg_name].value.find('|')+1:]:
                str(par[index_cpu_name].comment).split()[1].split(';') if par[index_cpu_name].comment else ''})

    return return_sl, return_sl_diff


def is_read_pz(sheet, sl_ai: dict, sl_ae: dict):

    cells = sheet['A1': get_column_letter(50) + str(50)]
    index_alg_name = is_f_ind(cells[0], 'Алгоритмическое имя')
    index_rus_name = is_f_ind(cells[0], 'Наименование параметра')
    index_cpu_name = is_f_ind(cells[0], 'CPU')
    index_type_protect = is_f_ind(cells[0], 'Тип защиты')
    index_unit = is_f_ind(cells[0], 'Единица измерения')
    index_cond = is_f_ind(cells[0], 'Условия защиты')
    index_check_pz = is_f_ind(cells[0], 'Проверяется при ПЗ')

    cells = sheet['A2': get_column_letter(50) + str(sheet.max_row)]

    # Словарь Защит, в котором ключ - cpu, значение - кортеж списков [алг имя, рус. имя, единицы измерения]
    sl_pz = {}
    # Для каждого параметра на листе ...
    for par in cells:
        # Узнаём к какому контроллеру принадлежит параметр
        cpu_name_par = par[index_cpu_name].value
        # Если встретили пустую строку, то прерываем
        if cpu_name_par is None:
            break
        if par[index_type_protect].value not in 'АОссАОбсВОссВОбсАОНО':
            continue
        else:
            # Если в словарь защит нет инфы по контроллеру, то добавляем пустой кортеж
            if cpu_name_par not in sl_pz:
                sl_pz[cpu_name_par] = ()
            # Далее узнаём единицы измерения защиты
            if par[index_unit].value == '-999.0':
                tmp_eunit = str(par[index_unit].comment)[str(par[index_unit].comment).find(' ') + 1:
                                                         str(par[index_unit].comment).find('by')]
                # Для каждого слова в условии защиты ищем первую BND| и определяем по словарям
                # AI, и AE какие единицы измерения писать, на тот случай, если комменты не правдивы
                for word in par[index_cond].value.split():
                    if 'BND|' in word:
                        check = word[word.find('|')+1:word.rfind('_')]
                        if f'AI_{check}' in sl_ai.get(cpu_name_par, {}):
                            tmp_eunit = sl_ai[cpu_name_par][f'AI_{check}']
                        elif f'AE_{check}' in sl_ae.get(cpu_name_par, {}):
                            tmp_eunit = sl_ae[cpu_name_par][f'AE_{check}']
                        break
            else:
                tmp_eunit = par[index_unit].value
            # В словарь Защит соответствующего контроллера добавляем [алг имя, рус. имя, единицы измерения]
            rus_name = f'{par[index_type_protect].value}. {par[index_rus_name].value}'
            sl_pz[cpu_name_par] += ([par[index_alg_name].value, rus_name, tmp_eunit,
                                     f'Проверяется при ПЗ - {par[index_check_pz].value}'],)

    # В словаре защит алгоритмическое имя меняем на A+ номер
    num_pz = 0
    for plc in sl_pz:
        for protect in range(len(sl_pz[plc])):
            sl_pz[plc][protect][0] = 'A' + str(num_pz).zfill(3)
            num_pz += 1

    # Из функции возвращаем словарь, в котором ключ - cpu, значение - кортеж алг. имён A+000 и т.д.
    return_sl_pz = {key: tuple([prot[0] for prot in value]) for key, value in sl_pz.items()}
    # sl_pz_xml = {cpu: {алг_имя(A000): (тип защиты в студии, рус.имя, ед измерения)}}
    sl_pz_xml = {cpu: dict(zip(return_sl_pz[cpu], [('PZ.PZ_PLC_View',) + tuple(val[1:]) for val in value]))
                 for cpu, value in sl_pz.items()}
    return return_sl_pz, sl_pz_xml


def is_read_signals(sheet, sl_wrn_di, sl_cpu_spec: dict):
    return_wrn = {}
    return_ts = {}
    return_ppu = {}
    return_alr = {}
    return_modes = {}
    return_alg = {}
    sl_signal_type = {
        'ТС': return_ts,
        'ТС (без условий)': return_ts,
        'ТС (сообщение)': return_ts,
        'ТС (подрежим)': return_ts,
        'ГР': return_ppu,
        'ХР': return_ppu,
        'АОсс': return_alr,
        'АО': return_alr,
        'НО': return_alr,
        'АОбс': return_alr,
        'ВОсс': return_alr,
        'ВОбс': return_alr,
        'АС (без условий)': return_alr,
        'BOOL': return_alg,
        'INT': return_alg,
        'FLOAT': return_alg,
        'ПС': return_wrn,
        'ПС (без условий)': return_wrn,
        'Режим': return_modes
    }
    sl_type_wrn = {
        'Да (по наличию)': 'WRN_On.WRN_On_PLC_View',
        'Да (по отсутствию)': 'WRN_Off.WRN_Off_PLC_View'
    }
    sl_cpu_archive = {}
    cells = sheet['A1': get_column_letter(50) + str(50)]
    index_alg_name = is_f_ind(cells[0], 'Алгоритмическое имя')
    index_rus_name = is_f_ind(cells[0], 'Наименование параметра')
    index_cpu_name = is_f_ind(cells[0], 'CPU')
    index_type_protect = is_f_ind(cells[0], 'Тип защиты')
    # index_unit = is_f_ind(cells[0], 'Единица измерения')

    cells = sheet['A2': get_column_letter(50) + str(sheet.max_row)]

    for par in cells:
        if par[index_rus_name].value is None:
            break
        # Если тип защиты есть в словаре типовых сигналов:
        protect = par[index_type_protect].value
        if protect in sl_signal_type:
            if par[index_cpu_name].value not in sl_signal_type[protect]:
                sl_signal_type[protect][par[index_cpu_name].value] = {}
            tuple_update = (f"{protect.replace(' (без условий)','')}. {par[index_rus_name].value}"
                            if protect in 'АОссАОбсВОссВОбсАОНО' or 'АС' in protect
                            else par[index_rus_name].value, )
            if protect in ('BOOL', 'INT', 'FLOAT'):
                tuple_update = (f'ALG.ALG_{protect}_PLC_View',) + tuple_update
            elif 'ПС' in protect:
                tuple_update += ('Да (по наличию)',)
            key_update = ('_'.join(par[index_alg_name].value.split('|')) if protect in ('BOOL', 'INT', 'FLOAT')
                          else par[index_alg_name].value[par[index_alg_name].value.find('|') + 1:])
            if 'Режим' in protect:
                sl_signal_type[protect][par[index_cpu_name].value].update(
                    {'regNum': ('MODES.regNum_PLC_View', 'Номер режима'),
                     # 'regName': ('MODES.String_PLC_View', 'Имя режима'),
                     'regColor': ('MODES.regNum_PLC_View', 'Цвет режима'),
                     # 'MsgName': ('MODES.String_PLC_View', 'Состояние'),
                     'MsgNameShow': ('MODES.MODES_NoMsg_PLC_View', 'Показывать состояние'),
                     # 'SubName': ('MODES.String_PLC_View', 'Состояние'),
                     'SubNameShow': ('MODES.MODES_NoMsg_PLC_View', 'Состояние'),
                     })
                tuple_update = ('MODES.MODES_PLC_View', f'Режим "{par[index_rus_name].value}"')
            sl_signal_type[protect][par[index_cpu_name].value].update({key_update: tuple_update})
        if 'Запись архива' in protect:
            if par[index_cpu_name].value not in sl_cpu_archive:
                sl_cpu_archive[par[index_cpu_name].value] = {}
            alg_name = par[index_alg_name].value[par[index_alg_name].value.find('|') + 1:]
            sl_cpu_archive[par[index_cpu_name].value].update({alg_name: f"{par[index_rus_name].value}"})

        '''
        пока оставил данный кусок кода как наследие от старого кода, возможно подскажет     
        elif par[type_protect].value in 'АОссАОбсВОссВОбсАОНО' or 'АС' in par[type_protect].value:
            if 'АС' in par[type_protect].value:
                tmp_alr[par[alg_name].value[par[alg_name].value.find('|') + 1:]] = ('АС. ' +
                                                                                    is_cor_chr(par[par_name].value),
                                                                                    'АС')
            else:
                tmp_alr[par[alg_name].value[par[alg_name].value.find('|') + 1:]] = (par[type_protect].value + '. ' +
                                                                                    is_cor_chr(par[par_name].value),
                                                                                    'Защита')
        '''
    # Есть более изящное решение, если приравнять сразу return_wrn = sl_wrn_di в самом начале, но нужно проверить!!!
    for cpu in return_wrn:
        if cpu in sl_wrn_di:
            return_wrn[cpu].update(sl_wrn_di[cpu])
    return_wrn = {cpu: {alg: (sl_type_wrn.get(val[1]), val[0])
                        for alg, val in value.items()} for cpu, value in return_wrn.items()}

    # пробегаемся по словарям в приведённом ниже списке и добавляем в кортежи параметров соответствующие типы
    # да, немного костыль, но красивый
    for change_return in [(return_ts, 'TS.TS_PLC_View'),
                          (return_ppu, 'PPU.PPU_PLC_View'),
                          (return_alr, 'ALR.ALR_PLC_View')]:
        for cpu, sl_par in change_return[0].items():
            for par in sl_par:
                sl_par[par] = (change_return[1],) + sl_par[par]

    # Добавляем ТС, ППУ, ПС по умолчанию
    for plc, tuple_spec in sl_cpu_spec.items():
        if 'Main' in tuple_spec:
            # if plc not in return_ts:
            #     return_ts.update({plc: {}})
            # return_ts[plc].update({'Unlocking': ('TS.TS_PLC_View', 'Деблокировка')})
            # return_ts[plc].update({'StopGTU': ('TS.TS_PLC_View', 'Разрешена деблокировка защит')})
            # return_ts[plc].update({'ModePZ': ('TS.TS_PLC_View', 'Режим проверки защит')})
            # return_ts[plc].update({'ResetPZ': ('TS.TS_PLC_View', 'Сброс проверки защит')})
            if plc not in sl_cpu_archive:
                sl_cpu_archive.update({plc: {}})
            sl_cpu_archive[plc].update({'UserArh': 'Пользовательский'})
            if plc not in return_alg:
                return_alg.update({plc: {}})
            return_alg[plc].update({'ALG_ExtrCtrlOn': ('ALG.ALG_BOOL_PLC_View', 'Управление от внешней системы')})
            return_alg[plc].update({
                'ALG_ExtrQuery': ('ALG.ALG_BOOL_PLC_View', 'Запрос на управление от внешней системы')
            })
        if 'Main' in tuple_spec and 'PPU' in tuple_spec:
            if plc not in return_ts:
                return_ts.update({plc: {}})
            return_ts[plc].update({'CheckHR': ('TS.TS_PLC_View', 'Контроль ХР')})
            return_ts[plc].update({'CheckGR': ('TS.TS_PLC_View', 'Контроль ГР')})
            if plc not in return_alg:
                return_alg.update({plc: {}})
            return_alg[plc].update({'ALG_ShowPPU': ('ALG.ALG_BOOL_PLC_View', 'Проверка ППУ')})
        if 'PPU' in tuple_spec:
            if plc not in return_ppu:
                return_ppu.update({plc: {}})
            return_ppu[plc].update({
                f'TunNotRead_{plc}': ('PPU.PPU_PLC_View', f'Ошибка чтения настроечных файлов контроллера {plc}')
            })
        if 'Main' in tuple_spec and 'ALR' in tuple_spec:
            if plc not in return_wrn:
                return_wrn.update({plc: {}})
            return_wrn[plc].update({'AllAlrBlk': ('WRN_On.WRN_On_PLC_View', 'Защиты заблокированы')})
            return_wrn[plc].update({'NowAlrBlk': ('WRN_On.WRN_On_PLC_View', 'Сработала заблокированная защита')})
            if plc not in return_ts:
                return_ts.update({plc: {}})
            return_ts[plc].update({'Unlocking': ('TS.TS_PLC_View', 'Деблокировка')})
            return_ts[plc].update({'StopGTU': ('TS.TS_PLC_View', 'Разрешена деблокировка защит')})
            return_ts[plc].update({'ModePZ': ('TS.TS_PLC_View', 'Режим проверки защит')})
            return_ts[plc].update({'ResetPZ': ('TS.TS_PLC_View', 'Сброс проверки защит')})

    # Для каждой записи архива, которую нашли, добавляем две ТСки, которые тянем наверх
    for plc, sl_archive in sl_cpu_archive.items():
        if plc not in return_ts:
            return_ts.update({plc: {}})
        for alg, text_alg in sl_archive.items():
            return_ts[plc].update({
                f'{alg}_START': ('TS.TS_PLC_View', f'Запись архива по событию "{text_alg}" начата')
            })
            return_ts[plc].update({
                f'{alg}_STOP': ('TS.TS_PLC_View', f'Запись архива по событию "{text_alg}" закончена')
            })

    return return_ts, return_ppu, return_alr, return_alg, return_wrn, return_modes, sl_cpu_archive


def is_read_drv(sheet, sl_all_drv):
    sl_color_di = {'FF969696': '0', 'FF00B050': '1', 'FFFFFF00': '2', 'FFFF0000': '3'}
    # sl_type_drv = {
    #     'FLOAT': 'Types.DRV_AI.DRV_AI_PLC_View',
    #     'INT': 'Types.DRV_INT.DRV_INT_PLC_View',
    #     'BOOL': 'Types.DRV_DI.DRV_DI_PLC_View',
    #     'IEC': 'Types.DRV_AI.DRV_AI_PLC_View',
    #     'IECR': 'Types.DRV_AI.DRV_AI_PLC_View',
    #     'IECB': 'Types.DRV_DI.DRV_DI_PLC_View'
    # }
    sl_type_drv = {}
    if os.path.exists(os.path.join('Template_Alpha', 'Systemach', f'dict_type_drv')):
        with open(os.path.join('Template_Alpha', 'Systemach', f'dict_type_drv'), 'r', encoding='UTF-8') as f_signal:
            for line in f_signal:
                if '#' in line:
                    continue
                if not line.strip():
                    break
                line = line.strip()
                lst_line = [i.strip() for i in line.split(':')]
                sl_type_drv[lst_line[0]] = lst_line[1]
    else:
        print(f'Не найден файл сигналов Template_Alpha/Systemach/dict_type_drv, структуры модулей добавлены не будут')

    cells = sheet['A1': get_column_letter(50) + str(50)]

    index_alg_name = is_f_ind(cells[0], 'Алгоритмическое имя')
    index_rus_name = is_f_ind(cells[0], 'Наименование параметра')
    index_cpu_name = is_f_ind(cells[0], 'CPU')
    index_unit = is_f_ind(cells[0], 'Единица измерения')
    index_type_sig = is_f_ind(cells[0], 'Тип')
    index_type_msg = is_f_ind(cells[0], 'Тип сообщения')
    index_color_on = is_f_ind(cells[0], 'Цвет при наличии')
    index_color_off = is_f_ind(cells[0], 'Цвет при отсутствии')
    index_fracdig = is_f_ind(cells[0], 'Число знаков')
    index_drv = is_f_ind(cells[0], 'Драйвер')
    index_save_history = is_f_ind(cells[0], 'Сохранять в истории')

    cells = sheet['A2': get_column_letter(50) + str(sheet.max_row)]
    # return_sl_cpu_drv = {cpu: {(Драйвер, рус имя драйвера):
    # {алг.пар: (Тип переменной в студии, рус имя, тип сообщения, цвет отключения, цвет включения,
    # ед. изм., кол-во знаков, Сохраняемый - Да/Нет, кол-во знаков для истории) }}}
    return_sl_cpu_drv = {}
    # return_ios_drv = {(Драйвер, рус имя драйвера): {cpu:
    # {алг.пар: (Тип переменной в студии, рус имя, тип сообщения, цвет отключения, цвет включения,
    # ед. изм., кол-во знаков) }}}
    return_ios_drv = {}
    # sl_cpu_drv_signal = {cpu: {Драйвер: (кортеж переменных)}}
    sl_cpu_drv_signal = {}
    # Составляем множество контроллеров, у которых есть данные параметры
    # set_par_cpu = set()
    for par in cells:
        if par[0].value is None:
            break
        # Если указанный драйвер есть в словаре объявленных драйверов, то обрабатываем
        if par[index_drv].value in sl_all_drv:
            # set_par_cpu.add(par[index_cpu_name].value)
            # Получаем тип переменной (нужно для удобного отсечения и дальнейшего использования)
            type_sig_par = par[index_type_sig].value  # .replace(' (с имитацией)', '')  # учесть тип имитации!!!

            # Если в словаре sl_cpu_drv_signal нет инфы по cpu, то создаём для него внутренний пустой словарь
            if par[index_cpu_name].value not in sl_cpu_drv_signal:
                sl_cpu_drv_signal[par[index_cpu_name].value] = {}
            # Если в sl_cpu_drv_signal[cpu] нет инфы по драйверу, то создаём для него внутренний кортеж
            if par[index_drv].value not in sl_cpu_drv_signal[par[index_cpu_name].value]:
                sl_cpu_drv_signal[par[index_cpu_name].value][par[index_drv].value] = ()
            sl_cpu_drv_signal[par[index_cpu_name].value][par[index_drv].value] += (par[index_alg_name].value,)

            if par[index_cpu_name].value not in return_sl_cpu_drv:
                return_sl_cpu_drv[par[index_cpu_name].value] = {}

            if (par[index_drv].value, sl_all_drv.get(par[index_drv].value)) \
                    not in return_sl_cpu_drv[par[index_cpu_name].value]:
                return_sl_cpu_drv[par[index_cpu_name].value][(par[index_drv].value,
                                                              sl_all_drv.get(par[index_drv].value))] = {}
            if (par[index_drv].value, sl_all_drv.get(par[index_drv].value)) not in return_ios_drv:
                return_ios_drv[(par[index_drv].value, sl_all_drv.get(par[index_drv].value))] = {}

            if par[index_cpu_name].value \
                    not in return_ios_drv[(par[index_drv].value, sl_all_drv.get(par[index_drv].value))]:
                return_ios_drv[(par[index_drv].value,
                                sl_all_drv.get(par[index_drv].value))][par[index_cpu_name].value] = {}

            type_sig_in_tuple = sl_type_drv.get(type_sig_par, 'Types.').replace('Types.', '')
            type_msg_in_tuple = (par[index_type_msg].value if 'BOOL' in type_sig_par else '-')
            c_off_in_tuple = (sl_color_di.get(par[index_color_off].fill.start_color.index)
                              if 'BOOL' in type_sig_par else '0')
            c_on_in_tuple = (sl_color_di.get(par[index_color_on].fill.start_color.index)
                             if 'BOOL' in type_sig_par else '0')
            unit_in_tuple = (par[index_unit].value if 'BOOL' not in type_sig_par else '-')
            fracdig_in_tuple = (par[index_fracdig].value if 'FLOAT' in type_sig_par
                                                            or 'Расчетные' in type_sig_par
                                                            or 'IECR' in type_sig_par
                                                            or 'IEC' in type_sig_par else '0')
            history_in_tuple = f'Сохраняемый - {par[index_save_history].value}'
            frag_dig_hist_in_tuple = ''.join(
                [i for i in str(par[index_fracdig].comment)[:str(par[index_fracdig].comment).find('by')]
                 if i.isdigit()]) if par[index_fracdig].comment else par[index_fracdig].value

            tuple_par = (type_sig_in_tuple, par[index_rus_name].value, type_msg_in_tuple,
                         c_off_in_tuple, c_on_in_tuple, unit_in_tuple, fracdig_in_tuple, history_in_tuple,
                         frag_dig_hist_in_tuple)

            return_sl_cpu_drv[par[index_cpu_name].value][(par[index_drv].value, sl_all_drv.get(par[index_drv].value))][
                par[index_alg_name].value] = tuple_par

            return_ios_drv[(par[index_drv].value,
                            sl_all_drv.get(par[index_drv].value))][par[index_cpu_name].value][
                par[index_alg_name].value] = tuple_par

    return sl_cpu_drv_signal, return_sl_cpu_drv, return_ios_drv


def is_read_create_grh(sheet, sl_object_all):
    # Считываем файл-шаблон для дополнительных параметров алгоритма (нужен fracdig)

    # Словарь типов ПЛК-аспектов для алгоритмических переменных
    sl_type_alogritm = {
        #  if 'Старт режима' not in val[0] else 'GRH.GRH_BOOL_TS_PLC_View'
        'BOOL_TS': 'Types.GRH.GRH_BOOL_TS_PLC_View',
        'BOOL': 'Types.GRH.GRH_BOOL_PLC_View',
        'INT': 'Types.GRH.GRH_INT_PLC_View',
        'FLOAT': 'Types.GRH.GRH_FLOAT_PLC_View'
    }
    cells = sheet['A1': 'A' + str(sheet.max_row)]

    # словарь по cpu sl_alg_in_cpu = {cpu: sl_algoritm}
    sl_alg_in_cpu = {}
    # словарь команд по cpu sl_command_in_cpu = {cpu: sl_command}
    # sl_command = {(Режим_alg, русское имя режима): {номер шага: {команда_alg: русский текст команды}}}
    sl_command_in_cpu = {}

    # словарь режимов(в том числе и подрежимов),
    # sl_mod_cpu = {cpu: кортеж режимов(в том числе и подрежимов)-(содержит кортеж (алг, рус)}
    sl_mod_cpu = {}

    # Для каждого объекта...
    for objects in sl_object_all:
        # ...для каждого контроллера...
        for cpu in sl_object_all[objects]:
            # ...заполняем общий словарь по контроллерам
            sl_alg_in_cpu[cpu], sl_command_in_cpu[cpu], sl_mod_cpu[cpu] = is_load_algoritm(
                controller=cpu, cells=cells, sheet=sheet)
    # Превращаем в ТС все старты режимов (Возможно потом нужно будет для оттекстовки стартов из-под режимов)
    # print('sl_mod_cpu', sl_mod_cpu)

    # _all_mod = [x[0] for j in [val for _, val in sl_mod_cpu.items()] for x in j]
    # sl_alg_in_cpu = {
    #     cpu: {
    #         par: (tuple_par[0], 'BOOL_TS')
    #         if re.sub(r'_START$', '', re.sub(r"GRH\|.*?_", '', par)) in _all_mod and par.endswith('_START')
    #         else tuple_par for par, tuple_par in sl_alg_tag.items()} for cpu, sl_alg_tag in sl_alg_in_cpu.items()
    # }

    # print('sl_alg_in_cpu', sl_alg_in_cpu)
    # print(_all_mod)
    # for _ in sl_alg_in_cpu:
    #     for par, tt in sl_alg_in_cpu[_].items():
    #         print(par, tt)
    #     print()

    # Из функции возвращаем словарь, где ключ - cpu, а значение - кортеж переменных GRH
    # Также возвращаем словарь вида {cpu: {алг_пар: (тип переменной в студии, русское имя,
    # количество знаков, если есть)}}
    sl_ = dict(zip(sl_alg_in_cpu.keys(), [tuple(value.keys()) for value in sl_alg_in_cpu.values()]))
    # {cpu: {alg_par.replace('GRH|', ''): (
    #         sl_type_alogritm[val[1]].lstrip('Types.'), val[0]) for alg_par, val in value.items()}
    #         for cpu, value in sl_alg_in_cpu.items() if value}
    return {key: value for key, value in sl_.items() if value}, \
           {cpu: {alg_par.split('|')[1]: (sl_type_alogritm.get(val[1]).replace('Types.', ''), val[0]) +
                                         (tuple(val[2]) if len(val) == 3 else tuple())
                  for alg_par, val in value.items()}
            for cpu, value in sl_alg_in_cpu.items() if value}, \
           {key: value for key, value in sl_command_in_cpu.items() if value}, sl_mod_cpu


def check_diff_file(check_path, file_name_check, new_data, message_print):
    # Если в целевой(указанной) папке уже есть формируемый файл
    if os.path.exists(os.path.join(check_path, file_name_check)):
        # считываем имеющейся файл
        with open(os.path.join(check_path, file_name_check), 'r', encoding='UTF-8') as f_check:
            old_data = f_check.read()
        # Если отличаются
        if new_data != old_data:
            # Если нет папки Old, то создаём её
            if not os.path.exists(os.path.join(check_path, 'Old')):
                os.mkdir(os.path.join(check_path, 'Old'))
            # Переносим старую файл в папку Old
            os.replace(os.path.join(check_path, file_name_check),
                       os.path.join(check_path, 'Old', file_name_check))
            sleep(0.5)
            # Записываем новый файл
            with open(os.path.join(check_path, file_name_check), 'w', encoding='UTF-8') as f_wr:
                f_wr.write(new_data)
            # пишем, что надо заменить
            print(message_print)
            with open('Required_change.txt', 'a', encoding='UTF-8') as f_change:
                f_change.write(f'{datetime.datetime.now()} - {message_print}\n')
    # Если в целевой(указанной) папке нет формируемого файла, то создаём его и пишем, что заменить
    else:
        with open(os.path.join(check_path, file_name_check), 'w', encoding='UTF-8') as f_wr:
            f_wr.write(new_data)
        print(message_print)
        with open('Required_change.txt', 'a', encoding='UTF-8') as f_change:
            f_change.write(f'{datetime.datetime.now()} - {message_print}\n')


# Функция для индексов
# Функция получения алгоритмического имени в нижнем регистре (вместе с разделителем |, если такой есть)
# из строки в словаре Трея
def get_variable_lower(line):
    line = line.split(',')
    if isinstance(line[0], str):
        var = ''.join([i for i in line[0] if i not in '#'])
        return var.lower()
    else:
        return 'Странная переменная'


# Функция для индексов
# Функция получения алгоритмического имени (вместе с разделителем |, если такой есть)
# из строки в словаре Трея
def get_variable(line):
    line = line.split(',')
    if isinstance(line[0], str):
        var = ''.join([i for i in line[0] if i not in '#'])
        return var
    else:
        return 'Странная переменная'


# Функция для замены нескольких значений
def multiple_replace(target_str):
    replace_values = {'\n': ' ', '_x000D_': ''}
    # получаем заменяемое: подставляемое из словаря в цикле
    for i, j in replace_values.items():
        # меняем все target_str на подставляемое
        target_str = target_str.replace(i, j)
    return target_str


# функция для поиска индекса нужного столбца
def is_f_ind(cell, name_col, target_count=1):
    count = 1
    for i in range(len(cell)):
        if cell[i].value is None:
            continue
        if multiple_replace(cell[i].value) == name_col:
            if count == target_count:
                return i
            else:
                count += 1
    return 0


def f_ind_json(target_str):
    replace_values = {'\"': '', '\'': ''}
    # получаем заменяемое: подставляемое из словаря в цикле
    for i, j in replace_values.items():
        # меняем все target_str на подставляемое
        target_str = target_str.replace(i, j)
    return target_str


def is_create_rlock(sl_object_all: dict, sl_cpu_archive: dict):
    # print(sl_object_all, '\n', sl_cpu_archive)
    root_aspect = ET.Element('omx', xmlns="system", xmlns_dp="automation.deployment",
                             xmlns_snmp="automation.snmp", xmlns_ct="automation.control")

    child_obj_rlock = ET.SubElement(root_aspect, 'ct_object', name="Service.rlock",
                                    access_level="public")
    ET.SubElement(child_obj_rlock, 'ct_parameter', name='setAlarm', type='bool', direction='in', access_level="public")
    ET.SubElement(child_obj_rlock, 'ct_parameter', name='setDescription', type='string',
                  direction='in', access_level="public")
    ET.SubElement(child_obj_rlock, 'ct_parameter', name='High', type='uint64', direction='in', access_level="public")
    ET.SubElement(child_obj_rlock, 'ct_parameter', name='Low', type='uint64', direction='in', access_level="public")
    ET.SubElement(child_obj_rlock, 'ct_parameter', name='setUserAlarm', type='bool',
                  direction='in', access_level="public")
    # Для каждого объекта
    for objects, sl_plc in sl_object_all.items():
        for plc in sl_plc:
            if plc in sl_cpu_archive:
                for signal_arch, rus_name_arch in sl_cpu_archive[plc].items():
                    ET.SubElement(child_obj_rlock, 'ct_parameter', name=f'Low_{signal_arch}{objects[0]}_tmp',
                                  type='uint64', direction='in',
                                  access_level="public")
                    ET.SubElement(child_obj_rlock, 'ct_parameter', name=f'High_{signal_arch}{objects[0]}_tmp',
                                  type='uint64', direction='in',
                                  access_level="public")
                    ET.SubElement(child_obj_rlock, 'ct_subject-ref', name=f'{objects[0]}_{signal_arch}_START',
                                  object=f"{objects[0]}.System.TS.{signal_arch}_START", const_access="false",
                                  aspected="false")
                    ET.SubElement(child_obj_rlock, 'ct_subject-ref', name=f'{objects[0]}_{signal_arch}_STOP',
                                  object=f"{objects[0]}.System.TS.{signal_arch}_STOP", const_access="false",
                                  aspected="false")
                    object_handler_start = ET.SubElement(
                        child_obj_rlock, 'ct_handler', name=f'{signal_arch}_Handler_START_{objects[0]}',
                        source_code=f'if ({objects[0]}_{signal_arch}_START.PLCSignals.Value)\n{{\n'
                                    f'\tcommit Low_{signal_arch}{objects[0]}_tmp = '
                                    f'Variant.ToUint8(DateTime.UtcNow() - 18000000000,0);'
                                    f'//Время до начала события - по + 18000000000,0 - 30 минут\n}}'
                    )
                    ET.SubElement(object_handler_start, 'ct_trigger',
                                  on=f"{objects[0]}_{signal_arch}_START.PLCSignals.Value", cause="change")
                    object_handler_stop = ET.SubElement(
                        child_obj_rlock, 'ct_handler', name=f'{signal_arch}_Handler_STOP_{objects[0]}',
                        source_code=f'if ({objects[0]}_{signal_arch}_STOP.PLCSignals.Value)\n{{\n'
                                    f'\tcommit setDescription.Value = \"{objects[1]} {rus_name_arch}\";\n'
                                    f'\tcommit High_{signal_arch}{objects[0]}_tmp = '
                                    f'Variant.ToUint8(DateTime.UtcNow() + 18000000000,0);'
                                    f'//Время после события - по + 18000000000,0 - 30 минут\n'
                                    f'\tcommit Low = Low_{signal_arch}{objects[0]}_tmp;\n'
                                    f'\tcommit High = High_{signal_arch}{objects[0]}_tmp;\n'
                                    f'\tcommit setAlarm.Value = true;\n'
                                    f'\tcommit setAlarm.Value = false;\n}}'
                    )
                    ET.SubElement(object_handler_stop, 'ct_trigger',
                                  on=f"{objects[0]}_{signal_arch}_STOP.PLCSignals.Value", cause="change")
    # Общий хендлер для архивов пользователя
    object_handler_user = ET.SubElement(
        child_obj_rlock, 'ct_handler', name=f'Hanlder_User_one',
        source_code=f'if (setUserAlarm.Value)\n{{\n'
                    f'\tcommit High = Variant.ToUint8(DateTime.UtcNow() + 18000000000,0);'
                    f'//Время после начала события - по + 18000000000,0 - 30 минут\n'
                    f'\tcommit Low = Variant.ToUint8(DateTime.UtcNow() - 18000000000,0);'
                    f'//Время до начала события - по + 18000000000,0 - 30 минут\n'
                    f'\tcommit setAlarm.Value = true;\n'
                    f'\tcommit setAlarm.Value = false;\n}}'
    )
    ET.SubElement(object_handler_user, 'ct_trigger', on=f"setUserAlarm", cause="change")

    # Нормируем и записываем IOS-аспект
    temp = ET.tostring(root_aspect).decode('UTF-8')

    check_diff_file(check_path=os.path.join('File_for_Import', 'IOS_Aspect_in_ApplicationServer'),
                    file_name_check=f'rlock_object(импортируется в приложение сервера).omx-export',
                    new_data=multiple_replace_xml(lxml.etree.tostring(lxml.etree.fromstring(temp),
                                                                      pretty_print=True, encoding='unicode')),
                    message_print=f'Требуется заменить IOS-Аспект архивов по событиям (механизм rlock). '
                                  f'Файл rlock_object.omx-export')

    root_aspect = ET.Element('omx', xmlns="system", xmlns_dp="automation.deployment",
                             xmlns_snmp="automation.snmp", xmlns_ct="automation.control")

    child_interval_lock = ET.SubElement(root_aspect, 'ct_object', name="IntervalLock",
                                        access_level="public")
    obj_high = ET.SubElement(child_interval_lock, 'ct_parameter', name='High', type='uint64',
                             direction='out', access_level="public")
    ET.SubElement(obj_high, 'attribute', type='unit.Server.Attributes.History',
                  value=f"Enable=\"True\" ServerTime=\"False\"")

    obj_low = ET.SubElement(child_interval_lock, 'ct_parameter', name='Low', type='uint64',
                            direction='out', access_level="public")
    ET.SubElement(obj_low, 'attribute', type='unit.Server.Attributes.History',
                  value=f"Enable=\"True\" ServerTime=\"False\"")
    obj_write_lock = ET.SubElement(child_interval_lock, 'ct_parameter', name='WriteLock', type='bool',
                                   direction='out', access_level="public")
    ET.SubElement(obj_write_lock, 'attribute', type='unit.Server.Attributes.History',
                  value=f"Enable=\"True\" ServerTime=\"False\"")
    ET.SubElement(obj_write_lock, 'attribute', type='unit.Server.Attributes.Alarm',
                  value=f'{{"Condition":{{"IsEnabled":"true",'
                        f'"Subconditions":[{{"AckStrategy":1,"IsEnabled":true,'
                        f'"Message":"Архивация выполнена",'
                        f'"Severity":500,"Type":2}},'
                        f'{{"AckStrategy":1,"IsEnabled":false,'
                        f'"Message":"___",'
                        f'"Severity":500,"Type":3}}],'
                        f'"Type":2}}}}')
    ET.SubElement(child_interval_lock, 'ct_parameter', name='Description', type='string',
                  direction='out', access_level="public")
    ET.SubElement(child_interval_lock, 'ct_subject-ref', name=f'rlock', object=f"rlock", const_access="false",
                  aspected="false")
    ET.SubElement(child_interval_lock, 'ct_subject-ref', name=f'_State', object=f"State", const_access="false",
                  aspected="false")

    object_handler = ET.SubElement(
        child_interval_lock, 'ct_handler', name=f'Handler',
        source_code=f'if (rlock.setAlarm.Value && _State.Server.Value) {{\n'
                    f'\tcommit Description = rlock.setDescription;\n'
                    f'\tcommit High = rlock.High;\n'
                    f'\tcommit Low = rlock.Low;\n'
                    f'\tcommit WriteLock = true;\n}}'
    )
    ET.SubElement(object_handler, 'ct_trigger',
                  on=f"rlock.setAlarm", cause="update")

    # Нормируем и записываем IOS-аспект
    temp = ET.tostring(root_aspect).decode('UTF-8')

    check_diff_file(check_path=os.path.join('File_for_Import', 'IOS_Aspect_in_ApplicationServer'),
                    file_name_check=f'IntervalLock(импортируется в структуру Historian).omx-export',
                    new_data=multiple_replace_xml(lxml.etree.tostring(lxml.etree.fromstring(temp),
                                                                      pretty_print=True, encoding='unicode')),
                    message_print=f'Требуется заменить IOS-Аспект архивов по событиям в модуле истории(механизм rlock).'
                                  f'Файл IntervalLock.omx-export')


def is_create_service_snmp(server_name_rez: str):
    root_aspect = ET.Element('omx', xmlns="system", xmlns_dp="automation.deployment",
                             xmlns_snmp="automation.snmp", xmlns_ct="automation.control")
    # Добавляем SNMP
    child_snmp = ET.SubElement(root_aspect, 'ct_object', name='Service.SNMP', access_level="public")
    ET.SubElement(child_snmp, 'attribute', type='unit.Server.Attributes.Replicate', value=f"false")
    for comp, par in {
        'Computer1': ('SNMP_Agent.Application.SNMP',),
        'Computer2': (f'{server_name_rez}.SNMP_Agent.Application.SNMP',),
        'ComputerLocal': ('SNMP_AgentLocal.Application.SNMP',)
    }.items():
        child_comp = ET.SubElement(child_snmp, 'ct_object', name=f'{comp}', base_type="Types.SNMP.Server_IOS_View",
                                   access_level="public", aspect="Types.IOS_Aspect",
                                   original=f"{par[0]}")
        ET.SubElement(child_comp, 'ct_init-ref', ref="_Server_PLC", target=f"{par[0]}")

    ET.SubElement(child_snmp, 'ct_subject-ref', name=f'ConnectionStatus', object=f"Redundancy.Channel1",
                  const_access="false",
                  aspected="false")
    ET.SubElement(child_snmp, 'ct_subject-ref', name=f'ConnectionStatus2', object=f"Redundancy.Channel2",
                  const_access="false", aspected="false")
    ET.SubElement(child_snmp, 'ct_parameter', name='prevConnect1',
                  type='bool', direction='out', access_level="public")
    ET.SubElement(child_snmp, 'ct_parameter', name='prevConnect2',
                  type='bool', direction='out', access_level="public")
    ET.SubElement(child_snmp, 'ct_parameter', name='tmpValue',
                  type='string', direction='out', access_level="public")
    ET.SubElement(child_snmp, 'ct_parameter', name='tmpValue2',
                  type='string', direction='out', access_level="public")
    child_counter = ET.SubElement(child_snmp, 'ct_parameter', name='Counter',
                                  type='uint64', direction='out', access_level="public")
    ET.SubElement(child_counter, 'attribute', type='unit.Server.Attributes.Replicate', value=f"true")
    ET.SubElement(child_counter, 'attribute', type='unit.System.Attributes.InitialValue', value=f"1000")
    ET.SubElement(child_counter, 'attribute', type='unit.Server.Attributes.InitialQuality', value=f"192")
    ET.SubElement(child_snmp, 'ct_parameter', name='Value',
                  type='string', direction='out', access_level="public")
    ET.SubElement(child_snmp, 'ct_parameter', name='Value2',
                  type='string', direction='out', access_level="public")
    ET.SubElement(child_snmp, 'ct_timer', name='Timer', period='100')
    ET.SubElement(child_snmp, 'ct_subject-ref', name=f'State', object=f"State", const_access="false",
                  aspected="false")

    object_handler_connectionlost = ET.SubElement(
        child_snmp, 'ct_handler', name=f'Handler_ConnectionLost',
        source_code=f'if (ConnectionStatus.ConnectionEstablished == false && '
                    f'ConnectionStatus2.ConnectionEstablished == false)\n'
                    f'{{\n'
                    f'	localName: string = ComputerLocal.computerName.Value;\n'
                    f'	remoteName: string = "";	\n'
                    f'	if (localName == Computer1.computerName.Value) \n'
                    f'	{{\n'
                    f'		remoteName = Computer2.computerName.Value;\n'
                    f'	}} else \n'
                    f'	{{\n'
                    f'		remoteName = Computer1.computerName.Value;\n'
                    f'	}}\n'
                    f'	tmpValue2 = "";\n'
                    f'	tmpValue = "<Subcondition Type=\\"Dynamic\\" Enabled=\\"1\\" SoundEnabled=\\"0\\" '
                    f'AckRequired =\\"1\\" Sound=\\"\\" Message=\\". На сервере '" + localName +"' '
                    'отсутствует связь с резервным сервером '"+ remoteName +"'.\\" Severity=\\"40\\"/>";\n'
                    f'}}\n'
                    f'else\n'
                    f'{{\n'
                    f'	tmpValue = "<Subcondition Type=\\"Dynamic\\" Enabled=\\"0\\" />";\n\n}}\n'
                    f'if (ConnectionStatus.ConnectionEstablished != prevConnect1 && '
                    f'ConnectionStatus2.ConnectionEstablished != prevConnect2) {{\n'
                    f'	Counter = 0;\n'
                    f'	prevConnect1 = ConnectionStatus.ConnectionEstablished;\n'
                    f'	prevConnect2 = ConnectionStatus2.ConnectionEstablished;\n'
                    f'}}\n'
    )
    ET.SubElement(object_handler_connectionlost, 'ct_trigger',
                  on=f"ConnectionStatus.ConnectionEstablished", cause="update")
    ET.SubElement(object_handler_connectionlost, 'ct_trigger',
                  on=f"ConnectionStatus2.ConnectionEstablished", cause="update")

    object_handler_switchactive = ET.SubElement(
        child_snmp, 'ct_handler', name=f'Handler_SwitchActive',
        source_code=f'if (State.Server && (ConnectionStatus.ConnectionEstablished || '
                    f'ConnectionStatus2.ConnectionEstablished))\n'
                    f'{{\n'
                    f'	localName: string = ComputerLocal.computerName.Value;\n'
                    f'  remoteName: string = "";\n'
                    f'  if (localName == Computer1.computerName.Value)\n'
                    f'{{\n'
                    f'		remoteName = Computer2.computerName.Value;\n'
                    f'	}} else \n'
                    f'	{{\n'
                    f'		remoteName = Computer1.computerName.Value;\n'
                    f'	}};\n'
                    f'	Counter = 0;\n\n'
                    f'	tmpValue2 = "<Subcondition Type=\\"Dynamic\\" Enabled=\\"1\\" SoundEnabled=\\"0\\"  '
                    f'AckRequired =\\"1\\" Sound=\\"\\" Message=\\". '
                    f'Произошел резервный переход. Активный сервер '" + localName +"'\\" Severity=\\"40\\"/>";\n'
                    f'}}\n'

    )
    ET.SubElement(object_handler_switchactive, 'ct_trigger',
                  on=f"State.Server", cause="update")

    object_handler_switchred = ET.SubElement(
        child_snmp, 'ct_handler', name=f'Handler_SwitchRed',
        source_code=f'if (State.Server && (ConnectionStatus.ConnectionEstablished || '
                    f'ConnectionStatus2.ConnectionEstablished))\n'
                    f'{{\n'
                    f'	localName: string = ComputerLocal.computerName.Value;\n'
                    f'  remoteName: string = "";\n'
                    f'  if (localName == Computer1.computerName.Value)\n'
                    f'{{\n'
                    f'		remoteName = Computer2.computerName.Value;\n'
                    f'	}} else \n'
                    f'	{{\n'
                    f'		remoteName = Computer1.computerName.Value;\n'
                    f'	}};\n'
                    f'	Counter = 0;\n\n'
                    f'	tmpValue = "<Subcondition Type=\\"Dynamic\\" Enabled=\\"1\\" SoundEnabled=\\"0\\"  '
                    f'AckRequired =\\"1\\" Sound=\\"\\" Message=\\". '
                    f'Произошел резервный переход. Резервный сервер '" + remoteName +"'\\" Severity=\\"40\\"/>";\n'
                    f'}}\n'

    )
    ET.SubElement(object_handler_switchred, 'ct_trigger',
                  on=f"State.Server", cause="update")

    object_handler_1 = ET.SubElement(
        child_snmp, 'ct_handler', name=f'Handler',
        source_code=f'Counter = Counter + 1;\n'
                    f'if (Counter == 20) \n'
                    f'{{\n'
                    f'	if (ConnectionStatus.ConnectionEstablished || ConnectionStatus2.ConnectionEstablished) {{\n'
                    f'		Value = "<Subcondition Type=\\"Dynamic\\" Enabled=\\"0\\" />";\n'
                    f'	Value2 = "<Subcondition Type=\\"Dynamic\\" Enabled=\\"0\\" />";\n'
                    f'	}}\n'
                    f'}};\n\n'
                    f'if (Counter == 21)\n'
                    f'{{\n'
                    f'	Value = tmpValue;\n'
                    f'	Value2 = tmpValue2;\n'
                    f'}};\n\n'
                    f'if (Counter == 70)\n'
                    f'{{\n'
                    f'	if (ConnectionStatus.ConnectionEstablished || ConnectionStatus2.ConnectionEstablished) {{\n'
                    f'		Value = "<Subcondition Type=\\"Dynamic\\" Enabled=\\"0\\" />";\n'
                    f'	Value2 = "<Subcondition Type=\\"Dynamic\\" Enabled=\\"0\\" />";	\n'
                    f'	}}\n}}\n'


    )
    ET.SubElement(object_handler_1, 'ct_trigger',
                  on=f"Timer", cause="update")

    # Пишем ссылку в SYS
    ET.SubElement(root_aspect, 'ct_subject-ref', name=f'SYS._SNMP', object=f"Service.SNMP",
                  const_access="false",
                  aspected="false")
    for par in ('Value', 'Value2'):
        child_value = ET.SubElement(root_aspect, 'ct_parameter',
                                    attrib={
                                        'name': f'SYS.{par}', 'type': "string",
                                        'direction': "out", 'access-level': "public"
                                    })
        ET.SubElement(child_value, 'attribute', type='unit.Server.Attributes.Alarm',
                      value=f'{{"Condition":{{"IsEnabled":"true",'
                            f'"Subconditions":[{{"AckStrategy":1,"IsEnabled":true,'
                            f'"Message":"Проверка1",'
                            f'"Severity":1,"Type":1}}],"Type":1}}}}'
                      )
        child_handler = ET.SubElement(root_aspect, 'ct_handler', name=f'SYS.Handler_{par}',
                                      source_code=f"commit {par} = _SNMP.{par};")
        ET.SubElement(child_handler, 'ct_trigger', on=f'_SNMP.{par}', cause="change")

    # Нормируем и записываем IOS-аспект
    temp = ET.tostring(root_aspect).decode('UTF-8')

    check_diff_file(check_path=os.path.join('File_for_Import', 'IOS_Aspect_in_ApplicationServer'),
                    file_name_check=f'file_service_snmp.omx-export',
                    new_data=multiple_replace_xml(lxml.etree.tostring(lxml.etree.fromstring(temp),
                                                                      pretty_print=True, encoding='unicode')),
                    message_print=f'Требуется заменить IOS-Аспект SNMP (сообщения резервирования).'
                                  f'Файл file_service_snmp.omx-export')
    return


def is_read_out(sheet, type_signal):
    # return_sl = {cpu: {алг_пар: (ТИП ALG в студии, русское имя, Нестандартный канал Да/Нет, модуль, канал)}}
    return_sl = {}
    return_sl_diff = {}
    # return_sl_sday = {cpu: {Префикс папки.alg_par: русское имя}}
    # return_sl_sday = {}
    # return_sl_mnemo = {cpu: {узел: список параметров узла}}
    return_sl_mnemo = {}
    sl_plc_aspect = {'CDO': 'In.In_PLC_View'}
    # print('Зашли в функцию, максимальное число колонок - ', sheet.max_column)
    cells = sheet['A1': get_column_letter(100) + str(100)]  # sheet.max_column
    index_alg_name = is_f_ind(cells[0], 'Алгоритмическое имя')
    index_rus_name = is_f_ind(cells[0], 'Наименование параметра')
    # index_unit = is_f_ind(cells[0], 'Единицы измерения')
    # index_short_name = is_f_ind(cells[0], 'Короткое наименование')
    # index_frag_dig = is_f_ind(cells[0], 'Количество знаков')
    index_cpu_name = is_f_ind(cells[0], 'CPU')
    index_res = is_f_ind(cells[0], 'Резервный')
    # index_node_mnemo = is_f_ind(cells[0], 'Узел')
    # index_sday = is_f_ind(cells[0], 'Используется в ведомости')
    index_no_standart = is_f_ind(cells[0], 'Нестандартный канал')
    index_num_module = is_f_ind(cells[0], 'Номер модуля')
    index_num_canal = is_f_ind(cells[0], 'Номер канала')

    cells = sheet['A2': get_column_letter(100) + str(sheet.max_row)]
    # Составляем множество контроллеров, у которых есть данные параметры
    # set_par_cpu = set()
    for par in cells:
        if par[0].value is None:
            break
        # set_par_cpu.add(par[index_cpu_name].value)
        # Проверка комментария
        # if par[index_cpu_name].comment:
        #     print(str(par[index_cpu_name].comment).split()[1].split(';'))
        if par[index_res].value == 'Нет':
            if par[index_cpu_name].value not in return_sl:
                # return_sl = {cpu: {алг_пар: (русское имя, ед изм., короткое имя, количество знаков,
                # количество знаков для истории)}}
                return_sl[par[index_cpu_name].value] = {}

            # Обрабатываем и составляем словарь отличий между параметрами по объектам
            if par[index_cpu_name].value not in return_sl_diff:
                # return_sl = {cpu: {алг_пар: (русское имя, ед изм., короткое имя, количество знаков,
                # количество знаков для истории)}}
                return_sl_diff[par[index_cpu_name].value] = {}
            return_sl_diff[par[index_cpu_name].value].update({
                par[index_alg_name].value.replace('|', '_'):
                    str(par[index_cpu_name].comment).split()[1].split(';') if par[index_cpu_name].comment else ''})

            # if par[index_sday].value == 'Да':
            #     if par[index_cpu_name].value not in return_sl_sday:
            #         return_sl_sday[par[index_cpu_name].value] = {}
            #     return_sl_sday[par[index_cpu_name].value].update(
            #         {f"{type_signal}.{par[index_alg_name].value.replace('|', '_')}": par[index_rus_name].value})

            # frag_dig_hist = ''.join([i for i
            #                          in str(par[index_frag_dig].comment)[:str(par[index_frag_dig].comment).find('by')]
            #                          if i.isdigit()]) if par[index_frag_dig].comment else par[index_frag_dig].value

            return_sl[par[index_cpu_name].value].update({par[index_alg_name].value.replace('DO|', ''): (
                sl_plc_aspect.get(type_signal, 'Пустой тип'),
                par[index_rus_name].value,
                f'{par[index_num_module].value}.Canal_{par[index_num_canal].value}'
                if par[index_no_standart].value == 'Нет' else 'Нестандартный канал'
                # par[index_no_standart].value,
                # par[index_num_module].value,
                # par[index_num_canal].value
            )})
            # Если не парсим уставки, то заполняем словарь для мнемосхемы
            if 'SP|' not in par[index_alg_name].value and par[index_cpu_name].value not in return_sl_mnemo:
                return_sl_mnemo[par[index_cpu_name].value] = {}
            if 'SP|' not in par[index_alg_name].value and \
                    'Проверка выходных сигналов' not in return_sl_mnemo[par[index_cpu_name].value]:
                return_sl_mnemo[par[index_cpu_name].value]['Проверка выходных сигналов'] = list()
            if 'SP|' not in par[index_alg_name].value:
                return_sl_mnemo[par[index_cpu_name].value]['Проверка выходных сигналов'].append(
                    par[index_alg_name].value.replace('DO|', '')
                )

    return return_sl, return_sl_mnemo, return_sl_diff


def is_read_pru(sheet):
    # return_sl = {cpu: {алг_пар: (Тип параметра в студии, русское имя, )}}
    return_sl = {}
    return_sl_diff = {}
    cells = sheet['A1': get_column_letter(50) + str(50)]
    index_alg_name = is_f_ind(cells[0], 'Алгоритмическое имя')
    index_rus_name = is_f_ind(cells[0], 'Наименование параметра')
    index_cpu_name = is_f_ind(cells[0], 'CPU')
    index_type = is_f_ind(cells[0], 'Тип')
    index_direction = is_f_ind(cells[0], 'Направление')

    cells = sheet['A2': get_column_letter(50) + str(sheet.max_row)]
    # Составляем множество контроллеров, у которых есть данные параметры
    # set_par_cpu = set()
    for par in cells:
        if par[0].value is None:
            break
        if par[index_type].value != 'BOOL' or par[index_direction].value != 'Команда от ПРУ':
            continue
        # set_par_cpu.add(par[index_cpu_name].value)
        if par[index_cpu_name].value not in return_sl:
            # return_sl = {cpu: {алг_пар: (Тип кнопки в студии, русское имя, )}}
            return_sl[par[index_cpu_name].value] = {}
        return_sl[par[index_cpu_name].value].update({
            par[index_alg_name].value.replace('|', '_', 1): (
                'PRU.PRU_PLC_View',
                f'Команда от ПРУ: {par[index_rus_name].value}'
            )
        })
        # Проверка комментария
        # if par[index_cpu_name].comment:
        #     print(str(par[index_cpu_name].comment).split()[1].split(';'))
        # Обрабатываем и составляем словарь отличий между параметрами по объектам
        # if par[index_cpu_name].value not in return_sl_diff:
        #     # return_sl = {cpu: {алг_пар: (русское имя, ед изм., короткое имя, количество знаков,
        #     # количество знаков для истории)}}
        #     return_sl_diff[par[index_cpu_name].value] = {}
        # return_sl_diff[par[index_cpu_name].value].update({
        #     'BTN_' + par[index_alg_name].value[par[index_alg_name].value.find('|')+1:]:
        #         str(par[index_cpu_name].comment).split()[1].split(';') if par[index_cpu_name].comment else ''})

    return return_sl
