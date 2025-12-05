import openpyxl
import os
import sys
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.styles.borders import Border, Side
# from openpyxl.worksheet.header_footer import HeaderFooter
from openpyxl.workbook.defined_name import DefinedName


def create_reports_sday_old_format(sl_object_all: dict, node_param_rus: str, sl_param: dict):
    # sl_object_all =
    # { (Объект, рус имя объекта, индекс объекта): {контроллер: (ip основной, ip резервный, индекс объекта)} }
    # sl_param = {cpu: {Префикс папки.alg_par: русское имя}}
    # num_cel = 8
    # num_par = 1

    thin_border = Border(left=Side(style='thin'),  # thick - толстые границы, thin - тонкие
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    tuple_sh = ('W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH')
    str_sh = 'IJKLMNOPQRST'

    # Для каждого объекта
    for obj in sl_object_all:
        num_cel = 8
        num_par = 1
        # Создаём эксельку
        wb = openpyxl.Workbook()
        # Выбираем активный лист
        sheet = wb.active
        sheet.title = 'Лист 1'
        # Заполняем стартовую информацию
        for col, w in {'F': 7, 'G': 105, 'H': 15}.items():
            sheet.column_dimensions[col].width = w
        sheet['E2'] = f'="{obj[0]}."'
        for cell, value in {'A2': '=NOW()', 'B2': '="№ ГПА"', 'A3': '=reportdate', 'A4': '=smena', 'A5': '="smena=1"',
                            'A6': '="smena=2"', 'B3': '="дата"', 'B4': '="смена"', 'B5': '="дневная"',
                            'B6': '="ночная"', 'C7': '="узел"', 'D7': '="ед.изм"', 'E7': '="значение"',
                            'D3': '="начало пути"', 'E3': '="Получение данных с OPC UA@"',
                            'D4': '="значение"', 'E4': '=".Value;1"', 'E5': '=".Value.100;1"',
                            'A9': '=YEAR(A3)', 'B9': '="год"', 'A10': '=MONTH(A3)', 'B10': '="месяц"',
                            'A11': '=DAY(A3)', 'B11': '="день"', 'A12': '=DATE(A9, A10, A11)',
                            'A14': '=IF(A4=1, TIME(8, 0, 0), TIME(20, 0, 0))', 'A15': '=TIME(0, 3, 0)',
                            'B14': '="выбор смены"', 'B15': '="диапазон"'}.items():
            sheet[cell] = value
        # Заполняем временные формулы
        for cell, value in {'I5': '=A12+A14', 'I6': '=I5+$A$15'}.items():
            sheet[cell] = value
        cell_ind_str = 'JKLMNOPQRST'
        formul_ind_str = 'IJKLMNOPQRS'
        for i in range(len(cell_ind_str)):
            sheet[f'{cell_ind_str[i]}5'] = f'={formul_ind_str[i]}5+1/24'
            sheet[f'{cell_ind_str[i]}6'] = f'={cell_ind_str[i]}5+$A$15'
        for coll in 'IJKLMNOPQRST':
            sheet[f'{coll}4'] = f'=IF($A$2>{coll}5, true, false)'
        # Выставляем форматы
        for cell, value in {'A2': 22, 'A3': 22, 'A12': 22, 'A14': 21, 'A15': 21}.items():
            sheet[cell].number_format = numbers.BUILTIN_FORMATS[value]
        for coll in 'IJKLMNOPQRST':
            sheet[f'{coll}5'].number_format = numbers.BUILTIN_FORMATS[22]
            sheet[f'{coll}6'].number_format = numbers.BUILTIN_FORMATS[22]
        # Заполняем шапку формул данных
        sheet['V5'] = '="текущие значения"'
        for i in range(len(tuple_sh)):
            sheet[f'{tuple_sh[i]}5'] = f'={str_sh[i]}7'
            sheet[f'{tuple_sh[i]}5'].number_format = numbers.BUILTIN_FORMATS[22]
        # Создаём заголовок
        sheet.row_dimensions[2].height = 20
        sheet.row_dimensions[1].height = 25
        sheet[f'G2'] = f'="Сменная ведомость {obj[1]} на "'
        sheet[f'G2'].font = Font(size=14, name='Arial', bold=True)
        sheet[f'G2'].alignment = Alignment(horizontal="right", vertical="top")
        sheet[f'H2'] = f'=A3'
        sheet[f'H2'].font = Font(size=14, name='Arial', bold=True)
        sheet[f'H2'].alignment = Alignment(horizontal="center", vertical="top")
        # sheet[f'H2'].number_format = numbers.BUILTIN_FORMATS[15]
        sheet[f'H2'].number_format = 'dd.mm.yyyy'

        for cell, i in {'F7': '="№"', 'G7': '="Наименование  "',
                        'H7': '="ед.изм"'}.items():
            sheet[cell] = i
            sheet[cell].font = Font(size=12, name='Arial', bold=True)
            sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
            sheet[cell].border = thin_border
        # Проставляем время каждый час в первой шапке
        for coll in str_sh:
            sheet[f'{coll}7'] = f'=TIME(HOUR({coll}5), 0, 0)'
            sheet[f'{coll}7'].font = Font(size=12, name='Arial', bold=True)
            sheet[f'{coll}7'].alignment = Alignment(horizontal="center", vertical="center")
            sheet[f'{coll}7'].border = thin_border
            sheet[f'{coll}7'].number_format = numbers.BUILTIN_FORMATS[20]
            sheet.column_dimensions[coll].width = 10

        # Для каждого контроллера...
        for cpu, sl_par in sl_param.items():
            # ...при условии, что объект содержит текущий контроллер...
            if cpu in sl_object_all[obj]:
                # ...для каждого параметра...
                for par, property_par in sl_par.items():
                    # Если отсчитали 17 параметров, то добавляем шапку для следующего листа
                    if not (num_par - 1) % 33 and num_par != 1:
                        num_cel += 2
                        for _ in range(3):
                            sheet.row_dimensions[num_cel].height = 35
                            for cell, i in {'G': '="должность"', 'J': '="ФИО"', 'M': '="подпись"'}.items():
                                sheet[f'{cell}{num_cel}'] = i
                                sheet[f'{cell}{num_cel}'].font = Font(size=12, name='Arial')
                                sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
                            for coll in 'GHIJKLMN':
                                sheet[f'{coll}{num_cel}'].border = Border(top=Side(style='thin'))
                            num_cel += 1

                        num_cel += 1
                        sheet.row_dimensions[num_cel].height = 25
                        sheet[f'G{num_cel}'] = f'="Сменная ведомость {obj[1]} на "'
                        sheet[f'G{num_cel}'].font = Font(size=16, name='Arial', bold=True)
                        sheet[f'G{num_cel}'].alignment = Alignment(horizontal="right", vertical="top")
                        sheet[f'H{num_cel}'] = f'=H2'
                        sheet[f'H{num_cel}'].font = Font(size=16, name='Arial', bold=True)
                        sheet[f'H{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
                        # sheet[f'H{num_cel}'].number_format = numbers.BUILTIN_FORMATS[15]
                        sheet[f'H{num_cel}'].number_format = 'dd.mm.yyyy'
                        num_cel += 2
                        sheet.row_dimensions[num_cel].height = 20
                        for cell, i in {'F': '="№"', 'G': '="Наименование  "',
                                        'H': '="ед.изм"'}.items():
                            sheet[f'{cell}{num_cel}'] = i
                            sheet[f'{cell}{num_cel}'].font = Font(size=14, name='Arial', bold=True)
                            sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="center")
                            sheet[f'{cell}{num_cel}'].border = thin_border
                        for coll in 'IJKLMNOPQRST':
                            sheet[f'{coll}{num_cel}'] = f'={coll}7'
                            sheet[f'{coll}{num_cel}'].font = Font(size=14, name='Arial', bold=True)
                            sheet[f'{coll}{num_cel}'].alignment = Alignment(horizontal="center", vertical="center")
                            sheet[f'{coll}{num_cel}'].border = thin_border
                            sheet[f'{coll}{num_cel}'].number_format = numbers.BUILTIN_FORMATS[20]
                        num_cel += 1
                    # Заполняем параметры
                    sheet[f'C{num_cel}'] = f'="{par}"'
                    sheet[f'F{num_cel}'] = f'="{num_par}"'
                    rus_name = property_par.replace("\"", "\"\"")
                    sheet[f'G{num_cel}'] = f'="{rus_name}  "'  # Русское наименование

                    # Заполняем формулы для вывода данных
                    sheet[f'V{num_cel}'] = f'=CurrAttrValue(E{num_cel}, 0)'
                    for i in range(len(tuple_sh)):
                        sheet[f'{tuple_sh[i]}{num_cel}'] = f'=ArchiveAttributeValue($E{num_cel}, ' \
                                                           f'0, {str_sh[i]}$5, {str_sh[i]}$6, 1)'
                        # sheet[f'{tuple_sh[i]}{num_cel}'] = f'=ValueOnDate($E{num_cel}, 0, {str_sh[i]}$5)'
                        # print(11)
                    for i in range(len(str_sh)):
                        sheet[f'{str_sh[i]}{num_cel}'] = f'=IF({str_sh[i]}$4, IF(ISNUMBER({tuple_sh[i]}{num_cel}), ' \
                                                         f'{tuple_sh[i]}{num_cel}, $V{num_cel}), "-")'
                        sheet[f'{str_sh[i]}{num_cel}'].number_format = numbers.BUILTIN_FORMATS[2]
                        sheet[f'{str_sh[i]}{num_cel}'].font = Font(size=14, name='Arial')
                        sheet[f'{str_sh[i]}{num_cel}'].border = thin_border
                    # Заполняем ссылки
                    for i, value in {'D': f'=CONCATENATE($E$3, $E$2, $C{num_cel}, $E$5)',
                                     'E': f'=CONCATENATE($E$3, $E$2, $C{num_cel}, $E$4)'}.items():
                        sheet[f'{i}{num_cel}'] = value
                    # Заполняем формулы
                    for i, value in {'H': f'=CurrAttrValue(D{num_cel}, 0)'}.items():
                        sheet[f'{i}{num_cel}'] = value

                    # Устанавливаем для ячеек размер шрифта 14, Arial и выставляем границу
                    for i in 'FGH':
                        sheet[f'{i}{num_cel}'].font = Font(size=14, name='Arial')
                        sheet[f'{i}{num_cel}'].border = thin_border
                    # Выводим в центр нумерацию, значения и единицы измерения
                    for i, vert_alig in {'F': 'center', 'G': 'left', 'H': 'center', 'S': 'center'}.items():
                        sheet[f'{i}{num_cel}'].alignment = Alignment(horizontal=vert_alig, vertical="center",
                                                                     wrap_text=True)
                    sheet.row_dimensions[num_cel].height = 20
                    num_cel += 1
                    num_par += 1
        # Прячем столбцы с привязками
        for col in ('A', 'B', 'C', 'D', 'E'):
            sheet.column_dimensions[col].hidden = True
        for row in (4, 5, 6):
            sheet.row_dimensions[row].hidden = True
        for col in ('V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK',
                    'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX'):
            sheet.column_dimensions[col].hidden = True
        # Добавляем подпись в конце
        num_cel += 2
        for _ in range(3):
            sheet.row_dimensions[num_cel].height = 35
            for cell, i in {'G': '="должность"', 'J': '="ФИО"', 'M': '="подпись"'}.items():
                sheet[f'{cell}{num_cel}'] = i
                sheet[f'{cell}{num_cel}'].font = Font(size=12, name='Arial')
                sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
            for coll in 'GHIJKLMN':
                sheet[f'{coll}{num_cel}'].border = Border(top=Side(style='thin'))
            num_cel += 1

        wb.save(os.path.join('File_for_Import', 'Reports', f'{node_param_rus} {obj[1]}.xlsx'))
    return


def create_reports_pz(sl_object_all: dict, node_param_rus: str, node_alg_name: str, sl_param: dict,
                      reports_protocol_version: int):
    # sl_object_all =
    # { (Объект, рус имя объекта, индекс объекта): {контроллер: (ip основной, ip резервный, индекс объекта)} }
    # sl_param = {cpu: {алг_имя(A000): (тип защиты в студии, рус.имя, ед измерения)}}
    # num_cel = 4
    # num_par = 1
    igrek = 10
    igrek_delta = 10

    thin_border = Border(left=Side(style='thin'),  # thick - толстые границы, thin - тонкие
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    use_opc_ua = False
    match reports_protocol_version:
        case 0:  # 0 - OPC UA, 1 - Alpha.Link
            use_opc_ua = True
        case 1:  # 0 - OPC UA, 1 - Alpha.Link
            use_opc_ua = False
        case _:
            use_opc_ua = True
    # sheet['B1'] = '=".Value;1"' if use_opc_ua else '=".Value"'
    # sheet['A2'] = '="Получение данных с OPC UA@"' if use_opc_ua else '="Получение данных Alpha.Link@"'
    # Для каждого объекта
    for obj in sl_object_all:
        num_cel_start = 4
        num_cel = 4
        num_par = 1
        # Создаём эксельку
        wb = openpyxl.Workbook()
        # Выбираем активный лист
        sheet = wb.active
        sheet.title = 'Лист 1'
        # Заполняем стартовую информацию
        for col, w in {'P': 5, 'Q': 50}.items():
            sheet.column_dimensions[col].width = w
        sheet['A1'] = f'="{obj[0]}."'

        if use_opc_ua:
            data_lst = {'A2': '="Получение данных с OPC UA@"', 'B2': '=".TRDELAY;1"', 'C2': '=".TDELAY;1"',
                            'D2': '=".SETPOINT;1"', 'E2': '=".VALUE;1"', 'F2': '=".VALUE.5501;1"',
                            'G2': '=".VALUE.100;1"', 'H2': '=".BLOCKED;1"', 'I2': '=".CHECK;1"',
                            'J2': '=".CHECKVALUE;1"'}
        else:
            data_lst = {'A2': '="Получение данных Alpha.Link@"', 'B2': '=".TRDELAY"', 'C2': '=".TDELAY"',
                        'D2': '=".SETPOINT"', 'E2': '=".VALUE"', 'F2': '=".VALUE.5501;1"',
                        'G2': '=".VALUE.100"', 'H2': '=".BLOCKED"', 'I2': '=".CHECK"',
                        'J2': '=".CHECKVALUE"'}

        for cell, value in data_lst.items():
            sheet[cell] = value
        for cell, value in {'B3': '="Таймер"', 'C3': '="Задержка"', 'D3': '="Уставка"', 'E3': '="Значение"',
                            'F3': '="Название"', 'G3': '="ед.измерения"'}.items():
            sheet[cell] = value
        # Создаём заголовок
        sheet.row_dimensions[3].height = 13
        sheet.row_dimensions[1].height = 17
        sheet['Q1'] = f'="Протокол проверки защит {obj[1]} на  "'
        sheet['Q1'].font = Font(size=14, name='Arial', bold=True)
        sheet['Q1'].alignment = Alignment(horizontal="right", vertical="top")
        sheet['R1'] = f'=NOW()'
        sheet['R1'].font = Font(size=14, name='Arial', bold=True)
        sheet['R1'].alignment = Alignment(horizontal="center", vertical="top")
        # sheet['R1'].number_format = numbers.BUILTIN_FORMATS[15]
        sheet['R1'].number_format = 'dd.mm.yyyy'
        sheet['S1'] = f'=NOW()'
        sheet['S1'].font = Font(size=14, name='Arial', bold=True)
        sheet['S1'].alignment = Alignment(horizontal="left", vertical="top")
        sheet['S1'].number_format = numbers.BUILTIN_FORMATS[20]
        for cell, i in {'P3': '="№"', 'Q3': '="Наименование защиты  "',
                        'R3': '="Таймер"', 'S3': '="Задержка"', 'T3': '="Уставка"', 'U3': '="Значение"',
                        'V3': '="Eд.изм"', 'W3': '="Отметка о проверке"'}.items():
            sheet[cell] = i
            sheet[cell].font = Font(size=12, name='Arial', bold=True)
            sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
            sheet[cell].border = thin_border

        # Для каждого контроллера...
        tuple_page = tuple()
        num_page = 1
        for cpu, sl_par in sl_param.items():
            # ...при условии, что объект содержит текущий контроллер...
            if cpu in sl_object_all[obj]:
                # ...для каждого параметра...
                for par, property_par in sl_par.items():
                    if 'Проверяется при ПЗ - Нет' in property_par:
                        continue
                    # Если отсчитали igrek параметров, то добавляем шапку для следующего листа
                    if not (num_par - 1) % igrek and num_par != 1:
                        num_cel += 1
                        tuple_page += (num_page,)
                        num_page += 1
                        igrek_delta = 10
                        for _ in range(3):
                            sheet.row_dimensions[num_cel].height = 12
                            # for cell, i in {'Q': '="должность"', 'S': '="ФИО"', 'U': '="подпись"'}.items():
                            #     sheet[f'{cell}{num_cel}'] = i
                            #     sheet[f'{cell}{num_cel}'].font = Font(size=5, name='Arial')
                            #     sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
                            # for coll in 'QRSTU':
                            #     sheet[f'{coll}{num_cel}'].border = Border(top=Side(style='thin'))
                            num_cel += 1

                        num_cel += 0
                        sheet.row_dimensions[num_cel].height = 17
                        sheet[f'Q{num_cel}'] = f'="Протокол проверки защит {obj[1]} на "'
                        sheet[f'Q{num_cel}'].font = Font(size=14, name='Arial', bold=True)
                        sheet[f'Q{num_cel}'].alignment = Alignment(horizontal="right", vertical="top")
                        sheet[f'R{num_cel}'] = f'=R1'
                        sheet[f'R{num_cel}'].font = Font(size=14, name='Arial', bold=True)
                        sheet[f'R{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
                        # sheet[f'R{num_cel}'].number_format = numbers.BUILTIN_FORMATS[15]
                        sheet[f'R{num_cel}'].number_format = 'dd.mm.yyyy'
                        sheet[f'S{num_cel}'] = f'=S1'
                        sheet[f'S{num_cel}'].font = Font(size=14, name='Arial', bold=True)
                        sheet[f'S{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
                        sheet[f'S{num_cel}'].number_format = numbers.BUILTIN_FORMATS[20]
                        num_cel += 2
                        sheet.row_dimensions[num_cel].height = 15
                        for cell, i in {'P': '="№"', 'Q': '="Наименование защиты  "',
                                        'R': '="Таймер"', 'S': '="Задержка"', 'T': '="Уставка"', 'U': '="Значение"',
                                        'V': '="Eд.изм"', 'W': '="Отметка о проверке"'}.items():
                            sheet[f'{cell}{num_cel}'] = i
                            sheet[f'{cell}{num_cel}'].font = Font(size=12, name='Arial', bold=True)
                            sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="center")
                            sheet[f'{cell}{num_cel}'].border = thin_border
                        num_cel += 1
                    # Заполняем параметры
                    sheet[f'A{num_cel}'] = f'="{node_alg_name}.{par}"'
                    sheet[f'P{num_cel}'] = f'="{num_par}"'
                    rus_name = property_par[1].replace("\"", "\"\"")
                    e_unit = property_par[2].replace("\"", "\"\"")
                    sheet[f'Q{num_cel}'] = f'="{rus_name}  "'  # Русское наименование
                    # Заполняем ссылки
                    for i, value in {'B': f'=CONCATENATE($A$2, $A$1, $A{num_cel}, B$2)',
                                     'C': f'=CONCATENATE($A$2, $A$1, $A{num_cel}, C$2)',
                                     'D': f'=CONCATENATE($A$2, $A$1, $A{num_cel}, D$2)',
                                     'E': f'=CONCATENATE($A$2, $A$1, $A{num_cel}, E$2)',
                                     #'F': f'=CONCATENATE($A$2, $A$1, $A{num_cel}, F$2)',
                                     # 'G': f'=CONCATENATE($A$2, $A$1, $A{num_cel}, G$2)',
                                     'H': f'=CONCATENATE($A$2, $A$1, $A{num_cel}, H$2)',
                                     'I': f'=CONCATENATE($A$2, $A$1, $A{num_cel}, I$2)',
                                     'J': f'=CONCATENATE($A$2, $A$1, $A{num_cel}, J$2)'}.items():
                        sheet[f'{i}{num_cel}'] = value
                    # Заполняем формулы
                    for i, value in {'K': f'=CurrAttrValue(D{num_cel}, 0)', 'L': f'=CurrAttrValue(E{num_cel}, 0)',
                                     'M': f'=CurrAttrValue(H{num_cel}, 0)', 'N': f'=CurrAttrValue(I{num_cel}, 0)',
                                     'O': f'=CurrAttrValue(J{num_cel}, 0)'}.items():
                        sheet[f'{i}{num_cel}'] = value
                    # Заполняем инфу по параметрам
                    for i, value in {'R': (f'=IF(N{num_cel}, S{num_cel}, "")', 17),
                                     'S': (f'=CurrAttrValue(C{num_cel}, 0)', 15),
                                     'T': (f'=IF(K{num_cel}=-200, "д.вх.", K{num_cel})', 13),
                                     'U': (f'=IF(L{num_cel}=-200, "д.вх.", IF(N{num_cel}, O{num_cel}, L{num_cel}))',
                                           15),
                                     # 'V': (f'=CurrAttrValue(G{num_cel}, 0)', 12),
                                     'V': (f'="{e_unit}"', 12),
                                     'W': (f'=IF(M{num_cel}, "Блокирована", IF(N{num_cel}, "Проверено", "-"))',
                                           25)}.items():
                        sheet[f'{i}{num_cel}'] = value[0]
                        sheet.column_dimensions[i].width = value[1]
                    # Выставляем формат
                    for i in 'RS':
                        sheet[f'{i}{num_cel}'].number_format = numbers.BUILTIN_FORMATS[2]
                    # Устанавливаем для ячеек размер шрифта 14, Arial и выставляем границу
                    for i in ('P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W'):
                        sheet[f'{i}{num_cel}'].font = Font(size=12, name='Arial')
                        sheet[f'{i}{num_cel}'].border = thin_border

                    frag_dig_par = int(property_par[4])
                    frag_dig = '' if not frag_dig_par else '.' + '0' * frag_dig_par
                    sheet[f'T{num_cel}'].number_format = f'0{frag_dig}'  # '0.000'
                    sheet[f'U{num_cel}'].number_format = f'0{frag_dig}'  # '0.000'
                    # Выводим в центр нумерацию, значения и единицы измерения
                    for i, vert_alig in {'P': 'center', 'Q': 'left', 'R': 'center', 'S': 'center',
                                         'T': 'center', 'U': 'center', 'V': 'center', 'W': 'center'}.items():
                        sheet[f'{i}{num_cel}'].alignment = Alignment(horizontal=vert_alig, vertical="center",
                                                                     wrap_text=True)
                    sheet.row_dimensions[num_cel].height = 38
                    num_cel += 1
                    num_par += 1
                    igrek_delta -= 1
        # Прячем столбцы с привязками
        for col in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O'):
            sheet.column_dimensions[col].hidden = True
        # Добавляем подпись в конце
        tuple_page += (num_page,)
        # num_cel += 1
        # for _ in range(3):
        #     sheet.row_dimensions[num_cel].height = 12
        #     for cell, i in {'Q': '="должность"', 'S': '="ФИО"', 'U': '="подпись"'}.items():
        #         sheet[f'{cell}{num_cel}'] = i
        #         sheet[f'{cell}{num_cel}'].font = Font(size=5, name='Arial')
        #         sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
        #     for coll in 'QRSTU':
        #         sheet[f'{coll}{num_cel}'].border = Border(top=Side(style='thin'))
        #     num_cel += 1

        for _ in range(igrek_delta):
            sheet.row_dimensions[num_cel].height = 38
            num_cel += 1
        cc = igrek + num_cel_start
        for _ in tuple_page:
            # print(_, cc)
            # print(sheet[f'G{_}'].value, end=' ')
            # tmp = sheet[f'G{_}'].value.replace('=', '').replace('"', '')
            sheet[f'W{cc}'].value = f'="{_}/{len(tuple_page)}"'
            sheet[f'W{cc}'].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            cc += (igrek + 7)
        # Добавляем колонтитул
        footer_text = ('_' * 144 + ' ' * 2 + 'должность' + ' ' * 130 + 'ФИО' + ' ' * 50 + 'подпись' + ' ' * 50) * 3
        sheet.oddFooter.center.text = footer_text
        sheet.oddFooter.center.size = 10
        sheet.oddFooter.center.font = "Arial,Bold"
        # Сохраняем
        wb.save(os.path.join('File_for_Import', 'Reports', f'{node_param_rus} {obj[1]}.xlsx'))
    return


def create_reports(sl_object_all: dict, node_param_rus: str, node_alg_name: str, sl_param: dict,
                   reports_protocol_version: int):
    sl_r = {'AI': 'значений измеряемых параметров', 'AE': 'значений расчётных параметров',
            'System.CNT': 'накопленных значений по наработке'}
    igrek = 20  # 16
    igrek_delta = 20
    # sl_object_all =
    # { (Объект, рус имя объекта, индекс объекта): {контроллер: (ip основной, ip резервный, индекс объекта)} }

    # sl_param = {cpu: {алг_пар: (тип параметра в студии, русское имя, ед изм., короткое имя, количество знаков)}}
    # num_cel = 4
    # num_par = 1

    thin_border = Border(left=Side(style='thin'),  # thick - толстые границы, thin - тонкие
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    use_opc_ua = False
    match reports_protocol_version:
        case 0: # 0 - OPC UA, 1 - Alpha.Link
            use_opc_ua = True
        case 1:  # 0 - OPC UA, 1 - Alpha.Link
            use_opc_ua = False
        case _:
            use_opc_ua = True

    # Для каждого объекта
    for obj in sl_object_all:
        num_cel_start = 4
        num_cel = 4
        num_par = 1
        # Создаём эксельку
        wb = openpyxl.Workbook()
        # Выбираем активный лист
        sheet = wb.active
        sheet.title = 'Лист 1'
        # ВЕРХНИЙ КОЛОНТИТУЛ ПОКА ПОД ВОПРОСОМ
        # text_header = f'="Срезь {sl_r.get(node_alg_name)} {obj[1]} на "'
        # sheet.oddHeader.center.text = text_header  # Центральный верхний колонтитул
        # sheet.oddHeader.center.size = 16
        # text_footer = '''__________________________________________________________________________________
        #             должность                                                    ФИО                                   подпись'''
        # text_footer = "Проверка текста"*2 + "│" + " "*13 + "должность" + " "*52 + "ФИО"+ " "*35 + "подпись"
        # sheet.oddFooter.center.text = text_footer # Центральный нижний колонтитул
        # sheet.oddFooter.center.size = 10
        # Заполняем стартовую информацию
        sheet.column_dimensions['D'].width = 7 #5.86
        sheet.column_dimensions['E'].width = 100 # 94.86
        sheet.column_dimensions['F'].width = 25 # 19.86
        sheet.column_dimensions['G'].width = 15 #14.86
        sheet['A1'] = f'="{obj[0]}."'
        sheet['B1'] = '=".Value;1"' if use_opc_ua else '=".Value"'
        sheet['A2'] = '="Получение данных с OPC UA@"' if use_opc_ua else '="Получение данных Alpha.Link@"'
        # sheet['B2'] = '=".Value.100;1"'
        # Создаём заголовок
        sheet.row_dimensions[3].height = 20
        sheet.row_dimensions[1].height = 35
        sheet['E1'] = f'="Срез {sl_r.get(node_alg_name)} {obj[1]} на "'
        sheet['E1'].font = Font(size=16, name='Arial', bold=True)
        sheet['E1'].alignment = Alignment(horizontal="right", vertical="top")
        sheet['F1'] = f'=NOW()'
        sheet['F1'].font = Font(size=16, name='Arial', bold=True)
        sheet['F1'].alignment = Alignment(horizontal="center", vertical="top")
        # sheet['F1'].number_format = numbers.BUILTIN_FORMATS[15]
        sheet['F1'].number_format = 'dd.mm.yyyy'
        sheet['G1'] = f'=NOW()'
        sheet['G1'].font = Font(size=16, name='Arial', bold=True)
        sheet['G1'].alignment = Alignment(horizontal="left", vertical="top")
        sheet['G1'].number_format = numbers.BUILTIN_FORMATS[20]
        for cell, i in {'D3': '="№"', 'E3': '="Наименование параметра  "',
                        'F3': '="Значение"', 'G3': '="Ед. изм."'}.items():
            sheet[cell] = i
            sheet[cell].font = Font(size=14, name='Arial', bold=True)
            sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
            sheet[cell].border = thin_border
        tuple_page = tuple()
        num_page = 1
        # Для каждого контроллера...
        for cpu, sl_par in sl_param.items():
            # ...при условии, что объект содержит текущий контроллер...
            if cpu in sl_object_all[obj]:
                # ...для каждого параметра...
                for par, property_par in sl_par.items():
                    # Если отсчитали игрек параметров, то добавляем шапку для следующего листа
                    if not (num_par - 1) % igrek and num_par != 1:
                        # sheet[f'G{num_cel}'] = f'="{num_page}"'
                        tuple_page += (num_page,)
                        num_page += 1
                        igrek_delta = 20
                        num_cel += 2
                        # sheet.row_dimensions[num_cel].height = 35
                        # for cell, i in {'E': '="должность"', 'F': '="ФИО"', 'G': '="подпись"'}.items():
                        #     sheet[f'{cell}{num_cel}'] = i
                        #     sheet[f'{cell}{num_cel}'].font = Font(size=16, name='Arial', bold=True)
                        #     sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
                        #     sheet[f'{cell}{num_cel}'].border = Border(top=Side(style='thin'))

                        num_cel += 1
                        sheet.row_dimensions[num_cel].height = 35
                        sheet[f'E{num_cel}'] = f'="Срез {sl_r.get(node_alg_name)} {obj[1]} на "'
                        sheet[f'E{num_cel}'].font = Font(size=16, name='Arial', bold=True)
                        sheet[f'E{num_cel}'].alignment = Alignment(horizontal="right", vertical="top")
                        sheet[f'F{num_cel}'] = f'=F1'
                        sheet[f'F{num_cel}'].font = Font(size=16, name='Arial', bold=True)
                        sheet[f'F{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
                        # sheet[f'F{num_cel}'].number_format = numbers.BUILTIN_FORMATS[15]
                        sheet[f'F{num_cel}'].number_format = 'dd.mm.yyyy'
                        sheet[f'G{num_cel}'] = f'=G1'
                        sheet[f'G{num_cel}'].font = Font(size=16, name='Arial', bold=True)
                        sheet[f'G{num_cel}'].alignment = Alignment(horizontal="left", vertical="top")
                        sheet[f'G{num_cel}'].number_format = numbers.BUILTIN_FORMATS[20]
                        num_cel += 2
                        sheet.row_dimensions[num_cel].height = 20
                        for cell, i in {'D': '="№"', 'E': '="Наименование параметра  "',
                                        'F': '="Значение"', 'G': '="Ед. изм"'}.items():
                            sheet[f'{cell}{num_cel}'] = i
                            sheet[f'{cell}{num_cel}'].font = Font(size=14, name='Arial', bold=True)
                            sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="center")
                            sheet[f'{cell}{num_cel}'].border = thin_border
                        num_cel += 1

                    sheet[f'C{num_cel}'] = f'="{node_alg_name}.{par}"'
                    sheet[f'D{num_cel}'] = f'="{num_par}"'
                    rus_name = property_par[1].replace("\"", "\"\"")
                    sheet[f'E{num_cel}'] = f'="{rus_name}  "'  # Русское наименование
                    # sheet[f'A{num_cel}'] = f'=CONCATENATE($A$2, $A$1, C{num_cel}, $B$2)'
                    sheet[f'B{num_cel}'] = f'=CONCATENATE($A$2, $A$1, C{num_cel}, $B$1)'
                    sheet[f'F{num_cel}'] = f'=CurrAttrValue(B{num_cel}, 0)'
                    # Проверить работу на живых данных
                    if node_alg_name == 'System.CNT':
                        sheet[f'F{num_cel}'].number_format = '0.000'
                    else:
                        frag_dig_par =  int(property_par[4][0])
                        frag_dig = '' if not frag_dig_par else '.' + '0'* frag_dig_par
                        # print(rus_name, '0' + frag_dig)
                        sheet[f'F{num_cel}'].number_format = f'0{frag_dig}' # '0.000'

                    e_unit = (('-' if 'Swap' in par else 'час') if node_alg_name == 'System.CNT'
                              else property_par[2].replace("\"", "\"\"") ) # else f'=CurrAttrValue(A{num_cel}, 0)')
                    sheet[f'G{num_cel}'] = f'="{e_unit}  "'
                    # Устанавливаем для ячеек размер шрифта 14, Arial и выставляем границу
                    for i in ('C', 'D', 'E', 'A', 'B', 'F', 'G'):  # ('B', 'C', 'D', 'E', 'F', 'G'):
                        sheet[f'{i}{num_cel}'].font = Font(size=14, name='Arial')
                        sheet[f'{i}{num_cel}'].border = thin_border
                    for i in ('C', 'D', 'E', 'A', 'B', 'F', 'G'):  # ('A', ):
                        sheet[f'{i}{num_cel}'].font = Font(size=13.999, name='Arial')
                        sheet[f'{i}{num_cel}'].border = thin_border
                    # Выводим в центр нумерацию, значения и единицы измерения
                    for i, vert_alig in {'D': 'center', 'E': 'left', 'F': 'center', 'G': 'center'}.items():
                        sheet[f'{i}{num_cel}'].alignment = Alignment(horizontal=vert_alig, vertical="center",
                                                                     wrap_text=True)
                    sheet.row_dimensions[num_cel].height = 18
                    num_cel += 1
                    num_par += 1
                    igrek_delta -= 1
        # Прячем столбцы с привязками
        for col in ('A', 'B', 'C'):
            sheet.column_dimensions[col].hidden = True
        # Добавляем подпись в конце
        # sheet[f'G{num_cel}'] = f'="{num_page}"'
        tuple_page += (num_page,)
        # num_cel += 2
        # sheet.row_dimensions[num_cel].height = 35
        # for cell, i in {'E': '="должность"', 'F': '="ФИО"', 'G': '="подпись"'}.items():
        #     sheet[f'{cell}{num_cel}'] = i
        #     sheet[f'{cell}{num_cel}'].font = Font(size=16, name='Arial', bold=True)
        #     sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
        #     sheet[f'{cell}{num_cel}'].border = Border(top=Side(style='thin'))

        # print(tuple_page, node_alg_name, obj)
        # print(node_alg_name, igrek_delta)
        # for _ in range(igrek_delta):
        #     sheet.row_dimensions[num_cel].height = 20
        #     num_cel += 1
        # cc = igrek + num_cel_start
        # for _ in tuple_page:
        #     # print(_, cc)
        #     # print(sheet[f'G{_}'].value, end=' ')
        #     # tmp = sheet[f'G{_}'].value.replace('=', '').replace('"', '')
        #     sheet[f'G{cc}'].value = f'="{_}/{len(tuple_page)}"'
        #     sheet[f'G{cc}'].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        #     cc += (igrek + 4)
        # # Добавляем колонтитул
        # footer_text = ('_' * 144 + ' ' * 2 + 'должность' + ' ' * 130 + 'ФИО' + ' ' * 50 + 'подпись' + ' ' * 50) * 1
        # sheet.oddFooter.center.text = footer_text
        # sheet.oddFooter.center.size = 10
        # sheet.oddFooter.center.font = "Arial,Bold"
        # Добавляем колонтитул
        footer_text = ('_' * 144 + ' ' * 2 + 'должность' + ' ' * 130 + 'ФИО' + ' ' * 50 + 'подпись' + ' ' * 50) * 1
        sheet.oddFooter.center.text = footer_text
        sheet.oddFooter.center.size = 10
        sheet.oddFooter.center.font = "Arial,Bold"

        # text_header = f'="Срез {sl_r.get(node_alg_name)} {obj[1]} на "'
        # sheet.oddHeader.center.text = text_header # Центральный верхний колонтитул
        # sheet.oddHeader.center.size = 16
        # sheet.oddHeader.center.font = "Arial,Bold"

        # Если нет папки File_for_Import/Reports, то создадим её
        if not os.path.exists(os.path.join('File_for_Import', 'Reports')):
            os.mkdir(os.path.join('File_for_Import', 'Reports'))
        # Сохраняем
        wb.save(os.path.join('File_for_Import', 'Reports', f'{node_param_rus} {obj[1]}.xlsx'))

    return



def create_reports_sday(sl_object_all: dict, node_param_rus: str, reports_protocol_version: int):
    igrek = 21
    igrek_delta = 24
    sl_param = dict()
    #   os.path.dirname(sys.argv[0])
    if os.path.exists(os.path.join(os.path.abspath(os.curdir), 'SCADA', 'Daily')):
        all_obj = tuple([i[0] for i in sl_object_all])
        for file in os.listdir(os.path.join(os.path.abspath(os.curdir), 'SCADA', 'Daily')):
            # print(file, file[:file.find('_')] in all_obj and file.endswith('DailyList.csv'))
            if file[:file.find('_')] in all_obj and file.endswith('DailyList.csv'):
                # print(f"---------------------{file[:file.find('_')]}--------------------------")
                with open(os.path.join(os.path.abspath(os.curdir), 'SCADA', 'Daily', file), 'r') as f_:
                    for line in f_:
                        if 'Ошибка' in line or ';' not in line:
                            continue
                        tmp_list = line.strip().split(';')
                        if 'find' in tmp_list:
                            continue
                        # print(tmp_list)
                        folder_tag = tmp_list[1]
                        obj = folder_tag[:folder_tag.find('.')]
                        ind_end_str = folder_tag.rfind('.Value.V') if folder_tag.endswith('.Value.V') else len(folder_tag)
                        tag = folder_tag[folder_tag.find('.') + 1:ind_end_str]
                        description = tmp_list[2]
                        eunit = tmp_list[3]
                        precision = eunit[:eunit.find('|')]
                        eunit = '-' if 'Не используется' in eunit else eunit[eunit.find('|')+1:]
                        short_name = tmp_list[0]

                        if obj not in sl_param:
                            sl_param[obj] = {}
                        if tag not in sl_param[obj]:
                            sl_param[obj][tag] = (short_name, description, eunit, precision)
                            # print(tag, short_name, description, eunit, precision)
    else:
        print('Папка SCADA/Daily не найдена')

    sl_name_object = {alg_name[0]: alg_name[1] for alg_name in sl_object_all}
    # print(sl_param)

    thin_border = Border(left=Side(style='thin'),  # thick - толстые границы, thin - тонкие
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    tuple_sh = ('W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH')
    str_sh = 'IJKLMNOPQRST'

    sl_param = {i: value for i, value in sl_param.items() if i}

    use_opc_ua = False
    match reports_protocol_version:
        case 0:  # 0 - OPC UA, 1 - Alpha.Link
            use_opc_ua = True
        case 1:  # 0 - OPC UA, 1 - Alpha.Link
            use_opc_ua = False
        case _:
            use_opc_ua = True

    # sheet['B1'] = '=".Value;1"' if use_opc_ua else '=".Value"'
    # sheet['A2'] = '="Получение данных с OPC UA@"' if use_opc_ua else '="Получение данных Alpha.Link@"'

    # Для каждого объекта
    for obj, sl_par in sl_param.items():
        # for p in sl_par:
        #     print(p)
        num_cel_start = 8
        num_cel = 8
        num_par = 1
        # Создаём эксельку
        wb = openpyxl.Workbook()
        # Выбираем активный лист
        sheet = wb.active
        sheet.title = 'Лист 1'
        # Заполняем стартовую информацию
        for col, w in {'F': 5, 'G': 23, 'H': 8}.items():
            sheet.column_dimensions[col].width = w
        sheet['E2'] = f'="{obj}."'

        for cell, value in {'A2': '=NOW()', 'B2': '="№ ГПА"', 'A3': '=reportdate', 'A4': '=smena',
                            'A5': '="smena=1"',
                            'A6': '="smena=2"', 'B3': '="дата"', 'B4': '="смена"', 'B5': '="дневная"',
                            'B6': '="ночная"', 'C7': '="узел"', 'D7': '="ед.изм"', 'E7': '="значение"',
                            'D3': '="начало пути"',
                            'E3': '="Получение данных с OPC UA@"' if use_opc_ua else '="Получение данных Alpha.Link@"',
                            'D4': '="значение"',
                            'E4': '=".Value;1"' if use_opc_ua else '=".Value"',
                            'E5': '=".Value.100;1"',
                            'A9': '=YEAR(A3)', 'B9': '="год"', 'A10': '=MONTH(A3)', 'B10': '="месяц"',
                            'A11': '=DAY(A3)', 'B11': '="день"', 'A12': '=DATE(A9, A10, A11)',
                            'A14': '=IF(A4=1, TIME(8, 0, 0), TIME(20, 0, 0))', 'A15': '=TIME(0, 3, 0)',
                            'B14': '="выбор смены"', 'B15': '="диапазон"'}.items():
            sheet[cell] = value

        # Заполняем временные формулы
        for cell, value in {'I5': '=A12+A14', 'I6': '=I5+$A$15'}.items():  # 'I6': '=I5+$A$15'
            sheet[cell] = value
        cell_ind_str = 'JKLMNOPQRST'
        formul_ind_str = 'IJKLMNOPQRS'
        for i in range(len(cell_ind_str)):
            sheet[f'{cell_ind_str[i]}5'] = f'={formul_ind_str[i]}5+1/24'
            sheet[f'{cell_ind_str[i]}6'] = f'={cell_ind_str[i]}5+$A$15'
        for coll in 'IJKLMNOPQRST':
            sheet[f'{coll}4'] = f'=IF($A$2>{coll}5, true, false)'
        # Выставляем форматы
        for cell, value in {'A2': 22, 'A3': 22, 'A12': 22, 'A14': 21, 'A15': 21}.items():
            sheet[cell].number_format = numbers.BUILTIN_FORMATS[value]
        for coll in 'IJKLMNOPQRST':
            sheet[f'{coll}5'].number_format = numbers.BUILTIN_FORMATS[22]
            sheet[f'{coll}6'].number_format = numbers.BUILTIN_FORMATS[22]
        # Заполняем шапку формул данных
        sheet['V5'] = '="текущие значения"'
        for i in range(len(tuple_sh)):
            sheet[f'{tuple_sh[i]}5'] = f'={str_sh[i]}7'
            sheet[f'{tuple_sh[i]}5'].number_format = numbers.BUILTIN_FORMATS[22]

        # Создаём заголовок
        sheet.row_dimensions[2].height = 20
        sheet.row_dimensions[1].height = 25
        sheet.merge_cells('G2:L2')
        sheet[f'G2'] = f'="Сменная ведомость {sl_name_object.get(obj, str(777))} на "'
        sheet[f'G2'].font = Font(size=14, name='Arial', bold=True)
        sheet[f'G2'].alignment = Alignment(horizontal="right", vertical="top")
        sheet.merge_cells('M2:N2')
        sheet[f'M2'] = f'=A3'
        sheet[f'M2'].font = Font(size=14, name='Arial', bold=True)
        sheet[f'M2'].alignment = Alignment(horizontal="center", vertical="top")
        sheet[f'M2'].number_format = 'dd.mm.yyyy'

        sheet.row_dimensions[7].height = 15
        for cell, i in {'F7': '="№"', 'G7': '="Наименование  "',
                        'H7': '="ед.изм"'}.items():
            sheet[cell] = i
            sheet[cell].font = Font(size=10, name='Arial', bold=True)
            sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
            sheet[cell].border = thin_border

        # Проставляем время каждый час в первой шапке
        for coll in str_sh:
            sheet[f'{coll}7'] = f'=TIME(HOUR({coll}5), 0, 0)'
            sheet[f'{coll}7'].font = Font(size=10, name='Arial', bold=True)
            sheet[f'{coll}7'].alignment = Alignment(horizontal="center", vertical="center")
            sheet[f'{coll}7'].border = thin_border
            sheet[f'{coll}7'].number_format = numbers.BUILTIN_FORMATS[20]
            sheet.column_dimensions[coll].width = 9

        # ...для каждого параметра...
        tuple_page = tuple()
        num_page = 1
        for par, property_par in sl_par.items():
            # Если отсчитали igrek параметров, то добавляем шапку для следующего листа
            if not (num_par - 1) % igrek and num_par != 1:
                num_cel += 2
                tuple_page += (num_page,)
                num_page += 1
                igrek_delta = 24
                for _ in range(3):
                    # sheet.row_dimensions[num_cel].height = 35
                    # for cell, i in {'G': '="должность"', 'J': '="ФИО"', 'M': '="подпись"'}.items():
                    #     sheet[f'{cell}{num_cel}'] = i
                    #     sheet[f'{cell}{num_cel}'].font = Font(size=12, name='Arial')
                    #     sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
                    # for coll in 'GHIJKLMN':
                    #     sheet[f'{coll}{num_cel}'].border = Border(top=Side(style='thin'))
                    num_cel += 1

                num_cel += 2
                sheet.row_dimensions[num_cel].height = 25
                sheet.merge_cells(f'G{num_cel}:L{num_cel}')
                sheet[f'G{num_cel}'] = f'="Сменная ведомость {sl_name_object.get(obj, str(777))} на "'
                sheet[f'G{num_cel}'].font = Font(size=14, name='Arial', bold=True)
                sheet[f'G{num_cel}'].alignment = Alignment(horizontal="right", vertical="top")
                sheet.merge_cells(f'M{num_cel}:N{num_cel}')
                sheet[f'M{num_cel}'] = f'=M2'
                sheet[f'M{num_cel}'].font = Font(size=14, name='Arial', bold=True)
                sheet[f'M{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
                sheet[f'M{num_cel}'].number_format = 'dd.mm.yyyy'
                num_cel += 2
                sheet.row_dimensions[num_cel].height = 15
                for cell, i in {'F': '="№"', 'G': '="Наименование  "',
                                'H': '="ед.изм"'}.items():
                    sheet[f'{cell}{num_cel}'] = i
                    sheet[f'{cell}{num_cel}'].font = Font(size=10, name='Arial', bold=True)
                    sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="center")
                    sheet[f'{cell}{num_cel}'].border = thin_border
                for coll in 'IJKLMNOPQRST':
                    sheet[f'{coll}{num_cel}'] = f'={coll}7'
                    sheet[f'{coll}{num_cel}'].font = Font(size=10, name='Arial', bold=True)
                    sheet[f'{coll}{num_cel}'].alignment = Alignment(horizontal="center", vertical="center")
                    sheet[f'{coll}{num_cel}'].border = thin_border
                    sheet[f'{coll}{num_cel}'].number_format = numbers.BUILTIN_FORMATS[20]
                num_cel += 1
            # Заполняем параметры
            sheet[f'C{num_cel}'] = f'="{par}"'
            sheet[f'F{num_cel}'] = f'="{num_par}"'
            rus_name = property_par[0]
            sheet[f'G{num_cel}'] = f'="{rus_name}  "'  # Русское наименование

            # Заполняем формулы для вывода данных
            # Получение текущего значения, пока закомментировано, чтобы не вызывать дополнительные расчёты
            # sheet[f'V{num_cel}'] = f'=CurrAttrValue(E{num_cel}, 0)'
            for i in range(len(tuple_sh)):
                # sheet[f'{tuple_sh[i]}{num_cel}'] = f'=ArchiveAttributeValue($E{num_cel}, ' \
                #                                    f'0, {str_sh[i]}$5, {str_sh[i]}$6, 1)'
                sheet[f'{tuple_sh[i]}{num_cel}'] = f'=valueondate($E{num_cel}, 0, {str_sh[i]}$5, FALSE())'

                # sheet[f'{tuple_sh[i]}{num_cel}'] = f'=ValueOnDate($E{num_cel}, 0, {str_sh[i]}$5)'
            for i in range(len(str_sh)):
                # sheet[f'{str_sh[i]}{num_cel}'] = f'=IF({str_sh[i]}$4, IF(ISNUMBER({tuple_sh[i]}{num_cel}), ' \
                #                                  f'{tuple_sh[i]}{num_cel}, "нет"), "-")'  # предпоследний "-" это был $V{num_cel}
                sheet[f'{str_sh[i]}{num_cel}'] = f'=IF({str_sh[i]}$4, {tuple_sh[i]}{num_cel}, "-")'  # предпоследний "-" это был $V{num_cel}
                sheet[f'{str_sh[i]}{num_cel}'].number_format = numbers.BUILTIN_FORMATS[2]
                sheet[f'{str_sh[i]}{num_cel}'].font = Font(size=9, name='Arial')
                sheet[f'{str_sh[i]}{num_cel}'].border = thin_border
                sheet[f'{str_sh[i]}{num_cel}'].alignment = Alignment(horizontal="center", vertical="center")
            # Заполняем ссылки
            for i, value in {# 'D': f'=CONCATENATE($E$3, $E$2, $C{num_cel}, $E$5)',
                             'E': f'=CONCATENATE($E$3, $E$2, $C{num_cel}, $E$4)'}.items():
                sheet[f'{i}{num_cel}'] = value
            # Заполняем формулы
            for i, value in {'H': f'=CurrAttrValue(D{num_cel}, 0)'}.items():
                # sheet[f'{i}{num_cel}'] = value
                sheet[f'{i}{num_cel}'] = f'="{property_par[2]}"'
                # print(property_par)

            # Устанавливаем для ячеек размер шрифта 14, Arial и выставляем границу
            for i in 'FGH':
                sheet[f'{i}{num_cel}'].font = Font(size=10, name='Arial')
                sheet[f'{i}{num_cel}'].border = thin_border
            # Выводим в центр нумерацию, значения и единицы измерения
            for i, vert_alig in {'F': 'center', 'G': 'left', 'H': 'center', 'S': 'center'}.items():
                sheet[f'{i}{num_cel}'].alignment = Alignment(horizontal=vert_alig, vertical="center",
                                                             wrap_text=True)
            sheet.row_dimensions[num_cel].height = 15
            num_cel += 1
            num_par += 1
            igrek_delta -= 1

        # Прячем столбцы с привязками
        for col in ('A', 'B', 'C', 'D', 'E'):
            sheet.column_dimensions[col].hidden = True
        for row in (4, 5, 6):
            sheet.row_dimensions[row].hidden = True
        for col in ('U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK',
                    'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX'):
            sheet.column_dimensions[col].hidden = True
        # Добавляем подпись в конце
        num_cel += 2
        tuple_page += (num_page,)

        # for _ in range(3):
        #     sheet.row_dimensions[num_cel].height = 35
        #     for cell, i in {'G': '="должность"', 'J': '="ФИО"', 'M': '="подпись"'}.items():
        #         sheet[f'{cell}{num_cel}'] = i
        #         sheet[f'{cell}{num_cel}'].font = Font(size=12, name='Arial')
        #         sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
        #     for coll in 'GHIJKLMN':
        #         sheet[f'{      coll}{num_cel}'].border = Border(top=Side(style='thin'))
        #     num_cel += 1

        # Добавляем общие имена (локальные переменные)
        new_range = DefinedName('smena', attr_text=1, comment='2') # attr_text='"1"'
        wb.defined_names.add(new_range)
        new_range = DefinedName('reportdate', attr_text='"44987"', comment='0')
        wb.defined_names.add(new_range)

        # for _ in range(igrek_delta):
        #     sheet.row_dimensions[num_cel].height = 15
        #     num_cel += 1
        # cc = igrek + num_cel_start
        # for _ in tuple_page:
        #     # print(_, cc)
        #     # print(sheet[f'G{_}'].value, end=' ')
        #     # tmp = sheet[f'G{_}'].value.replace('=', '').replace('"', '')
        #     sheet[f'U{cc}'].value = f'="{_}/{len(tuple_page)}"'
        #     sheet[f'U{cc}'].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        #     cc += (igrek + 6)
        #
        # # Добавляем колонтитул
        # footer_text = ('_' * 144 + ' ' * 2 + 'должность' + ' ' * 130 + 'ФИО' + ' ' * 50 + 'подпись' + ' ' * 50) * 3
        # sheet.oddFooter.center.text = footer_text
        # sheet.oddFooter.center.size = 10
        # sheet.oddFooter.center.font = "Arial,Bold"
        # Добавляем колонтитул
        footer_text = ('_' * 144 + ' ' * 2 + 'должность' + ' ' * 130 + 'ФИО' + ' ' * 50 + 'подпись' + ' ' * 50) * 3
        sheet.oddFooter.center.text = footer_text
        sheet.oddFooter.center.size = 10
        sheet.oddFooter.center.font = "Arial,Bold"

        # Если нет папки File_for_Import/Reports, то создадим её
        if not os.path.exists(os.path.join('File_for_Import', 'Reports')):
            os.mkdir(os.path.join('File_for_Import', 'Reports'))
        # Сохраняем
        wb.save(
            os.path.join('File_for_Import', 'Reports', f'{node_param_rus} {sl_name_object.get(obj, str(777))}.xlsx')
        )

    return


def create_reports_sut(sl_object_all: dict, node_param_rus: str, reports_protocol_version: int):
    igrek = 21
    igrek_delta = 24
    sl_param = dict()
    #   os.path.dirname(sys.argv[0])
    if os.path.exists(os.path.join(os.path.abspath(os.curdir), 'SCADA', 'Daily')):
        all_obj = tuple([i[0] for i in sl_object_all])
        for file in os.listdir(os.path.join(os.path.abspath(os.curdir), 'SCADA', 'Daily')):
            # print(file, file[:file.find('_')] in all_obj and file.endswith('DailyList.csv'))
            if file[:file.find('_')] in all_obj and file.endswith('DailyList.csv'):
                # print(f"---------------------{file[:file.find('_')]}--------------------------")
                with open(os.path.join(os.path.abspath(os.curdir), 'SCADA', 'Daily', file), 'r') as f_:
                    for line in f_:
                        if 'Ошибка' in line or ';' not in line:
                            continue
                        tmp_list = line.strip().split(';')
                        if 'find' in tmp_list:
                            continue
                        # print(tmp_list)
                        folder_tag = tmp_list[1]
                        obj = folder_tag[:folder_tag.find('.')]
                        ind_end_str = folder_tag.rfind('.Value.V') if folder_tag.endswith('.Value.V') else len(folder_tag)
                        tag = folder_tag[folder_tag.find('.') + 1:ind_end_str]
                        description = tmp_list[2]
                        eunit = tmp_list[3]
                        precision = eunit[:eunit.find('|')]
                        eunit = '-' if 'Не используется' in eunit else eunit[eunit.find('|')+1:]
                        short_name = tmp_list[0]

                        if obj not in sl_param:
                            sl_param[obj] = {}
                        if tag not in sl_param[obj]:
                            sl_param[obj][tag] = (short_name, description, eunit, precision)
                            # print(tag, short_name, description, eunit, precision)
    else:
        print('Папка SCADA/Daily не найдена')

    sl_name_object = {alg_name[0]: alg_name[1] for alg_name in sl_object_all}
    # print(sl_param)

    thin_border = Border(left=Side(style='thin'),  # thick - толстые границы, thin - тонкие
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    tuple_sh = ('W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH')
    str_sh = 'IJKLMNOPQRST'

    sl_param = {i: value for i, value in sl_param.items() if i}

    use_opc_ua = False
    match reports_protocol_version:
        case 0:  # 0 - OPC UA, 1 - Alpha.Link
            use_opc_ua = True
        case 1:  # 0 - OPC UA, 1 - Alpha.Link
            use_opc_ua = False
        case _:
            use_opc_ua = True

    # sheet['B1'] = '=".Value;1"' if use_opc_ua else '=".Value"'
    # sheet['A2'] = '="Получение данных с OPC UA@"' if use_opc_ua else '="Получение данных Alpha.Link@"'

    # Для каждого объекта
    for obj, sl_par in sl_param.items():
        # for p in sl_par:
        #     print(p)
        num_cel_start = 8
        num_cel = 8
        num_par = 1
        # Создаём эксельку
        wb = openpyxl.Workbook()
        # Выбираем активный лист
        sheet = wb.active
        sheet.title = 'Лист 1'
        # Заполняем стартовую информацию
        for col, w in {'F': 5, 'G': 23, 'H': 8}.items():
            sheet.column_dimensions[col].width = w
        sheet['E2'] = f'="{obj}."'

        for cell, value in {'A2': '=NOW()', 'B2': '="№ ГПА"', 'A3': '=reportdate', 'A4': '=smena',
                            'A5': '="smena=1"',
                            'A6': '="smena=2"', 'B3': '="дата"', 'B4': '="смена"', 'B5': '="дневная"',
                            'B6': '="ночная"', 'C7': '="узел"', 'D7': '="ед.изм"', 'E7': '="значение"',
                            'D3': '="начало пути"',
                            'E3': '="Получение данных с OPC UA@"' if use_opc_ua else '="Получение данных Alpha.Link@"',
                            'D4': '="значение"',
                            'E4': '=".Value;1"' if use_opc_ua else '=".Value"',
                            'E5': '=".Value.100;1"',
                            'A9': '=YEAR(A3)', 'B9': '="год"', 'A10': '=MONTH(A3)', 'B10': '="месяц"',
                            'A11': '=DAY(A3)', 'B11': '="день"', 'A12': '=DATE(A9, A10, A11)',
                            'A14': '=TIME(0, 0, 0)', 'A15': '=TIME(0, 3, 0)',  # 'A14': '=IF(A4=1, TIME(8, 0, 0), TIME(20, 0, 0))'
                            'B14': '="выбор смены"', 'B15': '="диапазон"'}.items():
            sheet[cell] = value

        # Заполняем временные формулы
        for cell, value in {'I5': '=A12+A14', 'I6': '=I5+$A$15'}.items():  # 'I6': '=I5+$A$15'
            sheet[cell] = value
        cell_ind_str = 'JKLMNOPQRST'
        formul_ind_str = 'IJKLMNOPQRS'
        for i in range(len(cell_ind_str)):
            sheet[f'{cell_ind_str[i]}5'] = f'={formul_ind_str[i]}5+2/24'
            sheet[f'{cell_ind_str[i]}6'] = f'={cell_ind_str[i]}5+$A$15'
        for coll in 'IJKLMNOPQRST':
            sheet[f'{coll}4'] = f'=IF($A$2>{coll}5, true, false)'
        # Выставляем форматы
        for cell, value in {'A2': 22, 'A3': 22, 'A12': 22, 'A14': 21, 'A15': 21}.items():
            sheet[cell].number_format = numbers.BUILTIN_FORMATS[value]
        for coll in 'IJKLMNOPQRST':
            sheet[f'{coll}5'].number_format = numbers.BUILTIN_FORMATS[22]
            sheet[f'{coll}6'].number_format = numbers.BUILTIN_FORMATS[22]
        # Заполняем шапку формул данных
        sheet['V5'] = '="текущие значения"'
        for i in range(len(tuple_sh)):
            sheet[f'{tuple_sh[i]}5'] = f'={str_sh[i]}7'
            sheet[f'{tuple_sh[i]}5'].number_format = numbers.BUILTIN_FORMATS[22]

        # Создаём заголовок
        sheet.row_dimensions[2].height = 20
        sheet.row_dimensions[1].height = 25
        sheet.merge_cells('G2:L2')
        sheet[f'G2'] = f'="Суточная ведомость {sl_name_object.get(obj, str(777))} на "'
        sheet[f'G2'].font = Font(size=14, name='Arial', bold=True)
        sheet[f'G2'].alignment = Alignment(horizontal="right", vertical="top")
        sheet.merge_cells('M2:N2')
        sheet[f'M2'] = f'=A3'
        sheet[f'M2'].font = Font(size=14, name='Arial', bold=True)
        sheet[f'M2'].alignment = Alignment(horizontal="center", vertical="top")
        sheet[f'M2'].number_format = 'dd.mm.yyyy'

        sheet.row_dimensions[7].height = 15
        for cell, i in {'F7': '="№"', 'G7': '="Наименование  "',
                        'H7': '="ед.изм"'}.items():
            sheet[cell] = i
            sheet[cell].font = Font(size=10, name='Arial', bold=True)
            sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
            sheet[cell].border = thin_border

        # Проставляем время каждый час в первой шапке
        for coll in str_sh:
            sheet[f'{coll}7'] = f'=TIME(HOUR({coll}5), 0, 0)'
            sheet[f'{coll}7'].font = Font(size=10, name='Arial', bold=True)
            sheet[f'{coll}7'].alignment = Alignment(horizontal="center", vertical="center")
            sheet[f'{coll}7'].border = thin_border
            sheet[f'{coll}7'].number_format = numbers.BUILTIN_FORMATS[20]
            sheet.column_dimensions[coll].width = 9

        # ...для каждого параметра...
        tuple_page = tuple()
        num_page = 1
        for par, property_par in sl_par.items():
            # Если отсчитали igrek параметров, то добавляем шапку для следующего листа
            if not (num_par - 1) % igrek and num_par != 1:
                num_cel += 2
                tuple_page += (num_page,)
                num_page += 1
                igrek_delta = 24
                for _ in range(3):
                    # sheet.row_dimensions[num_cel].height = 35
                    # for cell, i in {'G': '="должность"', 'J': '="ФИО"', 'M': '="подпись"'}.items():
                    #     sheet[f'{cell}{num_cel}'] = i
                    #     sheet[f'{cell}{num_cel}'].font = Font(size=12, name='Arial')
                    #     sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
                    # for coll in 'GHIJKLMN':
                    #     sheet[f'{coll}{num_cel}'].border = Border(top=Side(style='thin'))
                    num_cel += 1

                num_cel += 2
                sheet.row_dimensions[num_cel].height = 25
                sheet.merge_cells(f'G{num_cel}:L{num_cel}')
                sheet[f'G{num_cel}'] = f'="Суточная ведомость {sl_name_object.get(obj, str(777))} на "'
                sheet[f'G{num_cel}'].font = Font(size=14, name='Arial', bold=True)
                sheet[f'G{num_cel}'].alignment = Alignment(horizontal="right", vertical="top")
                sheet.merge_cells(f'M{num_cel}:N{num_cel}')
                sheet[f'M{num_cel}'] = f'=M2'
                sheet[f'M{num_cel}'].font = Font(size=14, name='Arial', bold=True)
                sheet[f'M{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
                sheet[f'M{num_cel}'].number_format = 'dd.mm.yyyy'
                num_cel += 2
                sheet.row_dimensions[num_cel].height = 15
                for cell, i in {'F': '="№"', 'G': '="Наименование  "',
                                'H': '="ед.изм"'}.items():
                    sheet[f'{cell}{num_cel}'] = i
                    sheet[f'{cell}{num_cel}'].font = Font(size=10, name='Arial', bold=True)
                    sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="center")
                    sheet[f'{cell}{num_cel}'].border = thin_border
                for coll in 'IJKLMNOPQRST':
                    sheet[f'{coll}{num_cel}'] = f'={coll}7'
                    sheet[f'{coll}{num_cel}'].font = Font(size=10, name='Arial', bold=True)
                    sheet[f'{coll}{num_cel}'].alignment = Alignment(horizontal="center", vertical="center")
                    sheet[f'{coll}{num_cel}'].border = thin_border
                    sheet[f'{coll}{num_cel}'].number_format = numbers.BUILTIN_FORMATS[20]
                num_cel += 1
            # Заполняем параметры
            sheet[f'C{num_cel}'] = f'="{par}"'
            sheet[f'F{num_cel}'] = f'="{num_par}"'
            rus_name = property_par[0]
            sheet[f'G{num_cel}'] = f'="{rus_name}  "'  # Русское наименование

            # Заполняем формулы для вывода данных
            # Получение текущего значения, пока закомментировано, чтобы не вызывать дополнительные расчёты
            # sheet[f'V{num_cel}'] = f'=CurrAttrValue(E{num_cel}, 0)'
            for i in range(len(tuple_sh)):
                # sheet[f'{tuple_sh[i]}{num_cel}'] = f'=ArchiveAttributeValue($E{num_cel}, ' \
                #                                    f'0, {str_sh[i]}$5, {str_sh[i]}$6, 1)'
                sheet[f'{tuple_sh[i]}{num_cel}'] = f'=valueondate($E{num_cel}, 0, {str_sh[i]}$5, FALSE())'

                # sheet[f'{tuple_sh[i]}{num_cel}'] = f'=ValueOnDate($E{num_cel}, 0, {str_sh[i]}$5)'
            for i in range(len(str_sh)):
                # sheet[f'{str_sh[i]}{num_cel}'] = f'=IF({str_sh[i]}$4, IF(ISNUMBER({tuple_sh[i]}{num_cel}), ' \
                #                                  f'{tuple_sh[i]}{num_cel}, "нет"), "-")'  # предпоследний "-" это был $V{num_cel}
                sheet[f'{str_sh[i]}{num_cel}'] = f'=IF({str_sh[i]}$4, {tuple_sh[i]}{num_cel}, "-")'  # предпоследний "-" это был $V{num_cel}
                sheet[f'{str_sh[i]}{num_cel}'].number_format = numbers.BUILTIN_FORMATS[2]
                sheet[f'{str_sh[i]}{num_cel}'].font = Font(size=9, name='Arial')
                sheet[f'{str_sh[i]}{num_cel}'].border = thin_border
                sheet[f'{str_sh[i]}{num_cel}'].alignment = Alignment(horizontal="center", vertical="center")
            # Заполняем ссылки
            for i, value in {# 'D': f'=CONCATENATE($E$3, $E$2, $C{num_cel}, $E$5)',
                             'E': f'=CONCATENATE($E$3, $E$2, $C{num_cel}, $E$4)'}.items():
                sheet[f'{i}{num_cel}'] = value
            # Заполняем формулы
            for i, value in {'H': f'=CurrAttrValue(D{num_cel}, 0)'}.items():
                # sheet[f'{i}{num_cel}'] = value
                sheet[f'{i}{num_cel}'] = f'="{property_par[2]}"'
                # print(property_par)

            # Устанавливаем для ячеек размер шрифта 14, Arial и выставляем границу
            for i in 'FGH':
                sheet[f'{i}{num_cel}'].font = Font(size=10, name='Arial')
                sheet[f'{i}{num_cel}'].border = thin_border
            # Выводим в центр нумерацию, значения и единицы измерения
            for i, vert_alig in {'F': 'center', 'G': 'left', 'H': 'center', 'S': 'center'}.items():
                sheet[f'{i}{num_cel}'].alignment = Alignment(horizontal=vert_alig, vertical="center",
                                                             wrap_text=True)
            sheet.row_dimensions[num_cel].height = 15
            num_cel += 1
            num_par += 1
            igrek_delta -= 1

        # Прячем столбцы с привязками
        for col in ('A', 'B', 'C', 'D', 'E'):
            sheet.column_dimensions[col].hidden = True
        for row in (4, 5, 6):
            sheet.row_dimensions[row].hidden = True
        for col in ('U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK',
                    'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX'):
            sheet.column_dimensions[col].hidden = True
        # Добавляем подпись в конце
        num_cel += 2
        tuple_page += (num_page,)

        # for _ in range(3):
        #     sheet.row_dimensions[num_cel].height = 35
        #     for cell, i in {'G': '="должность"', 'J': '="ФИО"', 'M': '="подпись"'}.items():
        #         sheet[f'{cell}{num_cel}'] = i
        #         sheet[f'{cell}{num_cel}'].font = Font(size=12, name='Arial')
        #         sheet[f'{cell}{num_cel}'].alignment = Alignment(horizontal="center", vertical="top")
        #     for coll in 'GHIJKLMN':
        #         sheet[f'{      coll}{num_cel}'].border = Border(top=Side(style='thin'))
        #     num_cel += 1

        # Добавляем общие имена (локальные переменные)
        # new_range = DefinedName('smena', attr_text=1, comment='2') # attr_text='"1"'
        # wb.defined_names.add(new_range)
        new_range = DefinedName('reportdate', attr_text='"44987"', comment='0')
        wb.defined_names.add(new_range)

        # for _ in range(igrek_delta):
        #     sheet.row_dimensions[num_cel].height = 15
        #     num_cel += 1
        # cc = igrek + num_cel_start
        # for _ in tuple_page:
        #     # print(_, cc)
        #     # print(sheet[f'G{_}'].value, end=' ')
        #     # tmp = sheet[f'G{_}'].value.replace('=', '').replace('"', '')
        #     sheet[f'U{cc}'].value = f'="{_}/{len(tuple_page)}"'
        #     sheet[f'U{cc}'].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        #     cc += (igrek + 6)
        #
        # # Добавляем колонтитул
        # footer_text = ('_' * 144 + ' ' * 2 + 'должность' + ' ' * 130 + 'ФИО' + ' ' * 50 + 'подпись' + ' ' * 50) * 3
        # sheet.oddFooter.center.text = footer_text
        # sheet.oddFooter.center.size = 10
        # sheet.oddFooter.center.font = "Arial,Bold"
        # Добавляем колонтитул
        footer_text = ('_' * 144 + ' ' * 2 + 'должность' + ' ' * 130 + 'ФИО' + ' ' * 50 + 'подпись' + ' ' * 50) * 3
        sheet.oddFooter.center.text = footer_text
        sheet.oddFooter.center.size = 10
        sheet.oddFooter.center.font = "Arial,Bold"

        # Если нет папки File_for_Import/Reports, то создадим её
        if not os.path.exists(os.path.join('File_for_Import', 'Reports')):
            os.mkdir(os.path.join('File_for_Import', 'Reports'))
        # Сохраняем
        wb.save(
            os.path.join('File_for_Import', 'Reports', f'{node_param_rus} {sl_name_object.get(obj, str(777))}.xlsx')
        )

    return

